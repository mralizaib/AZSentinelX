"""
DVR/NVR Bulk Auto-Configuration — Blueprint
Accepts a CSV or Excel file with multiple device entries and runs the full
11-step configuration workflow for each device in a background thread.
"""
import io
import json
import logging
import smtplib
import threading
import time
import uuid
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from flask import (Blueprint, Response, jsonify, render_template,
                   request, send_file)
from flask_login import current_user, login_required

from config import Config
from models import DVRBulkJob, DVRConfigSession, db
from routes.permissions import make_blueprint_permission_check

logger = logging.getLogger(__name__)

dvr_bulk_bp = Blueprint('dvr_bulk', __name__)
dvr_bulk_bp.before_request(make_blueprint_permission_check('dvr_bulk'))

# ─────────────────────────────────────────────────────────────
# DB helpers
# ─────────────────────────────────────────────────────────────

def _load_job(job_id):
    return DVRBulkJob.query.get(job_id)


def _save_job(job):
    db.session.add(job)
    job.updated_at = datetime.utcnow()
    db.session.commit()


# ─────────────────────────────────────────────────────────────
# CSV / Excel parsing
# ─────────────────────────────────────────────────────────────

REQUIRED_COLS = {'ip', 'username', 'password'}
OPTIONAL_COLS = {
    'device_name': '',
    'port': '80',
    'use_https': 'false',
    'city': '',
    'state': '',
    'client_initials': '',
    'dlt_password': '',
    'cms_password': '',
}

COLUMN_ALIASES = {
    # ip
    'ip address': 'ip', 'url': 'ip', 'ip/url': 'ip', 'host': 'ip', 'address': 'ip',
    # port
    'port number': 'port',
    # username
    'user': 'username', 'admin user': 'username', 'admin username': 'username',
    # password
    'pass': 'password', 'admin password': 'password', 'admin pass': 'password',
    # device_name
    'name': 'device_name', 'device': 'device_name', 'dvr name': 'device_name',
    'nvr name': 'device_name', 'site name': 'device_name',
    # city / state
    'location city': 'city', 'location state': 'state',
    # client_initials
    'initials': 'client_initials', 'client': 'client_initials',
    # dlt_password
    'dlt pass': 'dlt_password', 'dlt pwd': 'dlt_password',
    # cms_password
    'cms pass': 'cms_password', 'cms pwd': 'cms_password',
    # use_https
    'https': 'use_https', 'ssl': 'use_https',
}


def _normalise_col(col):
    """Lower-strip and resolve aliases. Handles None gracefully."""
    if col is None:
        return '__unknown__'
    c = str(col).strip().lower()
    return COLUMN_ALIASES.get(c, c.replace(' ', '_'))


def _parse_file(file_storage):
    """Parse uploaded CSV or Excel.  Returns (devices: list[dict], error: str|None)."""
    filename = (file_storage.filename or '').lower()

    try:
        if filename.endswith('.csv'):
            import csv
            content = file_storage.read().decode('utf-8-sig', errors='replace')
            reader = csv.DictReader(io.StringIO(content))
            raw_rows = list(reader)
            headers = [h for h in (reader.fieldnames or []) if h is not None]
        elif filename.endswith(('.xlsx', '.xls')):
            try:
                import openpyxl
            except ImportError:
                return None, 'openpyxl is required to read Excel files'
            wb = openpyxl.load_workbook(io.BytesIO(file_storage.read()), data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                return None, 'Excel file is empty'
            headers = [str(h).strip() if h is not None else '' for h in rows[0]]
            raw_rows = []
            for row in rows[1:]:
                if any(v is not None and str(v).strip() != '' for v in row):
                    raw_rows.append({headers[i]: (str(row[i]).strip() if row[i] is not None else '')
                                      for i in range(min(len(headers), len(row)))})
        else:
            return None, 'Unsupported file type. Please upload a .csv or .xlsx file.'
    except Exception as e:
        return None, f'Could not parse file: {str(e)}'

    # Normalise header names — guard against None/empty keys from malformed files
    norm_map = {h: _normalise_col(h) for h in headers if h}

    devices = []
    errors = []
    for idx, raw in enumerate(raw_rows, start=2):
        # Skip None/empty keys that can appear in malformed CSVs (e.g. extra trailing commas)
        row = {norm_map.get(k, _normalise_col(k)): (v or '').strip()
               for k, v in raw.items() if k is not None and k != ''}
        # Skip blank rows
        if not any(row.values()):
            continue
        missing = [c for c in REQUIRED_COLS if not row.get(c)]
        if missing:
            errors.append(f'Row {idx}: missing required column(s): {", ".join(missing)}')
            continue
        # Apply defaults
        for col, default in OPTIONAL_COLS.items():
            if not row.get(col):
                row[col] = default
        # Derive device_name from ip if not provided
        if not row['device_name']:
            row['device_name'] = f"Device-{row['ip']}"
        # Normalise port / https
        try:
            row['port'] = int(row['port'] or 80)
        except ValueError:
            row['port'] = 80
        row['use_https'] = row.get('use_https', 'false').lower() in ('true', '1', 'yes')
        devices.append(row)

    if errors:
        return devices, '; '.join(errors)
    return devices, None


# ─────────────────────────────────────────────────────────────
# Single-device full configuration (runs in worker thread)
# ─────────────────────────────────────────────────────────────

def _configure_one_device(app, job_id, device_idx, device, global_city, global_state):
    """Run all 9 config steps for a single device. Updates job data in DB."""
    from dvr_api import DVRClient
    from routes.dvr_config import _build_report_data

    result = {
        'device_name': device['device_name'],
        'ip': device['ip'],
        'port': device['port'],
        'status': 'running',
        'error': None,
        'session_id': None,
        'steps': {},
        'overall': 'N/A',
        'started_at': datetime.utcnow().isoformat(),
        'completed_at': None,
    }

    city  = device.get('city')  or global_city  or 'Unknown'
    state = device.get('state') or global_state or 'Texas'

    def _update_job(extra=None):
        """Write current result back into the job record."""
        with app.app_context():
            job = _load_job(job_id)
            if job is None:
                return
            data = job.get_data()
            data['devices'][device_idx] = result
            if extra:
                data.update(extra)
            job.set_data(data)
            job.updated_at = datetime.utcnow()
            db.session.commit()

    try:
        client = DVRClient(
            ip=device['ip'],
            port=device['port'],
            username=device['username'],
            password=device['password'],
            use_https=device['use_https'],
        )

        # ── Step 1: Connect ───────────────────────────────────
        result['steps']['connect'] = {'status': 'running'}
        _update_job()
        ok = client.test_connection()
        if not ok:
            raise RuntimeError('Authentication failed — check IP, port, username, password')
        result['steps']['connect'] = {'status': 'done'}

        # Push the friendly name onto the device itself — previously this
        # was only stored locally for the bulk report and never sent to the
        # device, so the DVR/NVR's own name never actually changed.
        if device.get('device_name'):
            name_applied = client.configure_device_name(device['device_name'])
            if not name_applied:
                logger.warning(f"Could not set device name on {device['ip']}; continuing anyway")

        # ── Step 2: Discover ──────────────────────────────────
        result['steps']['discover'] = {'status': 'running'}
        _update_job()
        info = client.get_device_info()
        result['steps']['discover'] = {'status': 'done', 'info': info}

        # ── Step 3: Time / NTP ────────────────────────────────
        result['steps']['time'] = {'status': 'running'}
        _update_job()
        time_cfg = client.configure_time(city, state)
        time_cfg['city'] = city
        time_cfg['state'] = state
        result['steps']['time'] = {'status': 'done', 'config': time_cfg}

        # ── Step 4: Storage ───────────────────────────────────
        result['steps']['storage'] = {'status': 'running'}
        _update_job()
        hdds = client.get_storage_info()
        formatted = []
        for hdd in hdds:
            if hdd.get('status', '').lower() in ('uninitialized', 'uninit', 'error', ''):
                try:
                    client.format_hdd(hdd['id'])
                    formatted.append(hdd['id'])
                    time.sleep(2)
                except Exception as fe:
                    logger.warning(f"Bulk: format HDD {hdd['id']} on {device['ip']}: {fe}")
        result['steps']['storage'] = {'status': 'done', 'hdds': hdds, 'formatted': formatted}

        # ── Step 5: Streams ───────────────────────────────────
        result['steps']['streams'] = {'status': 'running'}
        _update_job()
        # Use real channel IDs from discovery; fall back to synthesised range.
        all_channel_ids = info.get('channel_ids') or []
        if not all_channel_ids:
            total_ch = int(info.get('total_channels', 0)) or 4
            all_channel_ids = [f"{ch}01" for ch in range(1, total_ch + 1)]
        # Filter to only online/active cameras — skip offline or disconnected channels.
        channel_ids = client.get_online_channel_ids(all_channel_ids)
        skipped_ids  = [cid for cid in all_channel_ids if cid not in channel_ids]
        stream_results = []
        for main_id in channel_ids:
            ch_label  = main_id[:-2] if len(main_id) > 2 else main_id
            main_ok   = client.configure_main_stream(main_id)
            sub_res   = client.configure_sub_stream(main_id)
            event_ok  = client.configure_event_stream(main_id)
            stream_results.append({
                'channel':       ch_label,
                'main':          'ok'    if main_ok else 'error',
                'sub':           sub_res['status'],
                'sub_resolution': sub_res['applied_resolution'],
                'sub_note':       sub_res['note'],
                'event':         'ok'    if event_ok else 'skipped',
                'online':        True,
            })
        for cid in skipped_ids:
            ch_label = cid[:-2] if len(cid) > 2 else cid
            stream_results.append({
                'channel': ch_label,
                'main':    'skipped',
                'sub':     'skipped',
                'event':   'skipped',
                'online':  False,
            })
        client.disable_channel_zero()
        result['steps']['streams'] = {
            'status': 'done',
            'channels': stream_results,
            'active_channels': len(channel_ids),
            'skipped_channels': len(skipped_ids),
        }

        # ── Step 6: Recording schedule ────────────────────────
        result['steps']['recording'] = {'status': 'running'}
        _update_job()
        rec_results = []
        for main_id in channel_ids:
            ch_label = main_id[:-2] if len(main_id) > 2 else main_id
            ok = client.configure_recording_schedule(main_id)
            rec_results.append({'channel': ch_label, 'status': 'ok' if ok else 'error', 'online': True})
        for cid in skipped_ids:
            ch_label = cid[:-2] if len(cid) > 2 else cid
            rec_results.append({'channel': ch_label, 'status': 'skipped', 'online': False})
        result['steps']['recording'] = {
            'status': 'done',
            'channels': rec_results,
            'active_channels': len(channel_ids),
            'skipped_channels': len(skipped_ids),
        }

        # ── Step 7: Holiday ───────────────────────────────────
        result['steps']['holiday'] = {'status': 'running'}
        _update_job()
        client.configure_holiday()
        result['steps']['holiday'] = {'status': 'done'}

        # ── Step 8: Users ─────────────────────────────────────
        result['steps']['users'] = {'status': 'running'}
        _update_job()
        initials = (device.get('client_initials') or 'XX').strip().upper()
        cms_pw     = device.get('cms_password')     or f"{initials}_cam12"
        dlt_pw     = device.get('dlt_password')     or f"{initials}9722IDT!"
        manager_pw = device.get('manager_password') or f"{initials}_man12"

        # create_manager: CSV column 'create_manager' or boolean field in device dict
        _cm_raw = device.get('create_manager', False)
        create_manager = _cm_raw if isinstance(_cm_raw, bool) else str(_cm_raw).lower() in ('true', 'yes', '1')

        user_accounts = [
            ('cms',  cms_pw, 'viewer',   'cms'),
            ('dlt',  dlt_pw, 'operator', 'dlt'),
        ]
        if create_manager:
            user_accounts.append(('manager', manager_pw, 'viewer', 'manager'))

        user_results = []
        for uname, upw, urole, uperm in user_accounts:
            try:
                uid = client.create_user(uname, upw, role=urole)
                client.set_user_permissions(uid, uname, uperm)
                user_results.append({'username': uname, 'password': upw,
                                     'role': urole, 'status': 'created'})
            except Exception as ue:
                user_results.append({'username': uname, 'password': upw,
                                     'role': urole, 'status': 'error', 'error': str(ue)})
        result['steps']['users'] = {
            'status': 'done', 'accounts': user_results,
            'client_initials': initials,
            'manager_created': create_manager,
        }

        # ── Step 9: Validate ──────────────────────────────────
        result['steps']['validation'] = {'status': 'running'}
        _update_job()
        expected = dict(time_cfg)
        expected['device_name'] = device.get('device_name', '')
        expected['channel_ids'] = channel_ids  # active channel IDs from stream step
        expected['users_created'] = [
            u['username'] for u in user_results
            if u.get('status') in ('created', 'created_permissions_failed')
        ]
        checks = client.validate_configuration(expected)
        passed = sum(1 for c in checks if c['status'] == 'pass')
        warned = sum(1 for c in checks if c['status'] == 'warn')
        failed = sum(1 for c in checks if c['status'] == 'fail')
        overall = 'PASS' if failed == 0 and warned == 0 else ('WARNING' if failed == 0 else 'FAIL')
        result['steps']['validation'] = {
            'status': 'done', 'checks': checks,
            'passed': passed, 'warned': warned, 'failed': failed, 'overall': overall,
            'completed_at': datetime.utcnow().isoformat(),
        }
        result['overall'] = overall

        # ── Save DVRConfigSession ─────────────────────────────
        with app.app_context():
            sess_id = f"bulk_{job_id}_{device_idx}_{int(time.time())}"
            sess_data = {
                'session_id': sess_id,
                'technician': 'bulk-job',
                'started_at': result['started_at'],
                'connection': {
                    'device_name': device['device_name'],
                    'ip': device['ip'],
                    'port': device['port'],
                    'username': device['username'],
                    'password': '***REDACTED***',  # never persist admin password
                    'use_https': device['use_https'],
                },
                'steps': {
                    'discovery': info,
                    'time': time_cfg,
                    'storage': {'hdds': hdds, 'formatted': formatted},
                    'streams': {
                        'channels': stream_results,
                        'channel_zero_disabled': True,
                        'active_channels': len(channel_ids),
                        'skipped_channels': len(skipped_ids),
                    },
                    'recording': {
                        'channels': rec_results,
                        'active_channels': len(channel_ids),
                        'skipped_channels': len(skipped_ids),
                    },
                    'holiday': {'name': 'Thanksgiving & Black Friday', 'status': 'configured'},
                    'users': {'client_initials': initials, 'accounts': user_results},
                    'validation': result['steps']['validation'],
                },
            }
            rec = DVRConfigSession(id=sess_id, bulk_job_id=job_id)
            rec.set_data(sess_data)
            db.session.add(rec)
            db.session.commit()
            result['session_id'] = sess_id

        result['status'] = 'done'
        result['completed_at'] = datetime.utcnow().isoformat()

    except Exception as e:
        logger.error(f"Bulk config device {device['ip']}: {e}", exc_info=True)
        result['status'] = 'failed'
        result['error'] = str(e)
        result['overall'] = 'FAIL'
        result['completed_at'] = datetime.utcnow().isoformat()

    _update_job()
    return result


def _run_bulk_job(app, job_id, devices, global_city, global_state, recipients, email_formats):
    """Worker thread — processes each device sequentially, then emails report."""
    with app.app_context():
        job = _load_job(job_id)
        if not job:
            return
        job.status = 'running'
        data = job.get_data()
        data['devices'] = [
            {
                'device_name': d['device_name'], 'ip': d['ip'],
                'port': d['port'], 'status': 'pending',
                'error': None, 'session_id': None, 'steps': {},
                'overall': 'N/A', 'started_at': None, 'completed_at': None,
            }
            for d in devices
        ]
        job.set_data(data)
        _save_job(job)

    succeeded = 0
    failed_count = 0

    for idx, device in enumerate(devices):
        res = _configure_one_device(app, job_id, idx, device, global_city, global_state)
        if res['status'] == 'done':
            succeeded += 1
        else:
            failed_count += 1

        with app.app_context():
            job = _load_job(job_id)
            if job:
                job.completed = idx + 1
                job.succeeded = succeeded
                job.failed_count = failed_count
                _save_job(job)

    # Mark job done
    with app.app_context():
        job = _load_job(job_id)
        if job:
            job.status = 'done'
            job.completed = len(devices)
            job.succeeded = succeeded
            job.failed_count = failed_count
            _save_job(job)

    # Send email if requested
    if recipients:
        try:
            _email_bulk_report(app, job_id, recipients, email_formats)
        except Exception as e:
            logger.error(f"Bulk email failed: {e}", exc_info=True)


# ─────────────────────────────────────────────────────────────
# Report builders
# ─────────────────────────────────────────────────────────────

def _build_bulk_excel(job_data, job):
    """Build a multi-sheet Excel: Overview + per-device sheets."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, PatternFill
    except ImportError:
        return None

    wb = openpyxl.Workbook()
    hdr_font  = Font(bold=True, color='FFFFFF')
    hdr_fill  = PatternFill('solid', fgColor='0056B3')
    pass_fill = PatternFill('solid', fgColor='C6EFCE')
    warn_fill = PatternFill('solid', fgColor='FFEB9C')
    fail_fill = PatternFill('solid', fgColor='FFC7CE')

    # ── Overview sheet ────────────────────────────────────────
    ws = wb.active
    ws.title = 'Overview'
    headers = ['#', 'Device Name', 'IP Address', 'Port', 'Overall', 'Passed',
               'Warnings', 'Failed', 'Status', 'Started', 'Completed', 'Session ID']
    ws.append(headers)
    for cell in ws[1]:
        cell.font = hdr_font
        cell.fill = hdr_fill

    devices = job_data.get('devices', [])
    for i, d in enumerate(devices, start=1):
        val = d.get('steps', {}).get('validation', {})
        overall = d.get('overall', 'N/A')
        row = [
            i, d['device_name'], d['ip'], d.get('port', 80),
            overall,
            val.get('passed', ''), val.get('warned', ''), val.get('failed', ''),
            d.get('status', ''),
            d.get('started_at', '')[:19] if d.get('started_at') else '',
            d.get('completed_at', '')[:19] if d.get('completed_at') else '',
            d.get('session_id', ''),
        ]
        ws.append(row)
        fill = pass_fill if overall == 'PASS' else (warn_fill if overall == 'WARNING' else fail_fill)
        for cell in ws[ws.max_row]:
            cell.fill = fill

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or '')), max((len(str(c.value or '')) for c in col[1:]), default=0)
        ) + 4

    # ── Per-device sheets ─────────────────────────────────────
    for d in devices:
        title = (d['device_name'] or d['ip'])[:31].replace('/', '-').replace('\\', '-')
        ws2 = wb.create_sheet(title=title)

        # Device info block
        ws2.append(['Field', 'Value'])
        for cell in ws2[1]:
            cell.font = hdr_font
            cell.fill = hdr_fill

        info = d.get('steps', {}).get('discover', {}).get('info', {})
        time_cfg = d.get('steps', {}).get('time', {}).get('config', {})
        users = d.get('steps', {}).get('users', {}).get('accounts', [])
        val = d.get('steps', {}).get('validation', {})

        fields = [
            ('Device Name',    d['device_name']),
            ('IP Address',     d['ip']),
            ('Port',           str(d.get('port', 80))),
            ('Manufacturer',   info.get('manufacturer', 'N/A')),
            ('Model',          info.get('model', 'N/A')),
            ('Serial Number',  info.get('serial_number', 'N/A')),
            ('Firmware',       info.get('firmware_version', 'N/A')),
            ('Total Channels', str(info.get('total_channels', 'N/A'))),
            ('Timezone',       time_cfg.get('iana_timezone', 'N/A')),
            ('NTP Server',     time_cfg.get('ntp_server', 'pool.ntp.org')),
            ('City/State',     f"{time_cfg.get('city','')}, {time_cfg.get('state','')}"),
            ('Overall Result', d.get('overall', 'N/A')),
            ('Status',         d.get('status', 'N/A')),
            ('Error',          d.get('error', '')),
            ('Completed At',   (d.get('completed_at', '') or '')[:19]),
        ]
        for row in fields:
            ws2.append(row)

        # Users sub-table
        ws2.append([])
        ws2.append(['Username', 'Password', 'Role', 'Status'])
        for cell in ws2[ws2.max_row]:
            cell.font = hdr_font
            cell.fill = hdr_fill
        for u in users:
            ws2.append([u.get('username', ''), u.get('password', ''),
                        u.get('role', ''), u.get('status', '')])

        # Validation sub-table (matches single DVR Excel: Status | Check | Detail | Action)
        ws2.append([])
        checks = val.get('checks', [])
        if checks:
            ws2.append(['Status', 'Validation Check', 'Detail', 'Corrective Action'])
            for cell in ws2[ws2.max_row]:
                cell.font = hdr_font
                cell.fill = hdr_fill
            ck_pass_fill = PatternFill('solid', fgColor='C6EFCE')
            ck_warn_fill = PatternFill('solid', fgColor='FFEB9C')
            ck_fail_fill = PatternFill('solid', fgColor='FFC7CE')
            for c in checks:
                icon = {'pass': '✅', 'warn': '⚠', 'fail': '❌'}.get(c['status'], '?')
                ws2.append([f"{icon} {c['status'].upper()}", c['check'], c['detail'], c.get('action', '')])
                row_fill = {'pass': ck_pass_fill, 'warn': ck_warn_fill, 'fail': ck_fail_fill}.get(c['status'])
                if row_fill:
                    for cell in ws2[ws2.max_row]:
                        cell.fill = row_fill

        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _h(s):
    """HTML-escape a string for safe embedding in report output."""
    import html
    return html.escape(str(s or ''), quote=True)


def _build_bulk_html_summary(job_data, job):
    """Build a standalone HTML report matching the single-DVR report format, one section per device."""
    devices = job_data.get('devices', [])
    total   = len(devices)
    passed  = sum(1 for d in devices if d.get('overall') == 'PASS')
    warned  = sum(1 for d in devices if d.get('overall') == 'WARNING')
    failed  = sum(1 for d in devices if d.get('overall') == 'FAIL')

    STATUS_ICON  = {'pass': '✅', 'warn': '⚠', 'fail': '❌'}
    STATUS_COLOR = {'pass': '#28a745', 'warn': '#ffc107', 'fail': '#dc3545'}
    OVERALL_COLOR = {'PASS': '#28a745', 'WARNING': '#ffc107', 'FAIL': '#dc3545'}

    # ── Summary table (index) ──────────────────────────────────
    summary_rows = ''
    for i, d in enumerate(devices, 1):
        overall = d.get('overall', 'N/A')
        color   = OVERALL_COLOR.get(overall, '#999')
        icon    = {'PASS': '✅', 'WARNING': '⚠', 'FAIL': '❌'}.get(overall, '?')
        val     = d.get('steps', {}).get('validation', {})
        err     = f"<br><small style='color:#dc3545;'>{_h(d.get('error',''))}</small>" if d.get('error') else ''
        summary_rows += f"""<tr>
          <td>{i}</td>
          <td><a href="#device-{i}"><strong>{_h(d['device_name'])}</strong></a>{err}</td>
          <td><code>{_h(d['ip'])}:{_h(d.get('port', 80))}</code></td>
          <td><span style="color:{color};font-weight:bold;">{icon} {_h(overall)}</span></td>
          <td>✅ {val.get('passed','')} &nbsp; ⚠ {val.get('warned','')} &nbsp; ❌ {val.get('failed','')}</td>
          <td>{_h(d.get('status',''))}</td>
        </tr>"""

    # ── Per-device detail sections ─────────────────────────────
    detail_sections = ''
    for i, d in enumerate(devices, 1):
        overall     = d.get('overall', 'N/A')
        oc          = OVERALL_COLOR.get(overall, '#999')
        oi          = {'PASS': '✅', 'WARNING': '⚠', 'FAIL': '❌'}.get(overall, '?')
        info        = d.get('steps', {}).get('discover', {}).get('info', {})
        time_cfg    = d.get('steps', {}).get('time',     {}).get('config', {})
        val         = d.get('steps', {}).get('validation', {})
        users       = d.get('steps', {}).get('users', {}).get('accounts', [])
        hdds        = d.get('steps', {}).get('storage', {}).get('hdds', [])
        checks      = val.get('checks', [])
        vs_passed   = val.get('passed', 0)
        vs_warned   = val.get('warned', 0)
        vs_failed   = val.get('failed', 0)

        # Device info grid
        info_grid = f"""
        <div class="info-grid">
          <div class="info-item"><label>Device Name</label><span>{_h(d.get('device_name','N/A'))}</span></div>
          <div class="info-item"><label>IP Address</label><span><code>{_h(d['ip'])}:{_h(d.get('port',80))}</code></span></div>
          <div class="info-item"><label>Manufacturer</label><span>{_h(info.get('manufacturer','N/A'))}</span></div>
          <div class="info-item"><label>Model</label><span>{_h(info.get('model','N/A'))}</span></div>
          <div class="info-item"><label>Device Type</label><span>{_h(info.get('device_type','N/A'))}</span></div>
          <div class="info-item"><label>Serial Number</label><span>{_h(info.get('serial_number','N/A'))}</span></div>
          <div class="info-item"><label>Firmware Version</label><span>{_h(info.get('firmware_version','N/A'))}</span></div>
          <div class="info-item"><label>MAC Address</label><span>{_h(info.get('mac_address','N/A'))}</span></div>
          <div class="info-item"><label>Total Channels</label><span>{_h(info.get('total_channels','N/A'))}</span></div>
          <div class="info-item"><label>Active Channels</label><span>{_h(info.get('active_channels','N/A'))}</span></div>
          <div class="info-item"><label>Location</label><span>{_h(time_cfg.get('city',''))}, {_h(time_cfg.get('state',''))}</span></div>
          <div class="info-item"><label>Timezone</label><span>{_h(time_cfg.get('iana_timezone','N/A'))}</span></div>
          <div class="info-item"><label>NTP Server</label><span>{_h(time_cfg.get('ntp_server','pool.ntp.org'))}</span></div>
          <div class="info-item"><label>Completed At</label><span>{_h((d.get('completed_at','') or '')[:19])}</span></div>
        </div>"""

        # HDD table
        hdd_rows = ''.join(
            f"<tr><td>{_h(h.get('id',''))}</td><td>{_h(h.get('name',''))}</td>"
            f"<td>{_h(h.get('capacity',''))} MB</td><td>{_h(h.get('status',''))}</td>"
            f"<td>{_h(h.get('free_space',''))} MB</td><td>{_h(h.get('property',''))}</td></tr>"
            for h in hdds
        ) or '<tr><td colspan="6" style="color:#999;">No HDD data available</td></tr>'

        # User accounts table
        user_rows = ''.join(
            f"<tr><td><strong>{_h(u.get('username',''))}</strong></td>"
            f"<td><code>{_h(u.get('password','N/A'))}</code></td>"
            f"<td>{_h(u.get('role','N/A'))}</td>"
            f"<td style=\"color:{'#28a745' if u.get('status')=='created' else '#dc3545'}\">"
            f"{_h(u.get('status','N/A'))}</td></tr>"
            for u in users
        ) or '<tr><td colspan="4" style="color:#999;">No user data available</td></tr>'

        # Validation table (identical format to single DVR report)
        check_rows = ''
        for c in checks:
            icon   = STATUS_ICON.get(c['status'], '?')
            color  = STATUS_COLOR.get(c['status'], '#999')
            action = f"<br><small style='color:#aaa;'>Action: {_h(c.get('action',''))}</small>" if c.get('action') else ''
            check_rows += f"""<tr>
              <td style="color:{color};font-size:1.2em;">{icon}</td>
              <td><strong>{_h(c['check'])}</strong></td>
              <td>{_h(c['detail'])}{action}</td>
            </tr>"""
        if not check_rows:
            check_rows = '<tr><td colspan="3" style="color:#999;">No validation data available</td></tr>'

        detail_sections += f"""
        <div id="device-{i}" class="device-section">
          <h2>Device {i} — {_h(d.get('device_name',''))} <code style="font-size:.6em;">{_h(d['ip'])}</code></h2>
          <div style="text-align:center;margin:16px 0;">
            <span class="badge" style="background:{oc};font-size:1.2em;">{oi} Overall: {_h(overall)}</span>
            &nbsp;
            <span class="badge" style="background:#28a745;">✅ {vs_passed} Passed</span>
            &nbsp;
            <span class="badge" style="background:#ffc107;">⚠ {vs_warned} Warnings</span>
            &nbsp;
            <span class="badge" style="background:#dc3545;">❌ {vs_failed} Failed</span>
          </div>
          {info_grid}

          <h3 style="color:#0056b3;margin-top:20px;">HDD Information</h3>
          <table><tr><th>ID</th><th>Name</th><th>Capacity</th><th>Status</th><th>Free Space</th><th>Property</th></tr>
          {hdd_rows}</table>

          <h3 style="color:#0056b3;margin-top:20px;">User Accounts</h3>
          <table><tr><th>Username</th><th>Password</th><th>Role</th><th>Status</th></tr>
          {user_rows}</table>

          <h3 style="color:#0056b3;margin-top:20px;">Validation Results</h3>
          <table><tr><th width="40">Status</th><th width="220">Check</th><th>Detail / Action</th></tr>
          {check_rows}</table>
        </div>
        <hr style="margin:30px 0;border-color:#ccc;">"""

    ts = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
body{{font-family:Arial,sans-serif;margin:30px;color:#222;}}
h1{{color:#0056b3;}}
h2{{color:#0056b3;border-bottom:2px solid #0056b3;padding-bottom:4px;margin-top:30px;}}
h3{{color:#0056b3;}}
table{{width:100%;border-collapse:collapse;margin-bottom:16px;}}
th{{background:#0056b3;color:#fff;padding:8px;text-align:left;}}
td{{padding:7px;border-bottom:1px solid #ddd;vertical-align:top;}}
tr:hover{{background:#f5f5f5;}}
.badge{{display:inline-block;padding:4px 12px;border-radius:4px;font-weight:bold;color:#fff;}}
.stats{{display:flex;gap:20px;margin:20px 0;flex-wrap:wrap;}}
.stat{{background:#f9f9f9;border:1px solid #ddd;border-radius:8px;padding:15px 25px;text-align:center;}}
.stat .n{{font-size:2em;font-weight:bold;}}
.info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:16px;}}
.info-item{{background:#f9f9f9;border:1px solid #ddd;border-radius:6px;padding:10px;}}
.info-item label{{font-size:0.8em;color:#666;display:block;}}
.info-item span{{font-size:1em;font-weight:bold;}}
code{{background:#eee;padding:2px 5px;border-radius:3px;font-size:0.95em;}}
.device-section{{margin-bottom:10px;}}
a{{color:#0056b3;}}
@media print{{.device-section{{page-break-inside:avoid;}}}}
</style></head><body>
<h1>🎥 ByteIT SentinalX — DVR/NVR Bulk AutoDeploy Report</h1>
<p>Generated: <strong>{ts} UTC</strong> &nbsp;|&nbsp; Job ID: <code>{job.id}</code></p>

<div class="stats">
  <div class="stat"><div class="n">{total}</div><div>Total Devices</div></div>
  <div class="stat" style="border-color:#28a745"><div class="n" style="color:#28a745">{passed}</div><div>✅ PASS</div></div>
  <div class="stat" style="border-color:#ffc107"><div class="n" style="color:#ffc107">{warned}</div><div>⚠ WARNING</div></div>
  <div class="stat" style="border-color:#dc3545"><div class="n" style="color:#dc3545">{failed}</div><div>❌ FAIL</div></div>
</div>

<h2>Summary</h2>
<table>
  <tr><th>#</th><th>Device Name</th><th>IP : Port</th><th>Overall</th><th>Validation</th><th>Status</th></tr>
  {summary_rows}
</table>

<h2>Device Detail Reports</h2>
{detail_sections}

<hr style="margin-top:40px;">
<p style="color:#999;font-size:0.85em;text-align:center;">
  ByteIT SentinalX — DVR/NVR Bulk AutoDeploy &nbsp;|&nbsp; {ts} UTC
</p>
</body></html>"""


def _email_bulk_report(app, job_id, recipients, formats):
    """Build and send the bulk report email."""
    with app.app_context():
        job = _load_job(job_id)
        if not job:
            return
        job_data = job.get_data()

    ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    total = job.total
    passed = job.succeeded
    failed_c = job.failed_count

    msg = MIMEMultipart('mixed')
    msg['Subject'] = (f"Bulk DVR/NVR Config Report — "
                      f"{total} devices | ✅ {passed} PASS | ❌ {failed_c} FAIL [{ts}]")
    msg['From'] = Config.SMTP_USERNAME
    msg['To'] = ', '.join(recipients)

    body = f"""<html><body style="font-family:Arial,sans-serif;">
<h2 style="color:#0056b3;">Bulk DVR/NVR Configuration Complete</h2>
<p><strong>Total Devices:</strong> {total}</p>
<p><strong>Passed:</strong> <span style="color:green">✅ {passed}</span></p>
<p><strong>Failed:</strong> <span style="color:red">❌ {failed_c}</span></p>
<p><strong>Job ID:</strong> <code>{job_id}</code></p>
<p><strong>Completed:</strong> {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')} UTC</p>
<hr><p style="color:#999;font-size:0.85em;">ByteIT SentinalX — AI DVR/NVR Bulk Auto-Configuration</p>
</body></html>"""
    msg.attach(MIMEText(body, 'html'))

    # Summary Excel
    if 'excel' in formats:
        try:
            buf = _build_bulk_excel(job_data, job)
            if buf:
                fname = f"Bulk_DVR_Config_{ts}.xlsx"
                part = MIMEApplication(buf.read(), Name=fname)
                part['Content-Disposition'] = f'attachment; filename="{fname}"'
                msg.attach(part)
        except Exception as e:
            logger.warning(f"Bulk Excel attachment: {e}")

    # Summary PDF
    if 'pdf' in formats:
        try:
            html = _build_bulk_html_summary(job_data, job)
            import weasyprint
            pdf_buf = io.BytesIO()
            weasyprint.HTML(string=html).write_pdf(pdf_buf)
            fname = f"Bulk_DVR_Config_{ts}.pdf"
            part = MIMEApplication(pdf_buf.getvalue(), Name=fname)
            part['Content-Disposition'] = f'attachment; filename="{fname}"'
            msg.attach(part)
        except Exception as e:
            logger.warning(f"Bulk PDF attachment: {e}")

    with app.app_context():
        pass  # just ensure context for import

    import smtplib
    with smtplib.SMTP(Config.SMTP_SERVER, Config.SMTP_PORT) as smtp:
        if Config.SMTP_USE_TLS:
            smtp.starttls()
        smtp.login(Config.SMTP_USERNAME, Config.SMTP_PASSWORD)
        smtp.sendmail(Config.SMTP_USERNAME, recipients, msg.as_string())
    logger.info(f"Bulk report sent to {recipients}")


# ─────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────

@dvr_bulk_bp.route('/dvr-bulk')
@login_required
def index():
    return render_template('dvr_bulk.html')


@dvr_bulk_bp.route('/api/dvr/bulk/template/<fmt>')
@login_required
def download_template(fmt):
    """Download a blank CSV or Excel template for bulk upload."""
    cols = ['device_name', 'ip', 'port', 'username', 'password',
            'city', 'state', 'client_initials', 'dlt_password', 'cms_password', 'use_https']
    sample = ['Main Office DVR', '192.168.1.100', '80', 'admin', 'admin123',
              'Houston', 'Texas', 'MOC', '', '', 'false']

    if fmt == 'csv':
        import csv
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(cols)
        w.writerow(sample)
        w.writerow(['Branch DVR', '192.168.1.101', '80', 'admin', 'admin123',
                    'Houston', 'Texas', 'MOC', '', '', 'false'])
        out = io.BytesIO(buf.getvalue().encode())
        out.seek(0)
        return send_file(out, as_attachment=True,
                         download_name='dvr_bulk_template.csv',
                         mimetype='text/csv')

    elif fmt == 'excel':
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill
        except ImportError:
            return jsonify({'error': 'openpyxl not available'}), 500

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Devices'
        ws.append(cols)
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='0056B3')
        ws.append(sample)
        ws.append(['Branch DVR', '192.168.1.101', '80', 'admin', 'admin123',
                   'Houston', 'Texas', 'MOC', '', '', 'false'])
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = 22

        # Instructions sheet
        ws2 = wb.create_sheet('Instructions')
        ws2.append(['Column', 'Required?', 'Description'])
        for cell in ws2[1]:
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill('solid', fgColor='0056B3')
        instructions = [
            ('device_name', 'Optional', 'Human-readable name for this DVR/NVR. Defaults to Device-<ip>.'),
            ('ip',          'Required', 'IP address or hostname of the device.'),
            ('port',        'Optional', 'HTTP port (default: 80). Use 443 with use_https=true.'),
            ('username',    'Required', 'Admin username (e.g. admin).'),
            ('password',    'Required', 'Admin password.'),
            ('city',        'Optional', 'City for timezone lookup. Falls back to global setting.'),
            ('state',       'Optional', 'US State for timezone lookup. Falls back to global setting.'),
            ('client_initials', 'Optional', 'Client initials used to generate dlt/cms passwords (e.g. MOC).'),
            ('dlt_password',    'Optional', 'Override auto-generated dlt password. Leave blank to auto-generate.'),
            ('cms_password',    'Optional', 'Override auto-generated cms password. Leave blank to auto-generate.'),
            ('use_https',   'Optional', 'true or false. Default: false.'),
        ]
        for row in instructions:
            ws2.append(row)
        for col in ws2.columns:
            ws2.column_dimensions[col[0].column_letter].width = 30

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name='dvr_bulk_template.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    return jsonify({'error': 'Unknown format'}), 400


@dvr_bulk_bp.route('/api/dvr/bulk/upload', methods=['POST'])
@login_required
def upload_devices():
    """Parse uploaded file and return device list preview."""
    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename:
        return jsonify({'success': False, 'error': 'Empty filename'}), 400

    devices, parse_error = _parse_file(f)
    if devices is None:
        return jsonify({'success': False, 'error': parse_error}), 400

    # Sanitise: strip passwords from preview
    preview = []
    for d in devices:
        preview.append({
            'device_name': d['device_name'],
            'ip': d['ip'],
            'port': d['port'],
            'username': d['username'],
            'has_password': bool(d.get('password')),
            'city': d.get('city', ''),
            'state': d.get('state', ''),
            'client_initials': d.get('client_initials', ''),
            'has_dlt_password': bool(d.get('dlt_password')),
            'has_cms_password': bool(d.get('cms_password')),
            'use_https': d.get('use_https', False),
        })

    return jsonify({
        'success': True,
        'count': len(devices),
        'devices': preview,
        'warnings': [parse_error] if parse_error else [],
    })


@dvr_bulk_bp.route('/api/dvr/bulk/start', methods=['POST'])
@login_required
def start_bulk():
    """
    Expects multipart/form-data with:
      - file: CSV or Excel
      - global_city, global_state: fallback location
      - recipients: comma-separated emails
      - email_formats: pdf,excel
    """
    from flask import current_app
    app = current_app._get_current_object()

    if 'file' not in request.files:
        return jsonify({'success': False, 'error': 'No file uploaded'}), 400
    f = request.files['file']

    global_city  = request.form.get('global_city', '').strip()
    global_state = request.form.get('global_state', '').strip()
    recipients_raw = request.form.get('recipients', '').strip()
    recipients = [r.strip() for r in recipients_raw.replace(';', ',').split(',') if r.strip()]
    email_formats_raw = request.form.get('email_formats', 'excel,pdf')
    email_formats = [x.strip() for x in email_formats_raw.split(',') if x.strip()]

    devices, parse_error = _parse_file(f)
    if devices is None:
        return jsonify({'success': False, 'error': parse_error}), 400
    if not devices:
        return jsonify({'success': False, 'error': 'No valid device rows found in file'}), 400

    job_id = str(uuid.uuid4())
    job = DVRBulkJob(
        id=job_id,
        created_by=current_user.username,
        status='pending',
        total=len(devices),
        completed=0,
        succeeded=0,
        failed_count=0,
    )
    job.set_data({'devices': [], 'global_city': global_city, 'global_state': global_state})
    db.session.add(job)
    db.session.commit()

    t = threading.Thread(
        target=_run_bulk_job,
        args=(app, job_id, devices, global_city, global_state, recipients, email_formats),
        daemon=True,
    )
    t.start()

    return jsonify({'success': True, 'job_id': job_id, 'total': len(devices)})


@dvr_bulk_bp.route('/api/dvr/bulk/progress/<job_id>', methods=['GET'])
@login_required
def bulk_progress(job_id):
    """Return current job progress (polled by the UI every 2 s)."""
    job = _load_job(job_id)
    if not job:
        return jsonify({'success': False, 'error': 'Job not found'}), 404

    data = job.get_data()
    devices = data.get('devices', [])

    # Build lightweight summary per device
    device_summary = []
    for d in devices:
        val = d.get('steps', {}).get('validation', {})
        device_summary.append({
            'device_name': d.get('device_name', ''),
            'ip': d.get('ip', ''),
            'status': d.get('status', 'pending'),
            'overall': d.get('overall', 'N/A'),
            'error': d.get('error'),
            'steps_done': [k for k, v in d.get('steps', {}).items()
                           if isinstance(v, dict) and v.get('status') == 'done'],
            'steps_running': [k for k, v in d.get('steps', {}).items()
                               if isinstance(v, dict) and v.get('status') == 'running'],
            'passed': val.get('passed', ''),
            'warned': val.get('warned', ''),
            'failed_checks': val.get('failed', ''),
        })

    pct = int(job.completed / max(job.total, 1) * 100)

    return jsonify({
        'success': True,
        'job_id': job_id,
        'status': job.status,
        'total': job.total,
        'completed': job.completed,
        'succeeded': job.succeeded,
        'failed_count': job.failed_count,
        'percent': pct,
        'devices': device_summary,
    })


@dvr_bulk_bp.route('/api/dvr/bulk/report/<job_id>/<fmt>', methods=['GET'])
@login_required
def bulk_report(job_id, fmt):
    """Download bulk report: excel or pdf."""
    job = _load_job(job_id)
    if not job:
        return jsonify({'success': False, 'error': 'Job not found'}), 404

    job_data = job.get_data()
    ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    filename = f"Bulk_DVR_Config_{ts}"

    if fmt == 'excel':
        buf = _build_bulk_excel(job_data, job)
        if not buf:
            return jsonify({'success': False, 'error': 'openpyxl not available'}), 500
        return send_file(buf, as_attachment=True,
                         download_name=f'{filename}.xlsx',
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif fmt == 'pdf':
        html = _build_bulk_html_summary(job_data, job)
        try:
            import weasyprint
            buf = io.BytesIO()
            weasyprint.HTML(string=html).write_pdf(buf)
            buf.seek(0)
            return send_file(buf, as_attachment=True,
                             download_name=f'{filename}.pdf',
                             mimetype='application/pdf')
        except Exception as e:
            logger.warning(f"Bulk PDF failed: {e}")
            return Response(html, mimetype='text/html',
                            headers={'Content-Disposition':
                                     f'attachment; filename="{filename}.html"'})

    elif fmt == 'json':
        job_data['job_id'] = job_id
        job_data['status'] = job.status
        buf = io.BytesIO(json.dumps(job_data, indent=2).encode())
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f'{filename}.json',
                         mimetype='application/json')

    return jsonify({'success': False, 'error': f'Unknown format: {fmt}'}), 400


@dvr_bulk_bp.route('/api/dvr/bulk/email/<job_id>', methods=['POST'])
@login_required
def email_bulk(job_id):
    """Email bulk report on demand."""
    data = request.get_json() or {}
    recipients = data.get('recipients', [])
    formats = data.get('formats', ['excel', 'pdf'])

    if not recipients:
        return jsonify({'success': False, 'error': 'No recipients specified'}), 400

    job = _load_job(job_id)
    if not job:
        return jsonify({'success': False, 'error': 'Job not found'}), 404

    try:
        from flask import current_app
        _email_bulk_report(current_app._get_current_object(), job_id, recipients, formats)
        return jsonify({'success': True, 'message': f'Report sent to {", ".join(recipients)}'})
    except Exception as e:
        logger.error(f"Email bulk report: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@dvr_bulk_bp.route('/api/dvr/bulk/jobs', methods=['GET'])
@login_required
def list_jobs():
    """List recent bulk jobs."""
    jobs = DVRBulkJob.query.order_by(DVRBulkJob.created_at.desc()).limit(20).all()
    result = []
    for j in jobs:
        result.append({
            'id': j.id,
            'created_by': j.created_by,
            'created_at': j.created_at.isoformat() if j.created_at else '',
            'status': j.status,
            'total': j.total,
            'completed': j.completed,
            'succeeded': j.succeeded,
            'failed_count': j.failed_count,
        })
    return jsonify({'jobs': result})
