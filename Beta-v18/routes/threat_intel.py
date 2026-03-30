from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for, send_file, Response
from flask_login import login_required, current_user
import logging
import json
import io
import csv
from datetime import datetime

logger = logging.getLogger(__name__)

threat_intel_bp = Blueprint('threat_intel', __name__)


@threat_intel_bp.route('/threat-intel')
@login_required
def index():
    from models import ThreatIntelItem, ThreatIntelConfig, db
    from threat_intel_service import FEED_SOURCES
    cfg = ThreatIntelConfig.get_instance()
    return render_template('threat_intel.html', cfg=cfg, feed_sources=FEED_SOURCES)


@threat_intel_bp.route('/api/threat-intel/items')
@login_required
def list_items():
    from models import ThreatIntelItem, db
    from threat_intel_service import FEED_SOURCES

    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 25, type=int)
    severity = request.args.get('severity', '')
    source = request.args.get('source', '')
    has_patch = request.args.get('has_patch', '')
    search = request.args.get('search', '').strip()
    exposed = request.args.get('exposed', '')

    query = ThreatIntelItem.query.order_by(ThreatIntelItem.published_at.desc())

    if severity:
        query = query.filter(ThreatIntelItem.severity == severity)
    if source:
        query = query.filter(ThreatIntelItem.source == source)
    if has_patch == 'true':
        query = query.filter(ThreatIntelItem.has_patch == True)
    if search:
        query = query.filter(ThreatIntelItem.title.ilike(f'%{search}%'))
    if exposed == 'true':
        from models import ThreatIntelCorrelation
        # Only list items where the vulnerability was *confirmed* present in the
        # environment (CVE found in Wazuh alert logs, or directly detected by the
        # Wazuh Vulnerability Detector).  Pure keyword/OS matches on external
        # feeds are excluded to prevent false positives in the exposed inventory.
        exposed_ids = db.session.query(ThreatIntelCorrelation.threat_intel_item_id).filter(
            ThreatIntelCorrelation.is_confirmed_present == True
        ).subquery()
        query = query.filter(ThreatIntelItem.id.in_(exposed_ids))

    total = query.count()
    items = query.offset((page - 1) * per_page).limit(per_page).all()

    results = []
    for item in items:
        analysis = {}
        if item.ai_analysis:
            try:
                analysis = json.loads(item.ai_analysis)
            except Exception:
                pass

        corr = item.correlation
        results.append({
            'id': item.id,
            'title': item.title,
            'description': item.description[:300] if item.description else '',
            'url': item.url,
            'source': item.source,
            'source_name': FEED_SOURCES.get(item.source, {}).get('name', item.source),
            'source_color': FEED_SOURCES.get(item.source, {}).get('color', 'secondary'),
            'published_at': item.published_at.strftime('%Y-%m-%d %H:%M') if item.published_at else '',
            'fetched_at': item.fetched_at.strftime('%Y-%m-%d %H:%M') if item.fetched_at else '',
            'severity': item.severity,
            'has_patch': item.has_patch,
            'has_mitigation': item.has_mitigation,
            'relevance_score': item.relevance_score,
            'ai_analyzed': item.ai_analyzed,
            'email_sent': item.email_sent,
            'cve_ids': item.get_cve_list(),
            'summary': analysis.get('summary', ''),
            'recommended_action': analysis.get('recommended_action', ''),
            'affected_count': corr.affected_count if corr else 0,
            'is_confirmed': corr.is_confirmed_present if corr else False,
        })

    return jsonify({
        'items': results,
        'total': total,
        'page': page,
        'per_page': per_page,
        'pages': (total + per_page - 1) // per_page,
    })


@threat_intel_bp.route('/api/threat-intel/stats')
@login_required
def stats():
    from models import ThreatIntelItem, db
    from sqlalchemy import func

    from models import ThreatIntelCorrelation
    total = ThreatIntelItem.query.count()
    by_severity = db.session.query(
        ThreatIntelItem.severity, func.count(ThreatIntelItem.id)
    ).group_by(ThreatIntelItem.severity).all()
    by_source = db.session.query(
        ThreatIntelItem.source, func.count(ThreatIntelItem.id)
    ).group_by(ThreatIntelItem.source).all()
    with_patch = ThreatIntelItem.query.filter_by(has_patch=True).count()
    emails_sent = ThreatIntelItem.query.filter_by(email_sent=True).count()
    analysed = ThreatIntelItem.query.filter_by(ai_analyzed=True).count()

    infra_exposed = db.session.query(func.count(ThreatIntelCorrelation.id)).filter(
        ThreatIntelCorrelation.is_confirmed_present == True
    ).scalar() or 0
    confirmed_in_env = db.session.query(func.count(ThreatIntelCorrelation.id)).filter(
        ThreatIntelCorrelation.is_confirmed_present == True
    ).scalar() or 0

    return jsonify({
        'total': total,
        'by_severity': dict(by_severity),
        'by_source': dict(by_source),
        'with_patch': with_patch,
        'emails_sent': emails_sent,
        'analysed': analysed,
        'infra_exposed': infra_exposed,
        'confirmed_in_env': confirmed_in_env,
    })


@threat_intel_bp.route('/api/threat-intel/refresh', methods=['POST'])
@login_required
def manual_refresh():
    from flask import current_app
    try:
        from threat_intel_service import run_full_refresh
        import threading
        app = current_app._get_current_object()

        def run():
            run_full_refresh(app)

        thread = threading.Thread(target=run, daemon=True)
        thread.start()
        return jsonify({'success': True, 'message': 'Refresh started in background. Check back shortly.'})
    except Exception as e:
        logger.error(f"Error triggering refresh: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@threat_intel_bp.route('/api/threat-intel/configure', methods=['POST'])
@login_required
def configure():
    from models import ThreatIntelConfig, db
    from threat_intel_service import FEED_SOURCES
    try:
        data = request.json or {}
        cfg = ThreatIntelConfig.get_instance()
        cfg.email_recipient = data.get('email_recipient', cfg.email_recipient)
        cfg.enabled = bool(data.get('enabled', cfg.enabled))
        cfg.notify_on_patch = bool(data.get('notify_on_patch', cfg.notify_on_patch))
        cfg.notify_on_critical = bool(data.get('notify_on_critical', cfg.notify_on_critical))
        min_rel = data.get('min_relevance', cfg.min_relevance)
        try:
            cfg.min_relevance = max(1, min(10, int(min_rel)))
        except (ValueError, TypeError):
            pass
        raw_sources = data.get('sources', cfg.get_sources())
        valid_sources = [s for s in raw_sources if s in FEED_SOURCES]
        cfg.set_sources(valid_sources if valid_sources else list(FEED_SOURCES.keys()))
        cfg.updated_at = datetime.utcnow()
        db.session.commit()
        return jsonify({'success': True})
    except Exception as e:
        logger.error(f"Error saving threat intel config: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@threat_intel_bp.route('/api/threat-intel/item/<int:item_id>')
@login_required
def get_item(item_id):
    from models import ThreatIntelItem
    from threat_intel_service import FEED_SOURCES
    item = ThreatIntelItem.query.get_or_404(item_id)
    analysis = {}
    if item.ai_analysis:
        try:
            analysis = json.loads(item.ai_analysis)
        except Exception:
            pass

    correlation_data = None
    corr = item.correlation
    if corr:
        correlation_data = {
            'correlated_at': corr.correlated_at.strftime('%Y-%m-%d %H:%M') if corr.correlated_at else '',
            'affected_count': corr.affected_count,
            'env_relevance_score': corr.env_relevance_score,
            'env_recommended_action': corr.env_recommended_action or '',
            'correlation_summary': corr.correlation_summary or '',
            'is_confirmed_present': corr.is_confirmed_present,
            'affected_agents': corr.get_affected_agents(),
        }

    return jsonify({
        'id': item.id,
        'title': item.title,
        'description': item.description or '',
        'url': item.url,
        'source': item.source,
        'source_name': FEED_SOURCES.get(item.source, {}).get('name', item.source),
        'source_color': FEED_SOURCES.get(item.source, {}).get('color', 'secondary'),
        'published_at': item.published_at.strftime('%Y-%m-%d %H:%M') if item.published_at else '',
        'fetched_at': item.fetched_at.strftime('%Y-%m-%d %H:%M') if item.fetched_at else '',
        'severity': item.severity,
        'has_patch': item.has_patch,
        'has_mitigation': item.has_mitigation,
        'relevance_score': item.relevance_score,
        'ai_analyzed': item.ai_analyzed,
        'email_sent': item.email_sent,
        'cve_ids': item.get_cve_list(),
        'summary': analysis.get('summary', ''),
        'recommended_action': analysis.get('recommended_action', ''),
        'correlation': correlation_data,
    })


@threat_intel_bp.route('/api/threat-intel/item/<int:item_id>/analyse', methods=['POST'])
@login_required
def analyse_single(item_id):
    from models import ThreatIntelItem, db
    from threat_intel_service import analyse_item_with_ai
    try:
        item = ThreatIntelItem.query.get_or_404(item_id)
        result = analyse_item_with_ai(item)
        item.ai_analysis = json.dumps(result)
        item.ai_analyzed = True
        item.severity = result.get('severity', item.severity)
        item.relevance_score = result.get('relevance_score', 5)
        item.has_patch = result.get('has_patch', item.has_patch)
        item.has_mitigation = result.get('has_mitigation', item.has_mitigation)
        db.session.commit()
        return jsonify({'success': True, 'analysis': result})
    except Exception as e:
        logger.error(f"Error analysing item {item_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@threat_intel_bp.route('/api/threat-intel/item/<int:item_id>/correlate', methods=['POST'])
@login_required
def correlate_single(item_id):
    """Correlate a single threat intel item against the internal infrastructure on demand."""
    from flask import current_app
    try:
        from threat_intel_correlator import correlate_single_item
        app = current_app._get_current_object()
        result = correlate_single_item(app, item_id)
        if 'error' in result:
            err = result['error']
            code = 404 if 'not found' in err.lower() else 503
            return jsonify({'success': False, 'error': err}), code
        return jsonify({'success': True, 'correlation': result})
    except Exception as e:
        logger.error(f"Error correlating item {item_id}: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@threat_intel_bp.route('/api/threat-intel/item/<int:item_id>/followup', methods=['POST'])
@login_required
def followup_question(item_id):
    """Answer a follow-up question about a specific threat intel item using AI."""
    from models import ThreatIntelItem
    try:
        item = ThreatIntelItem.query.get_or_404(item_id)
        data = request.json or {}
        question = (data.get('question') or '').strip()
        if not question:
            return jsonify({'success': False, 'error': 'No question provided'}), 400

        analysis = {}
        if item.ai_analysis:
            try:
                analysis = json.loads(item.ai_analysis)
            except Exception:
                pass

        from ai_insights import AIInsights
        ai = AIInsights()
        context = (
            f"Threat Intel Item:\n"
            f"Title: {item.title}\n"
            f"Severity: {item.severity}\n"
            f"CVEs: {', '.join(item.get_cve_list()) or 'None'}\n"
            f"Description: {(item.description or '')[:600]}\n"
            f"AI Summary: {analysis.get('summary', 'Not yet analysed')}\n"
            f"Recommended Action: {analysis.get('recommended_action', '')}\n\n"
            f"Answer the following question from a security analyst perspective, "
            f"focusing specifically on this threat. Be concise and actionable:\n\n"
            f"Question: {question}"
        )
        result = ai.analyze_alerts(alerts_data=[], analysis_prompt=context)
        answer = result.get('analysis', '').strip()
        if not answer:
            return jsonify({'success': False, 'error': 'AI did not return an answer. Ensure an AI provider is configured in Settings.'}), 500
        return jsonify({'success': True, 'answer': answer})
    except Exception as e:
        logger.error(f"Follow-up question failed for item {item_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@threat_intel_bp.route('/api/threat-intel/correlate-all', methods=['POST'])
@login_required
def correlate_all():
    """Run infrastructure correlation on all unprocessed items in background."""
    from flask import current_app
    import threading
    try:
        from threat_intel_correlator import correlate_items
        app = current_app._get_current_object()

        def run():
            correlate_items(app, max_items=50)

        threading.Thread(target=run, daemon=True).start()
        return jsonify({'success': True, 'message': 'Correlation started in background. Refresh items shortly.'})
    except Exception as e:
        logger.error(f"Error starting correlate-all: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


def _get_item_report_data(item_id):
    """Shared helper: fetch item + correlation data for report generation."""
    from models import ThreatIntelItem
    from threat_intel_service import FEED_SOURCES
    item = ThreatIntelItem.query.get_or_404(item_id)
    analysis = {}
    if item.ai_analysis:
        try:
            analysis = json.loads(item.ai_analysis)
        except Exception:
            pass
    corr = item.correlation
    agents = corr.get_affected_agents() if corr else []
    active_agents = [a for a in agents if a.get('match_source') == 'wazuh_vuln_detector' and a.get('vuln_status') == 'Active']
    solved_agents = [a for a in agents if a.get('match_source') == 'wazuh_vuln_detector' and a.get('vuln_status') == 'Solved']
    other_agents  = [a for a in agents if a.get('match_source') != 'wazuh_vuln_detector']
    return {
        'item': item,
        'analysis': analysis,
        'corr': corr,
        'agents': agents,
        'active_agents': active_agents,
        'solved_agents': solved_agents,
        'other_agents': other_agents,
        'source_name': FEED_SOURCES.get(item.source, {}).get('name', item.source),
    }


@threat_intel_bp.route('/api/threat-intel/item/<int:item_id>/download/csv')
@login_required
def download_csv(item_id):
    """Export vulnerability agent data as CSV."""
    try:
        d = _get_item_report_data(item_id)
        item = d['item']
        output = io.StringIO()
        writer = csv.writer(output)

        writer.writerow(['ByteIT SentinelX — Vulnerability Report'])
        writer.writerow(['Generated', datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')])
        writer.writerow(['CVE', ', '.join(item.get_cve_list()) or 'N/A'])
        writer.writerow(['Title', item.title])
        writer.writerow(['Severity', item.severity.upper()])
        writer.writerow(['Source', d['source_name']])
        writer.writerow(['Published', item.published_at.strftime('%Y-%m-%d') if item.published_at else ''])
        writer.writerow(['Patch Available', 'Yes' if item.has_patch else 'No'])
        writer.writerow([])

        if d['analysis']:
            writer.writerow(['AI Summary'])
            writer.writerow([d['analysis'].get('summary', '')])
            writer.writerow(['Recommended Action'])
            writer.writerow([d['analysis'].get('recommended_action', '')])
            writer.writerow([])

        if d['active_agents']:
            writer.writerow(['UNPATCHED AGENTS — Immediate Action Required'])
            writer.writerow(['Agent Name', 'IP Address', 'Location', 'OS', 'Status'])
            for a in d['active_agents']:
                writer.writerow([a.get('name',''), a.get('ip',''), a.get('location',''), a.get('os',''), 'UNPATCHED'])
            writer.writerow([])

        if d['solved_agents']:
            writer.writerow(['PATCHED AGENTS — Remediated'])
            writer.writerow(['Agent Name', 'IP Address', 'Location', 'OS', 'Status'])
            for a in d['solved_agents']:
                writer.writerow([a.get('name',''), a.get('ip',''), a.get('location',''), a.get('os',''), 'PATCHED'])
            writer.writerow([])

        if d['other_agents']:
            writer.writerow(['PLATFORM MATCH AGENTS — Investigate'])
            writer.writerow(['Agent Name', 'IP Address', 'Location', 'OS', 'Match Reason'])
            for a in d['other_agents']:
                reasons = '; '.join(a.get('reasons', []))
                writer.writerow([a.get('name',''), a.get('ip',''), a.get('location',''), a.get('os',''), reasons])

        cve_str = '_'.join(item.get_cve_list()[:1]) or f'item{item_id}'
        filename = f"vulnerability_report_{cve_str}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.error(f"CSV download error for item {item_id}: {e}")
        return jsonify({'error': str(e)}), 500


@threat_intel_bp.route('/api/threat-intel/item/<int:item_id>/download/xlsx')
@login_required
def download_xlsx(item_id):
    """Export vulnerability agent data as Excel XLSX."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        d = _get_item_report_data(item_id)
        item = d['item']
        wb = openpyxl.Workbook()

        # ── Summary Sheet ──────────────────────────────────────────
        ws = wb.active
        ws.title = 'Summary'

        hdr_fill   = PatternFill('solid', fgColor='0F2744')
        red_fill   = PatternFill('solid', fgColor='4A1010')
        green_fill = PatternFill('solid', fgColor='0A2A15')
        grey_fill  = PatternFill('solid', fgColor='1A1A2A')
        white_font = Font(color='FFFFFF', bold=True)
        red_font   = Font(color='FF6B6B', bold=True)
        green_font = Font(color='4ADE80', bold=True)
        label_font = Font(color='94A3B8', bold=False)
        val_font   = Font(color='E2E8F0')

        def set_cell(ws, row, col, value, font=None, fill=None, align=None, bold=False):
            cell = ws.cell(row=row, column=col, value=value)
            if font:  cell.font  = font
            if fill:  cell.fill  = fill
            if align: cell.alignment = align
            if bold and not font: cell.font = Font(bold=True, color='E2E8F0')
            return cell

        title_font = Font(color='FFFFFF', bold=True, size=14)
        ws.merge_cells('A1:F1')
        set_cell(ws, 1, 1, 'ByteIT SentinelX — Vulnerability Report', font=title_font, fill=hdr_fill, align=Alignment(horizontal='center'))
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:F2')
        set_cell(ws, 2, 1, f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}", font=Font(color='94A3B8', size=9), fill=PatternFill('solid', fgColor='0A0E1A'))

        meta_rows = [
            ('CVE(s)',             ', '.join(item.get_cve_list()) or 'N/A'),
            ('Title',             item.title),
            ('Severity',          item.severity.upper()),
            ('Source',            d['source_name']),
            ('Published',         item.published_at.strftime('%Y-%m-%d') if item.published_at else ''),
            ('Patch Available',   'Yes' if item.has_patch else 'No'),
            ('Env Relevance',     f"{d['corr'].env_relevance_score}/10" if d['corr'] else 'N/A'),
        ]
        row = 4
        for label, val in meta_rows:
            set_cell(ws, row, 1, label, font=label_font, fill=grey_fill)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            set_cell(ws, row, 2, val, font=val_font, fill=grey_fill)
            row += 1

        if d['analysis']:
            row += 1
            set_cell(ws, row, 1, 'AI Summary', font=white_font, fill=hdr_fill)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = set_cell(ws, row, 1, d['analysis'].get('summary', ''), font=val_font)
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 60

            row += 2
            set_cell(ws, row, 1, 'Recommended Action', font=white_font, fill=hdr_fill)
            ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
            c = set_cell(ws, row, 1, d['analysis'].get('recommended_action', ''), font=val_font)
            c.alignment = Alignment(wrap_text=True)
            ws.row_dimensions[row].height = 60

        ws.column_dimensions['A'].width = 22
        ws.column_dimensions['B'].width = 16
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 16
        ws.column_dimensions['E'].width = 24
        ws.column_dimensions['F'].width = 20

        # ── Unpatched Sheet ────────────────────────────────────────
        if d['active_agents']:
            ws2 = wb.create_sheet('Unpatched Agents')
            headers = ['Agent Name', 'IP Address', 'Location', 'OS', 'Status', 'Action']
            for ci, h in enumerate(headers, 1):
                c = ws2.cell(row=1, column=ci, value=h)
                c.font  = white_font
                c.fill  = PatternFill('solid', fgColor='6B0000')
                c.alignment = Alignment(horizontal='center')
            for ri, a in enumerate(d['active_agents'], 2):
                ws2.cell(row=ri, column=1, value=a.get('name','')).fill  = red_fill
                ws2.cell(row=ri, column=2, value=a.get('ip','')).fill    = red_fill
                ws2.cell(row=ri, column=3, value=a.get('location','')).fill = red_fill
                ws2.cell(row=ri, column=4, value=a.get('os','')).fill    = red_fill
                c = ws2.cell(row=ri, column=5, value='UNPATCHED')
                c.font = red_font; c.fill = red_fill
                ws2.cell(row=ri, column=6, value='Apply patch immediately').fill = red_fill
            for ci in range(1, 7):
                ws2.column_dimensions[get_column_letter(ci)].width = 20

        # ── Patched Sheet ──────────────────────────────────────────
        if d['solved_agents']:
            ws3 = wb.create_sheet('Patched Agents')
            headers = ['Agent Name', 'IP Address', 'Location', 'OS', 'Status']
            for ci, h in enumerate(headers, 1):
                c = ws3.cell(row=1, column=ci, value=h)
                c.font  = white_font
                c.fill  = PatternFill('solid', fgColor='064E3B')
                c.alignment = Alignment(horizontal='center')
            for ri, a in enumerate(d['solved_agents'], 2):
                for ci, key in enumerate(['name','ip','location','os'], 1):
                    ws3.cell(row=ri, column=ci, value=a.get(key,'')).fill = green_fill
                c = ws3.cell(row=ri, column=5, value='PATCHED')
                c.font = green_font; c.fill = green_fill
            for ci in range(1, 6):
                ws3.column_dimensions[get_column_letter(ci)].width = 20

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        cve_str = '_'.join(item.get_cve_list()[:1]) or f'item{item_id}'
        filename = f"vulnerability_report_{cve_str}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"XLSX download error for item {item_id}: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500


@threat_intel_bp.route('/api/threat-intel/item/<int:item_id>/download/pdf')
@login_required
def download_pdf(item_id):
    """Generate a management-ready PDF vulnerability report."""
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.platypus import (SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
                                        HRFlowable, KeepTogether)
        from reportlab.lib.enums import TA_LEFT, TA_CENTER, TA_RIGHT

        d = _get_item_report_data(item_id)
        item = d['item']

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4,
                                leftMargin=20*mm, rightMargin=20*mm,
                                topMargin=18*mm, bottomMargin=18*mm)

        # ── Colours ────────────────────────────────────────────────
        C_BG        = colors.HexColor('#0d1117')
        C_HEADER    = colors.HexColor('#0f2744')
        C_RED       = colors.HexColor('#dc3545')
        C_RED_DARK  = colors.HexColor('#6b0000')
        C_GREEN     = colors.HexColor('#198754')
        C_GREEN_DARK= colors.HexColor('#064e3b')
        C_ORANGE    = colors.HexColor('#fd7e14')
        C_YELLOW    = colors.HexColor('#ffc107')
        C_WHITE     = colors.white
        C_LIGHT     = colors.HexColor('#e2e8f0')
        C_MUTED     = colors.HexColor('#94a3b8')
        C_BORDER    = colors.HexColor('#1e3a5f')
        C_CELL_DARK = colors.HexColor('#0a0e1a')
        C_ROW_ALT   = colors.HexColor('#111827')

        # ── Styles ─────────────────────────────────────────────────
        sTitle    = ParagraphStyle('Title',   fontName='Helvetica',       fontSize=18, leading=22, textColor=C_WHITE, spaceAfter=4)
        sSub      = ParagraphStyle('Sub',     fontName='Helvetica',       fontSize=9,  leading=12, textColor=C_MUTED)
        sSection  = ParagraphStyle('Section', fontName='Helvetica',       fontSize=11, leading=14, textColor=C_WHITE, spaceBefore=8, spaceAfter=4)
        sBody     = ParagraphStyle('Body',    fontName='Helvetica',       fontSize=9,  leading=13, textColor=C_LIGHT)
        sLabel    = ParagraphStyle('Label',   fontName='Helvetica',       fontSize=8,  leading=10, textColor=C_MUTED)
        sBold     = ParagraphStyle('Bold',    fontName='Helvetica-Bold',  fontSize=9,  leading=13, textColor=C_WHITE)
        sWarning  = ParagraphStyle('Warning', fontName='Helvetica',       fontSize=9,  leading=13, textColor=colors.HexColor('#ff6b6b'))
        sSuccess  = ParagraphStyle('Success', fontName='Helvetica',       fontSize=9,  leading=13, textColor=colors.HexColor('#4ade80'))
        sCentre   = ParagraphStyle('Centre',  fontName='Helvetica',       fontSize=9,  leading=13, textColor=C_WHITE, alignment=TA_CENTER)

        sev_color = {
            'critical': C_RED, 'high': C_ORANGE, 'medium': C_YELLOW,
            'low': C_GREEN, 'unknown': C_MUTED
        }.get(item.severity.lower(), C_MUTED)

        story = []
        W = doc.width

        # ── Cover / Header ─────────────────────────────────────────
        hdr_data = [[
            Paragraph('<b>ByteIT SentinelX</b>', ParagraphStyle('h1', fontName='Helvetica-Bold', fontSize=14, textColor=C_WHITE)),
            Paragraph(f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}',
                      ParagraphStyle('h2', fontName='Helvetica', fontSize=8, textColor=C_MUTED, alignment=TA_RIGHT)),
        ]]
        hdr_tbl = Table(hdr_data, colWidths=[W*0.6, W*0.4])
        hdr_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), C_HEADER),
            ('TOPPADDING',    (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('LEFTPADDING',   (0,0), (0,-1),  14),
            ('RIGHTPADDING',  (-1,0),(-1,-1), 14),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(hdr_tbl)
        story.append(Spacer(1, 6*mm))

        # Title block
        story.append(Paragraph('VULNERABILITY INTELLIGENCE REPORT', sSub))
        story.append(Spacer(1, 2*mm))
        story.append(Paragraph(item.title, sTitle))
        story.append(Spacer(1, 2*mm))
        story.append(HRFlowable(width=W, thickness=1, color=C_BORDER))
        story.append(Spacer(1, 4*mm))

        # ── Meta grid ─────────────────────────────────────────────
        cves = ', '.join(item.get_cve_list()) or 'N/A'
        meta = [
            ['CVE(s)',    cves,                 'Severity', item.severity.upper()],
            ['Source',   d['source_name'],      'Published', item.published_at.strftime('%Y-%m-%d') if item.published_at else '–'],
            ['Patch',    'Available' if item.has_patch else 'Not available',
             'Env Relevance', f"{d['corr'].env_relevance_score}/10" if d['corr'] else 'N/A'],
        ]
        meta_rows_flat = []
        for row in meta:
            meta_rows_flat.append([
                Paragraph(row[0], sLabel),
                Paragraph(f'<b>{row[1]}</b>', sBold),
                Paragraph(row[2], sLabel),
                Paragraph(f'<b>{row[3]}</b>', sBold),
            ])

        meta_tbl = Table(meta_rows_flat, colWidths=[W*0.14, W*0.36, W*0.14, W*0.36])
        meta_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), C_CELL_DARK),
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [C_CELL_DARK, C_ROW_ALT]),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (-1,-1), 8),
            ('RIGHTPADDING',  (0,0), (-1,-1), 8),
            ('BOX',           (0,0), (-1,-1), 0.5, C_BORDER),
            ('INNERGRID',     (0,0), (-1,-1), 0.3, C_BORDER),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(meta_tbl)
        story.append(Spacer(1, 5*mm))

        # ── Action Required Banner ─────────────────────────────────
        active_count = len(d['active_agents'])
        solved_count = len(d['solved_agents'])
        total_wazuh  = active_count + solved_count

        if active_count > 0 and solved_count > 0:
            banner_color = C_RED
            banner_bg    = colors.HexColor('#2a0a0a')
            banner_text  = (f'<b>⚠ ACTION REQUIRED — PARTIAL REMEDIATION</b><br/>'
                            f'{active_count} of {total_wazuh} agent(s) are UNPATCHED. '
                            f'{solved_count} agent(s) have been remediated. '
                            f'Apply the patch to the remaining {active_count} unpatched agent(s) immediately.')
        elif active_count > 0:
            banner_color = C_RED
            banner_bg    = colors.HexColor('#2a0a0a')
            banner_text  = (f'<b>⚠ ACTION REQUIRED — IMMEDIATE PATCHING NEEDED</b><br/>'
                            f'Wazuh Vulnerability Detector has confirmed this CVE is UNPATCHED on '
                            f'{active_count} agent(s). Apply the available patch immediately.')
        elif solved_count > 0:
            banner_color = C_GREEN
            banner_bg    = colors.HexColor('#0a2a15')
            banner_text  = (f'<b>✓ NO ACTION REQUIRED — FULLY REMEDIATED</b><br/>'
                            f'All {solved_count} detected agent(s) have had the patch applied. '
                            f'This CVE is RESOLVED across your entire monitored environment.')
        elif d['other_agents']:
            banner_color = C_YELLOW
            banner_bg    = colors.HexColor('#2a1f00')
            banner_text  = (f'<b>⚠ INVESTIGATE — PLATFORM MATCH DETECTED</b><br/>'
                            f'{len(d["other_agents"])} agent(s) match this threat\'s platform/technology. '
                            f'This is not a confirmed incident — investigate individually to confirm.')
        else:
            banner_color = C_MUTED
            banner_bg    = C_CELL_DARK
            banner_text  = '<b>No infrastructure correlation data available.</b> Run infrastructure correlation to assess exposure.'

        banner_tbl = Table([[Paragraph(banner_text, ParagraphStyle('banner', fontName='Helvetica',
                                                                    fontSize=9, leading=14, textColor=C_WHITE))]],
                            colWidths=[W])
        banner_tbl.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), banner_bg),
            ('BOX',        (0,0), (-1,-1), 1.5, banner_color),
            ('TOPPADDING',    (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
            ('LEFTPADDING',   (0,0), (-1,-1), 12),
            ('RIGHTPADDING',  (0,0), (-1,-1), 12),
        ]))
        story.append(banner_tbl)
        story.append(Spacer(1, 5*mm))

        # ── AI Summary ─────────────────────────────────────────────
        if d['analysis'].get('summary'):
            story.append(Paragraph('AI Security Analysis', sSection))
            summary_tbl = Table([[Paragraph(d['analysis']['summary'], sBody)]], colWidths=[W])
            summary_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#0f2744')),
                ('BOX',        (0,0), (-1,-1), 0.5, C_BORDER),
                ('TOPPADDING',    (0,0), (-1,-1), 8),
                ('BOTTOMPADDING', (0,0), (-1,-1), 8),
                ('LEFTPADDING',   (0,0), (-1,-1), 10),
                ('RIGHTPADDING',  (0,0), (-1,-1), 10),
            ]))
            story.append(summary_tbl)
            story.append(Spacer(1, 3*mm))

            if d['analysis'].get('recommended_action'):
                action_tbl = Table([
                    [Paragraph('<b>Recommended Action</b>', ParagraphStyle('ra', fontName='Helvetica-Bold',
                                                                            fontSize=9, textColor=colors.HexColor('#fbbf24')))],
                    [Paragraph(d['analysis']['recommended_action'], sBody)],
                ], colWidths=[W])
                action_tbl.setStyle(TableStyle([
                    ('BACKGROUND', (0,0), (0,0), colors.HexColor('#1a1a00')),
                    ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#1a2f4a')),
                    ('LINEBEFORE', (0,0), (0,-1), 3, colors.HexColor('#f59e0b')),
                    ('TOPPADDING',    (0,0), (-1,-1), 7),
                    ('BOTTOMPADDING', (0,0), (-1,-1), 7),
                    ('LEFTPADDING',   (0,0), (-1,-1), 10),
                    ('RIGHTPADDING',  (0,0), (-1,-1), 10),
                    ('BOX',           (0,0), (-1,-1), 0.5, C_BORDER),
                ]))
                story.append(action_tbl)
                story.append(Spacer(1, 5*mm))

        # ── Env-specific action from correlation ───────────────────
        if d['corr'] and d['corr'].env_recommended_action:
            env_tbl = Table([
                [Paragraph('<b>Environment-Specific Action (Wazuh Correlation)</b>',
                           ParagraphStyle('ea', fontName='Helvetica-Bold', fontSize=9, textColor=colors.HexColor('#fbbf24')))],
                [Paragraph(d['corr'].env_recommended_action, sBody)],
            ], colWidths=[W])
            env_tbl.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (0,0), colors.HexColor('#1a1500')),
                ('BACKGROUND', (0,1), (-1,-1), colors.HexColor('#1a1f00')),
                ('LINEBEFORE', (0,0), (0,-1), 3, C_YELLOW),
                ('TOPPADDING',    (0,0), (-1,-1), 7),
                ('BOTTOMPADDING', (0,0), (-1,-1), 7),
                ('LEFTPADDING',   (0,0), (-1,-1), 10),
                ('RIGHTPADDING',  (0,0), (-1,-1), 10),
                ('BOX',           (0,0), (-1,-1), 0.5, C_BORDER),
            ]))
            story.append(env_tbl)
            story.append(Spacer(1, 5*mm))

        # ── Agent Tables ───────────────────────────────────────────
        def agent_table(title, agents, hdr_fill, status_label, status_color):
            if not agents:
                return
            story.append(Paragraph(title, sSection))
            tbl_data = [[
                Paragraph('<b>Agent Name</b>', sCentre),
                Paragraph('<b>IP Address</b>', sCentre),
                Paragraph('<b>Location</b>',   sCentre),
                Paragraph('<b>OS</b>',          sCentre),
                Paragraph('<b>Status</b>',      sCentre),
            ]]
            for i, a in enumerate(agents):
                row_bg = C_CELL_DARK if i % 2 == 0 else C_ROW_ALT
                tbl_data.append([
                    Paragraph(a.get('name','–'), sBody),
                    Paragraph(a.get('ip','–'),   sBody),
                    Paragraph(a.get('location','–'), sBody),
                    Paragraph(a.get('os','–'),   sBody),
                    Paragraph(f'<b>{status_label}</b>',
                               ParagraphStyle('sl', fontName='Helvetica-Bold', fontSize=9,
                                              textColor=status_color, alignment=TA_CENTER)),
                ])
            tbl = Table(tbl_data, colWidths=[W*0.25, W*0.18, W*0.18, W*0.25, W*0.14])
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), hdr_fill),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [C_CELL_DARK, C_ROW_ALT]),
                ('BOX',       (0,0), (-1,-1), 0.5, C_BORDER),
                ('INNERGRID', (0,0), (-1,-1), 0.3, C_BORDER),
                ('TOPPADDING',    (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('LEFTPADDING',   (0,0), (-1,-1), 6),
                ('RIGHTPADDING',  (0,0), (-1,-1), 6),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 4*mm))

        agent_table('Unpatched Agents — Immediate Action Required',
                    d['active_agents'], C_RED_DARK, 'UNPATCHED', C_RED)
        agent_table('Patched Agents — Remediation Complete',
                    d['solved_agents'], C_GREEN_DARK, 'PATCHED', C_GREEN)

        if d['other_agents']:
            story.append(Paragraph('Platform Match Agents — Investigate', sSection))
            tbl_data = [[
                Paragraph('<b>Agent Name</b>', sCentre),
                Paragraph('<b>IP Address</b>', sCentre),
                Paragraph('<b>OS</b>',          sCentre),
                Paragraph('<b>Match Reason</b>', sCentre),
            ]]
            for i, a in enumerate(d['other_agents']):
                tbl_data.append([
                    Paragraph(a.get('name','–'), sBody),
                    Paragraph(a.get('ip','–'),   sBody),
                    Paragraph(a.get('os','–'),   sBody),
                    Paragraph('; '.join(a.get('reasons',[])), sBody),
                ])
            tbl = Table(tbl_data, colWidths=[W*0.22, W*0.17, W*0.23, W*0.38])
            tbl.setStyle(TableStyle([
                ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#2a1f00')),
                ('ROWBACKGROUNDS', (0,1), (-1,-1), [C_CELL_DARK, C_ROW_ALT]),
                ('BOX',       (0,0), (-1,-1), 0.5, C_BORDER),
                ('INNERGRID', (0,0), (-1,-1), 0.3, C_BORDER),
                ('TOPPADDING',    (0,0), (-1,-1), 5),
                ('BOTTOMPADDING', (0,0), (-1,-1), 5),
                ('LEFTPADDING',   (0,0), (-1,-1), 6),
                ('RIGHTPADDING',  (0,0), (-1,-1), 6),
                ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ]))
            story.append(tbl)
            story.append(Spacer(1, 4*mm))

        # ── Footer ─────────────────────────────────────────────────
        story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER))
        story.append(Spacer(1, 2*mm))
        footer_data = [[
            Paragraph('CONFIDENTIAL — FOR MANAGEMENT USE ONLY', ParagraphStyle('fl', fontName='Helvetica',
                       fontSize=7, textColor=C_MUTED)),
            Paragraph(f'ByteIT SentinelX  ·  {datetime.utcnow().strftime("%Y-%m-%d")}',
                      ParagraphStyle('fr', fontName='Helvetica', fontSize=7, textColor=C_MUTED, alignment=TA_RIGHT)),
        ]]
        footer_tbl = Table(footer_data, colWidths=[W*0.6, W*0.4])
        footer_tbl.setStyle(TableStyle([('TOPPADDING',(0,0),(-1,-1),0),('BOTTOMPADDING',(0,0),(-1,-1),0)]))
        story.append(footer_tbl)

        def on_page(canvas_obj, doc_obj):
            canvas_obj.saveState()
            canvas_obj.setFillColor(C_BG)
            canvas_obj.rect(0, 0, A4[0], A4[1], fill=True, stroke=False)
            canvas_obj.restoreState()

        doc.build(story, onFirstPage=on_page, onLaterPages=on_page)
        buf.seek(0)
        cve_str = '_'.join(item.get_cve_list()[:1]) or f'item{item_id}'
        filename = f"vulnerability_report_{cve_str}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        return send_file(buf, mimetype='application/pdf', as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"PDF download error for item {item_id}: {e}", exc_info=True)
        return jsonify({'error': str(e)}), 500
