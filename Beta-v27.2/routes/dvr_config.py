"""
AI DVR/NVR Auto-Configuration — Blueprint
Handles the 11-step guided workflow for configuring Hikvision / Platinum / LTS devices.
"""
import hashlib
import io
import json
import logging
import re
import smtplib
import time
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import requests
from requests.auth import HTTPDigestAuth

from flask import (Blueprint, Response, jsonify, render_template,
                   request, send_file, session)
from flask_login import current_user, login_required

from config import Config
from models import DVRConfigSession, db
from routes.permissions import make_blueprint_permission_check

logger = logging.getLogger(__name__)

dvr_config_bp = Blueprint('dvr_config', __name__)
dvr_config_bp.before_request(make_blueprint_permission_check('dvr_config'))


# ─────────────────────────────────────────────
# Overall status computation (shared by validate + bulk)
# ─────────────────────────────────────────────

def _compute_overall(checks):
    """Compute 4-tier overall status from a list of validation check dicts.

    Tiers (in priority order):
      FAIL            — any check with severity='critical' has status='fail'
      PARTIAL SUCCESS — any non-critical check fails, OR any major-severity check warns
      WARNING         — only minor-severity warnings remain (no failures)
      PASS            — all checks pass, no warnings

    The 'severity' field on each check dict is set by dvr_api.validate_configuration:
      'critical' : connectivity / authentication failure
      'major'    : stream config, recording, HDD, user accounts
      'minor'    : timezone, NTP mode, device name, channel zero, etc.
    """
    critical_fail  = any(c['status'] == 'fail'  and c.get('severity') == 'critical' for c in checks)
    any_fail       = any(c['status'] == 'fail'  for c in checks)
    major_warn     = any(c['status'] == 'warn'  and c.get('severity', 'major') == 'major' for c in checks)
    any_warn       = any(c['status'] == 'warn'  for c in checks)

    if critical_fail:
        return 'FAIL'
    if any_fail or major_warn:
        return 'PARTIAL SUCCESS'
    if any_warn:
        return 'WARNING'
    return 'PASS'


# ─────────────────────────────────────────────
# Helper: build DVRClient from session data
# ─────────────────────────────────────────────

def _make_client(conn):
    from dvr_api import DVRClient
    return DVRClient(
        ip=conn['ip'],
        port=int(conn.get('port', 80)),
        username=conn['username'],
        password=conn['password'],
        use_https=conn.get('use_https', False),
    )


def _get_session_data(session_id):
    rec = DVRConfigSession.query.get(session_id)
    if not rec:
        return None
    return rec.get_data()


def _save_session_data(session_id, data):
    rec = DVRConfigSession.query.get(session_id)
    if not rec:
        rec = DVRConfigSession(id=session_id)
        db.session.add(rec)
    rec.set_data(data)
    rec.updated_at = datetime.utcnow()
    db.session.commit()


# ─────────────────────────────────────────────
# Main Page
# ─────────────────────────────────────────────

@dvr_config_bp.route('/dvr-config')
@login_required
def index():
    from dvr_api import STATE_TIMEZONE, TZ_TO_HIKVISION
    return render_template(
        'dvr_config.html',
        state_timezone_json=json.dumps(STATE_TIMEZONE),
        tz_to_hikvision_json=json.dumps(TZ_TO_HIKVISION),
    )


@dvr_config_bp.route('/dvr/history')
@login_required
def history_page():
    return render_template('dvr_history.html')


@dvr_config_bp.route('/dvr/templates')
@login_required
def templates_page():
    return render_template('dvr_templates.html')


@dvr_config_bp.route('/dvr/reports')
@login_required
def reports_page():
    return render_template('dvr_reports.html')


# ─────────────────────────────────────────────
# Step 1: Connect & Authenticate
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/connect', methods=['POST'])
@login_required
def connect():
    data = request.get_json() or {}
    required = ['device_name', 'ip', 'port', 'username', 'password']
    for f in required:
        if not data.get(f):
            return jsonify({'success': False, 'error': f'Missing field: {f}'}), 400

    try:
        client = _make_client(data)
        ok = client.test_connection()
        if not ok:
            return jsonify({'success': False,
                            'error': 'Authentication failed. Check IP, port, username and password.'}), 401

        # Push the friendly name onto the device itself — this previously
        # only lived in our local session/report data and never reached the
        # device, so the DVR/NVR's own name never actually changed.
        name_applied = client.configure_device_name(data['device_name'])
        if not name_applied:
            logger.warning(f"Could not set device name on {data['ip']}; continuing anyway")

        # Create a new config session
        sess_id = f"{current_user.id}_{int(time.time())}"
        sess_data = {
            'session_id': sess_id,
            'technician': current_user.username,
            'started_at': datetime.utcnow().isoformat(),
            'connection': {
                'device_name': data['device_name'],
                'ip': data['ip'],
                'port': data['port'],
                'username': data['username'],
                'password': data['password'],
                'use_https': data.get('use_https', False),
            },
            'steps': {},
        }
        _save_session_data(sess_id, sess_data)

        return jsonify({'success': True, 'session_id': sess_id})
    except Exception as e:
        logger.error(f"DVR connect error: {e}", exc_info=True)
        return jsonify({'success': False,
                        'error': f'Connection failed: {str(e)}. Verify the device is reachable.'}), 500


# ─────────────────────────────────────────────
# Step 2: Device Discovery
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/discover', methods=['POST'])
@login_required
def discover():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        info = client.get_device_info()

        sess['steps']['discovery'] = info
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'device_info': info})
    except Exception as e:
        logger.error(f"DVR discover error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 3: Location / Time / NTP
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-time', methods=['POST'])
@login_required
def configure_time():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    city = data.get('city', '')
    state = data.get('state', '')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        result = client.configure_time(city, state)
        result['city'] = city
        result['state'] = state

        sess['steps']['time'] = result
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'time_config': result})
    except Exception as e:
        logger.error(f"DVR time config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 4: Storage Detection & Formatting
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/storage', methods=['POST'])
@login_required
def storage_info():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        hdds = client.get_storage_info()

        sess['steps']['storage'] = {'hdds': hdds}
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'hdds': hdds})
    except Exception as e:
        logger.error(f"DVR storage error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@dvr_config_bp.route('/api/dvr/format-hdd', methods=['POST'])
@login_required
def format_hdd():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    hdd_id = data.get('hdd_id', '1')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        client.format_hdd(hdd_id)
        time.sleep(3)  # Brief pause; formatting is async on device
        hdds = client.get_storage_info()

        sess['steps']['storage'] = {'hdds': hdds, 'formatted': [hdd_id]}
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'message': f'HDD {hdd_id} formatting initiated.', 'hdds': hdds})
    except Exception as e:
        logger.error(f"DVR format HDD error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 5: Camera Stream Configuration
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-streams', methods=['POST'])
@login_required
def configure_streams():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        discovery = sess['steps'].get('discovery', {})
        # Prefer real channel IDs discovered from /Streaming/channels during
        # device discovery (e.g. ["101","201","301"]).  Fall back to synthesising
        # them from total_channels only when discovery didn't capture them.
        channel_ids = discovery.get('channel_ids') or []
        if not channel_ids:
            total_channels = int(discovery.get('total_channels', 0)) or 4
            channel_ids = [f"{ch}01" for ch in range(1, total_channels + 1)]

        results = []
        for main_id in channel_ids:
            ch_label = main_id[:-2] if len(main_id) > 2 else main_id
            main_res = client.configure_main_stream(main_id)
            sub_res  = client.configure_sub_stream(main_id)
            results.append({
                'channel':              ch_label,
                'main_stream':          main_res['status'],
                'main_resolution':      main_res['applied_resolution'],
                'main_bitrate':         main_res['applied_bitrate'],
                'main_note':            main_res['note'],
                'sub_stream':           sub_res['status'],
                'sub_resolution':       sub_res['applied_resolution'],
                'sub_bitrate':          sub_res['applied_bitrate'],
                'sub_note':             sub_res['note'],
            })

        client.disable_channel_zero()

        sess['steps']['streams'] = {'channels': results, 'channel_zero_disabled': True}
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'channels': results})
    except Exception as e:
        logger.error(f"DVR stream config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 6: Recording Schedule
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-recording', methods=['POST'])
@login_required
def configure_recording():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        discovery = sess['steps'].get('discovery', {})
        channel_ids = discovery.get('channel_ids') or []
        if not channel_ids:
            total_channels = int(discovery.get('total_channels', 0)) or 4
            channel_ids = [f"{ch}01" for ch in range(1, total_channels + 1)]

        results = []
        for main_id in channel_ids:
            ch_label = main_id[:-2] if len(main_id) > 2 else main_id
            ok = client.configure_recording_schedule(main_id)
            results.append({'channel': ch_label, 'status': 'configured' if ok else 'error'})

        sess['steps']['recording'] = {'channels': results}
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'channels': results})
    except Exception as e:
        logger.error(f"DVR recording config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 7: Holiday Schedule
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-holiday', methods=['POST'])
@login_required
def configure_holiday():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        ok = client.configure_holiday()

        holiday_info = {
            'name': 'Thanksgiving & Black Friday',
            'type': 'By Week',
            'start': '4th Monday of November',
            'end': '1st Monday of December',
            'recording_mode': 'Continuous',
            'status': 'configured' if ok else 'warn',
        }
        sess['steps']['holiday'] = holiday_info
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'holiday': holiday_info})
    except Exception as e:
        logger.error(f"DVR holiday config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 8 & 9: Users + Permissions
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-users', methods=['POST'])
@login_required
def configure_users():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    initials = (data.get('client_initials') or '').strip().upper()
    # User Selection: caller chooses which standard accounts to actually
    # create on the device. Defaults to both for backward compatibility.
    selected_users = data.get('selected_users')
    if not selected_users:
        selected_users = ['cms', 'dlt']
    selected_users = [u for u in selected_users if u in ('cms', 'dlt', 'manager')]

    # Manager account: optional. The operator either ticks the "create
    # manager" checkbox, or types a manager username/password directly into
    # the wizard — either counts as "credentials provided" and triggers
    # creation. If nothing is provided, the manager step is skipped silently
    # (not an error, not a failed check).
    manager_username_input = (data.get('manager_username') or '').strip()
    manager_password_input = (data.get('manager_password') or '').strip()
    create_manager = (
        bool(data.get('create_manager', False))
        or 'manager' in selected_users
        or bool(manager_username_input)
        or bool(manager_password_input)
    )
    if create_manager and 'manager' not in selected_users:
        selected_users.append('manager')

    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404
    if not initials:
        return jsonify({'success': False, 'error': 'Client initials are required'}), 400
    core_users = [u for u in selected_users if u in ('cms', 'dlt')]
    if not core_users:
        return jsonify({'success': False, 'error': 'Select at least one account to create'}), 400

    cms_password     = f"{initials}_cam12"
    dlt_password     = f"{initials}9722IDT!"
    # Default manager username is "manager" (editable via manager_username);
    # default password follows the client-initials standard (editable via
    # manager_password), e.g. "RCman_12" for client initials "RC".
    manager_username = manager_username_input or 'manager'
    manager_password = manager_password_input or f"{initials}man_12"

    admin_verification_password = (data.get('admin_verification_password') or '').strip() or None

    results = []
    try:
        client = _make_client(sess['connection'])
        # Inject admin verification password so the client can include
        # <adminPassword> in user-creation XML for devices that require it.
        if admin_verification_password:
            client.admin_verification_password = admin_verification_password

        def _account_status(result, created_label='created'):
            """Derive a human-readable status string from configure_standard_account result."""
            if not result.get('ok'):
                return 'error'
            already = result.get('already_existed', False)
            perm_st = result.get('permission_status', '')
            if perm_st == 'update_failed':
                return 'updated_permissions_failed' if already else 'created_permissions_failed'
            return 'updated' if already else created_label

        # CMS user (Viewer-level monitoring account).
        # Uses configure_standard_account: checks if the account exists first;
        # creates it if missing, updates permissions if they differ, and
        # continues without error either way.
        if 'cms' in selected_users:
            try:
                cms = client.configure_standard_account("cms", cms_password, "cms")
                status = _account_status(cms)
                entry = {
                    'username': 'cms',
                    'password': cms_password,
                    'role': 'viewer',
                    'status': status,
                    'already_existed': cms.get('already_existed', False),
                    'permission_status': cms.get('permission_status'),
                    'verified': cms.get('verified', False),
                    'permissions': 'Log Search, Live View, Playback, Video Export (Local); Log Search, Live View, Playback/Download (Remote)',
                }
                if status == 'error':
                    entry['error'] = cms.get('error', 'Unknown error')
                results.append(entry)
            except Exception as e:
                results.append({
                    'username': 'cms', 'password': cms_password,
                    'status': 'error', 'error': str(e),
                })

        # DLT user (Operator-level remote-admin account).
        # Same smart check-create-update flow as CMS.
        if 'dlt' in selected_users:
            try:
                dlt = client.configure_standard_account("dlt", dlt_password, "dlt")
                status = _account_status(dlt)
                entry = {
                    'username': 'dlt',
                    'password': dlt_password,
                    'role': 'operator',
                    'status': status,
                    'already_existed': dlt.get('already_existed', False),
                    'permission_status': dlt.get('permission_status'),
                    'verified': dlt.get('verified', False),
                    'permissions': 'Parameter Settings, Log Search, Shutdown/Reboot, Live View, Playback/Download (Remote)',
                }
                if status == 'error':
                    entry['error'] = dlt.get('error', 'Unknown error')
                results.append(entry)
            except Exception as e:
                results.append({
                    'username': 'dlt', 'password': dlt_password,
                    'status': 'error', 'error': str(e),
                })

        # Manager user (optional — Viewer-level, same permissions as CMS).
        # Creates the account if new, updates password if it already exists,
        # and only rewrites permissions when they differ from the required set.
        if 'manager' in selected_users:
            try:
                mgr = client.configure_standard_account(manager_username, manager_password, "manager")
                status = _account_status(mgr)
                entry = {
                    'username': manager_username,
                    'password': manager_password,
                    'role': 'viewer',
                    'status': status,
                    'already_existed': mgr.get('already_existed', False),
                    'permission_status': mgr.get('permission_status'),
                    'verified': mgr.get('verified', False),
                    'permissions': 'Log Search, Live View, Playback, Video Export (Local); Log Search, Live View, Playback/Download (Remote)',
                }
                if status == 'error':
                    entry['error'] = mgr.get('error', 'Unknown error')
                results.append(entry)
            except Exception as e:
                results.append({
                    'username': manager_username, 'password': manager_password,
                    'status': 'error', 'error': str(e),
                })

        sess['steps']['users'] = {
            'client_initials': initials,
            'selected_users': selected_users,
            'manager_created': 'manager' in selected_users,
            'accounts': results,
        }
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'users': results})
    except Exception as e:
        logger.error(f"DVR user config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 10: Validation
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/validate', methods=['POST'])
@login_required
def validate():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        expected = dict(sess['steps'].get('time', {}))
        expected['device_name'] = sess['connection'].get('device_name', '')
        users_step = sess['steps'].get('users', {})
        expected['users_created'] = [
            a['username'] for a in users_step.get('accounts', [])
            if a.get('status') in ('created', 'created_permissions_failed')
        ]
        # Use the same channel IDs discovered/configured earlier in the wizard
        # so the validation report checks every real channel on this device
        # instead of falling back to a single hardcoded channel.
        discovery = sess['steps'].get('discovery', {})
        expected['channel_ids'] = discovery.get('channel_ids') or []
        checks = client.validate_configuration(expected)

        passed = sum(1 for c in checks if c['status'] == 'pass')
        warned = sum(1 for c in checks if c['status'] == 'warn')
        failed = sum(1 for c in checks if c['status'] == 'fail')
        overall = _compute_overall(checks)

        sess['steps']['validation'] = {
            'checks': checks,
            'passed': passed,
            'warned': warned,
            'failed': failed,
            'overall': overall,
            'completed_at': datetime.utcnow().isoformat(),
        }
        _save_session_data(sess_id, sess)

        return jsonify({
            'success': True,
            'checks': checks,
            'summary': {'passed': passed, 'warned': warned, 'failed': failed, 'overall': overall},
        })
    except Exception as e:
        logger.error(f"DVR validate error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 11: Report Generation
# ─────────────────────────────────────────────

def _build_report_data(sess):
    conn = sess.get('connection', {})
    disc = sess['steps'].get('discovery', {})
    time_cfg = sess['steps'].get('time', {})
    storage = sess['steps'].get('storage', {})
    streams = sess['steps'].get('streams', {})
    recording = sess['steps'].get('recording', {})
    holiday = sess['steps'].get('holiday', {})
    users_step = sess['steps'].get('users', {})
    validation = sess['steps'].get('validation', {})

    started = datetime.fromisoformat(sess.get('started_at', datetime.utcnow().isoformat()))
    completed_str = validation.get('completed_at', datetime.utcnow().isoformat())
    completed = datetime.fromisoformat(completed_str)
    duration_sec = int((completed - started).total_seconds())
    duration = f"{duration_sec // 60}m {duration_sec % 60}s"

    return {
        'device_name': conn.get('device_name', 'N/A'),
        'ip_address': conn.get('ip', 'N/A'),
        'manufacturer': disc.get('manufacturer', 'N/A'),
        'model': disc.get('model', 'N/A'),
        'device_type': disc.get('device_type', 'N/A'),
        'serial_number': disc.get('serial_number', 'N/A'),
        'firmware_version': disc.get('firmware_version', 'N/A'),
        'total_channels': disc.get('total_channels', 'N/A'),
        'active_channels': disc.get('active_channels', 'N/A'),
        'mac_address': disc.get('mac_address', 'N/A'),
        'timezone': time_cfg.get('iana_timezone', 'N/A'),
        'ntp_server': time_cfg.get('ntp_server', 'pool.ntp.org'),
        'city': time_cfg.get('city', 'N/A'),
        'state': time_cfg.get('state', 'N/A'),
        'hdds': storage.get('hdds', []),
        'stream_channels': streams.get('channels', []),
        'channel_zero_disabled': streams.get('channel_zero_disabled', False),
        'recording_channels': recording.get('channels', []),
        'holiday': holiday,
        'users': users_step.get('accounts', []),
        'client_initials': users_step.get('client_initials', ''),
        'validation': validation.get('checks', []),
        'validation_summary': {
            'passed': validation.get('passed', 0),
            'warned': validation.get('warned', 0),
            'failed': validation.get('failed', 0),
            'overall': validation.get('overall', 'N/A'),
        },
        'technician': sess.get('technician', 'N/A'),
        'configured_at': completed.strftime('%Y-%m-%d %H:%M:%S'),
        'total_duration': duration,
    }


def _build_html_report(report):
    import html as _html_mod

    def _e(v):
        return _html_mod.escape(str(v or ''))

    # ── Colour maps ────────────────────────────────────────────────────────
    OVERALL_BG   = {'PASS': '#1a7f4b', 'WARNING': '#b07800', 'PARTIAL SUCCESS': '#0d6eac', 'FAIL': '#c0392b'}
    OVERALL_ICON = {'PASS': '✅', 'WARNING': '⚠️', 'PARTIAL SUCCESS': '🟡', 'FAIL': '❌'}
    OVERALL_LABEL = {
        'PASS':           'All checks passed — configuration complete.',
        'WARNING':        'Configuration complete with minor advisories.',
        'PARTIAL SUCCESS':'Configuration complete. Some settings applied via fallback or require review.',
        'FAIL':           'Critical failure — immediate attention required.',
    }
    ST_ICON  = {'pass': '✅', 'warn': '⚠️', 'fail': '❌'}
    ST_COLOR = {'pass': '#1a7f4b', 'warn': '#b07800', 'fail': '#c0392b'}
    ST_BG    = {'pass': '#eafaf1', 'warn': '#fffbea', 'fail': '#fdf3f2'}
    STREAM_ST_COLOR = {'configured': '#1a7f4b', 'configured_fallback': '#b07800', 'error': '#c0392b', 'skipped': '#6c757d'}
    STREAM_ST_ICON  = {'configured': '✅', 'configured_fallback': '⚠️', 'error': '❌', 'skipped': '—'}

    overall = report['validation_summary']['overall']
    vs      = report['validation_summary']
    hdr_bg  = OVERALL_BG.get(overall, '#555')
    hdr_icon= OVERALL_ICON.get(overall, '?')
    hdr_lbl = OVERALL_LABEL.get(overall, overall)

    # ── Validation checks table ────────────────────────────────────────────
    checks_rows = ''
    for c in report['validation']:
        icon  = ST_ICON.get(c['status'], '?')
        color = ST_COLOR.get(c['status'], '#555')
        bg    = ST_BG.get(c['status'], '#fff')
        action= (f'<div class="action-note">⟶ {_e(c["action"])}</div>'
                 if c.get('action') else '')
        checks_rows += f"""
        <tr style="background:{bg};">
          <td style="color:{color};font-size:1.15em;text-align:center;width:36px;">{icon}</td>
          <td style="font-weight:600;white-space:nowrap;">{_e(c['check'])}</td>
          <td>{_e(c['detail'])}{action}</td>
        </tr>"""

    # ── Users table ────────────────────────────────────────────────────────
    users_rows = ''
    for u in report['users']:
        ok  = u.get('status') == 'created'
        c   = '#1a7f4b' if ok else '#c0392b'
        ico = '✅' if ok else '❌'
        users_rows += f"""
        <tr>
          <td><strong>{_e(u['username'])}</strong></td>
          <td><code>{_e(u.get('password','N/A'))}</code></td>
          <td>{_e(u.get('role','N/A'))}</td>
          <td style="color:{c};font-weight:600;">{ico} {_e(u.get('status','N/A'))}</td>
        </tr>"""

    # ── HDD table ─────────────────────────────────────────────────────────
    hdds_rows = ''
    for h in report['hdds']:
        st    = h.get('status','').lower()
        ok    = st in ('ok','normal','active')
        c     = '#1a7f4b' if ok else '#c0392b'
        ico   = '✅' if ok else '❌'
        hdds_rows += (
            f"<tr>"
            f"<td>{_e(h['id'])}</td><td>{_e(h['name'])}</td>"
            f"<td>{_e(h['capacity'])} MB</td>"
            f"<td style='color:{c};font-weight:600;'>{ico} {_e(h['status'])}</td>"
            f"<td>{_e(h.get('free_space','N/A'))} MB</td>"
            f"<td>{_e(h['property'])}</td>"
            f"</tr>"
        )
    if not hdds_rows:
        hdds_rows = "<tr><td colspan='6' style='color:#999;text-align:center;'>No HDD data available</td></tr>"

    # ── Stream channels table ──────────────────────────────────────────────
    stream_rows = ''
    for sc in report.get('stream_channels', []):
        main_st  = sc.get('main_stream', sc.get('main', 'N/A'))
        main_res = sc.get('main_resolution', 'N/A')
        main_br  = sc.get('main_bitrate', 'N/A')
        sub_st   = sc.get('sub_stream',  sc.get('sub',  'N/A'))
        sub_res  = sc.get('sub_resolution', 'N/A')
        sub_br   = sc.get('sub_bitrate', 'N/A')
        notes    = '; '.join(filter(None, [sc.get('main_note',''), sc.get('sub_note','')]))
        mic = STREAM_ST_COLOR.get(main_st, '#555')
        mii = STREAM_ST_ICON.get(main_st, '?')
        sic = STREAM_ST_COLOR.get(sub_st,  '#555')
        sii = STREAM_ST_ICON.get(sub_st,   '?')
        stream_rows += (
            f"<tr>"
            f"<td style='font-weight:600;'>{_e(sc.get('channel','?'))}</td>"
            f"<td style='color:{mic};font-weight:600;'>{mii} {_e(main_st)}</td>"
            f"<td>{_e(main_res)}</td><td>{_e(main_br)} kbps</td>"
            f"<td style='color:{sic};font-weight:600;'>{sii} {_e(sub_st)}</td>"
            f"<td>{_e(sub_res)}</td><td>{_e(sub_br)} kbps</td>"
            f"<td style='color:#666;font-size:0.88em;'>{_e(notes)}</td>"
            f"</tr>"
        )
    if not stream_rows:
        stream_rows = "<tr><td colspan='8' style='color:#999;text-align:center;'>No stream data available</td></tr>"

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>DVR/NVR Configuration Report — {_e(report['device_name'])}</title>
<style>
  *{{box-sizing:border-box;margin:0;padding:0;}}
  body{{font-family:'Segoe UI',Arial,sans-serif;font-size:13px;color:#1a1a2e;background:#f4f6f9;}}
  .page{{max-width:960px;margin:0 auto;background:#fff;box-shadow:0 2px 16px rgba(0,0,0,.12);}}

  /* ── Header ── */
  .report-header{{background:{hdr_bg};color:#fff;padding:28px 36px 22px;}}
  .report-header .brand{{font-size:1.1em;opacity:.85;letter-spacing:.04em;text-transform:uppercase;margin-bottom:4px;}}
  .report-header h1{{font-size:1.7em;font-weight:700;margin-bottom:2px;}}
  .report-header .meta{{font-size:.88em;opacity:.82;margin-top:8px;}}
  .overall-banner{{background:rgba(0,0,0,.18);border-radius:8px;padding:14px 20px;margin-top:16px;
                   display:flex;align-items:center;gap:16px;flex-wrap:wrap;}}
  .overall-icon{{font-size:2em;}}
  .overall-text .title{{font-size:1.25em;font-weight:700;}}
  .overall-text .sub{{font-size:.9em;opacity:.9;margin-top:2px;}}

  /* ── Score cards ── */
  .score-bar{{display:flex;gap:0;border-bottom:3px solid #e8ecf0;}}
  .score-card{{flex:1;padding:14px 0;text-align:center;border-right:1px solid #e8ecf0;}}
  .score-card:last-child{{border-right:none;}}
  .score-card .num{{font-size:2em;font-weight:700;}}
  .score-card .lbl{{font-size:.78em;text-transform:uppercase;letter-spacing:.05em;color:#6c757d;margin-top:2px;}}
  .score-card.pass  .num{{color:#1a7f4b;}}
  .score-card.warn  .num{{color:#b07800;}}
  .score-card.fail  .num{{color:#c0392b;}}
  .score-card.total .num{{color:#0d6eac;}}

  /* ── Sections ── */
  .section{{padding:24px 36px;border-bottom:1px solid #e8ecf0;}}
  .section:last-child{{border-bottom:none;}}
  .section-title{{font-size:1em;font-weight:700;text-transform:uppercase;letter-spacing:.07em;
                  color:#0056b3;border-left:4px solid #0056b3;padding-left:10px;margin-bottom:16px;}}

  /* ── Info grid ── */
  .info-grid{{display:grid;grid-template-columns:repeat(3,1fr);gap:10px;}}
  .info-cell{{background:#f9fbff;border:1px solid #dce4f0;border-radius:6px;padding:10px 12px;}}
  .info-cell label{{font-size:.75em;color:#6c757d;display:block;text-transform:uppercase;letter-spacing:.04em;margin-bottom:3px;}}
  .info-cell span{{font-size:.95em;font-weight:600;color:#1a1a2e;word-break:break-all;}}

  /* ── Tables ── */
  table{{width:100%;border-collapse:collapse;font-size:.9em;}}
  th{{background:#0056b3;color:#fff;padding:9px 10px;text-align:left;font-weight:600;font-size:.82em;text-transform:uppercase;letter-spacing:.04em;}}
  td{{padding:8px 10px;border-bottom:1px solid #edf0f5;vertical-align:top;}}
  tr:last-child td{{border-bottom:none;}}
  .action-note{{font-size:.82em;color:#8a6200;margin-top:4px;background:#fffbea;padding:3px 6px;border-radius:3px;border-left:3px solid #ffc107;}}
  code{{background:#f0f4ff;color:#0d3c7a;padding:2px 6px;border-radius:3px;font-size:.88em;}}

  /* ── Footer ── */
  .report-footer{{background:#1a1a2e;color:#aab;padding:14px 36px;text-align:center;font-size:.8em;}}

  @media print{{
    body{{background:#fff;}}
    .page{{box-shadow:none;}}
    .section{{page-break-inside:avoid;}}
  }}
</style>
</head>
<body>
<div class="page">

<!-- HEADER -->
<div class="report-header">
  <div class="brand">REBIZ Sentinel X &mdash; AI DVR/NVR Auto-Configuration</div>
  <h1>Configuration Report</h1>
  <div class="meta">
    📍 {_e(report['device_name'])} &nbsp;·&nbsp;
    🌐 {_e(report['ip_address'])} &nbsp;·&nbsp;
    👤 {_e(report['technician'])} &nbsp;·&nbsp;
    📅 {_e(report['configured_at'])} &nbsp;·&nbsp;
    ⏱ {_e(report['total_duration'])}
  </div>
  <div class="overall-banner">
    <div class="overall-icon">{hdr_icon}</div>
    <div class="overall-text">
      <div class="title">Overall: {_e(overall)}</div>
      <div class="sub">{_e(hdr_lbl)}</div>
    </div>
  </div>
</div>

<!-- SCORE CARDS -->
<div class="score-bar">
  <div class="score-card pass"><div class="num">{vs['passed']}</div><div class="lbl">✅ Passed</div></div>
  <div class="score-card warn"><div class="num">{vs['warned']}</div><div class="lbl">⚠️ Warnings</div></div>
  <div class="score-card fail"><div class="num">{vs['failed']}</div><div class="lbl">❌ Failed</div></div>
  <div class="score-card total"><div class="num">{vs['passed']+vs['warned']+vs['failed']}</div><div class="lbl">Total Checks</div></div>
</div>

<!-- DEVICE INFO -->
<div class="section">
  <div class="section-title">Device Information</div>
  <div class="info-grid">
    <div class="info-cell"><label>Device Name</label><span>{_e(report['device_name'])}</span></div>
    <div class="info-cell"><label>IP Address</label><span>{_e(report['ip_address'])}</span></div>
    <div class="info-cell"><label>Manufacturer</label><span>{_e(report['manufacturer'])}</span></div>
    <div class="info-cell"><label>Model</label><span>{_e(report['model'])}</span></div>
    <div class="info-cell"><label>Device Type</label><span>{_e(report['device_type'])}</span></div>
    <div class="info-cell"><label>Serial Number</label><span>{_e(report['serial_number'])}</span></div>
    <div class="info-cell"><label>Firmware Version</label><span>{_e(report['firmware_version'])}</span></div>
    <div class="info-cell"><label>MAC Address</label><span>{_e(report['mac_address'])}</span></div>
    <div class="info-cell"><label>Total / Active Channels</label><span>{_e(report['total_channels'])} / {_e(report['active_channels'])}</span></div>
    <div class="info-cell"><label>Location</label><span>{_e(report['city'])}, {_e(report['state'])}</span></div>
    <div class="info-cell"><label>Timezone</label><span>{_e(report['timezone'])}</span></div>
    <div class="info-cell"><label>NTP Server</label><span>{_e(report['ntp_server'])}</span></div>
  </div>
</div>

<!-- VALIDATION RESULTS -->
<div class="section">
  <div class="section-title">Configuration Validation</div>
  <table>
    <tr><th style="width:36px;">Status</th><th style="width:240px;">Check</th><th>Detail / Action</th></tr>
    {checks_rows}
  </table>
</div>

<!-- STREAM CONFIGURATION -->
<div class="section">
  <div class="section-title">Stream Configuration</div>
  <table>
    <tr>
      <th>Ch</th>
      <th>Main Status</th><th>Main Resolution</th><th>Main Bitrate</th>
      <th>Sub Status</th><th>Sub Resolution</th><th>Sub Bitrate</th>
      <th>Notes</th>
    </tr>
    {stream_rows}
  </table>
</div>

<!-- HDD -->
<div class="section">
  <div class="section-title">Storage — HDD Information</div>
  <table>
    <tr><th>ID</th><th>Name</th><th>Capacity</th><th>Status</th><th>Free Space</th><th>Property</th></tr>
    {hdds_rows}
  </table>
</div>

<!-- USERS -->
<div class="section">
  <div class="section-title">User Accounts</div>
  <table>
    <tr><th>Username</th><th>Password</th><th>Role</th><th>Status</th></tr>
    {users_rows if users_rows else '<tr><td colspan="4" style="color:#999;text-align:center;">No user data available</td></tr>'}
  </table>
</div>

<!-- FOOTER -->
<div class="report-footer">
  REBIZ Sentinel X &mdash; AI DVR/NVR Auto-Configuration Module &nbsp;|&nbsp; Report generated {_e(report['configured_at'])}
</div>

</div>
</body></html>"""


def _build_excel_report(report):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    # Sheet 1: Device Info
    ws1 = wb.active
    ws1.title = "Device Info"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0056B3")

    ws1.append(["Field", "Value"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill

    fields = [
        ("Device Name", report['device_name']),
        ("IP Address", report['ip_address']),
        ("Manufacturer", report['manufacturer']),
        ("Model", report['model']),
        ("Device Type", report['device_type']),
        ("Serial Number", report['serial_number']),
        ("Firmware", report['firmware_version']),
        ("MAC Address", report['mac_address']),
        ("Total Channels", str(report['total_channels'])),
        ("Active Channels", str(report['active_channels'])),
        ("Location", f"{report['city']}, {report['state']}"),
        ("Timezone", report['timezone']),
        ("NTP Server", report['ntp_server']),
        ("Technician", report['technician']),
        ("Configured At", report['configured_at']),
        ("Duration", report['total_duration']),
    ]
    for row in fields:
        ws1.append(row)
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 40

    # Sheet 2: Stream Channels
    ws_streams = wb.create_sheet("Stream Channels")
    ws_streams.append(["Channel", "Main Status", "Main Resolution", "Main Bitrate (kbps)",
                        "Sub Status", "Sub Resolution", "Sub Bitrate (kbps)", "Notes"])
    for cell in ws_streams[1]:
        cell.font = header_font
        cell.fill = header_fill
    stream_status_fill = {
        'configured':          PatternFill("solid", fgColor="C6EFCE"),
        'configured_fallback': PatternFill("solid", fgColor="FFEB9C"),
        'error':               PatternFill("solid", fgColor="FFC7CE"),
    }
    for sc in report.get('stream_channels', []):
        main_st  = sc.get('main_stream', sc.get('main', 'N/A'))
        sub_st   = sc.get('sub_stream',  sc.get('sub',  'N/A'))
        notes    = '; '.join(filter(None, [sc.get('main_note', ''), sc.get('sub_note', '')]))
        ws_streams.append([
            sc.get('channel', '?'),
            main_st,
            sc.get('main_resolution', 'N/A'),
            sc.get('main_bitrate', 'N/A'),
            sub_st,
            sc.get('sub_resolution', 'N/A'),
            sc.get('sub_bitrate', 'N/A'),
            notes,
        ])
        row_fill = stream_status_fill.get(main_st) or stream_status_fill.get(sub_st)
        if row_fill:
            for cell in ws_streams[ws_streams.max_row]:
                cell.fill = row_fill
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_streams.column_dimensions[col].width = 24

    # Sheet 3: Validation
    ws2 = wb.create_sheet("Validation")
    ws2.append(["Status", "Check", "Detail", "Corrective Action"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for c in report['validation']:
        icon = {'pass': '✅', 'warn': '⚠', 'fail': '❌'}.get(c['status'], '?')
        ws2.append([f"{icon} {c['status'].upper()}", c['check'], c['detail'], c.get('action', '')])
    for col in ['A', 'B', 'C', 'D']:
        ws2.column_dimensions[col].width = 28

    # Sheet 3: Users
    ws3 = wb.create_sheet("User Accounts")
    ws3.append(["Username", "Password", "Role", "Status", "Permissions"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for u in report['users']:
        ws3.append([u['username'], u.get('password', 'N/A'), u.get('role', 'N/A'),
                     u.get('status', 'N/A'), u.get('permissions', '')])
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws3.column_dimensions[col].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@dvr_config_bp.route('/api/dvr/report/<fmt>', methods=['POST'])
@login_required
def generate_report(fmt):
    """Generate report in pdf / excel / json format."""
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    report = _build_report_data(sess)
    device_name = report['device_name'].replace(' ', '_')
    ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    filename = f"DVR_Config_{device_name}_{ts}"

    if fmt == 'json':
        buf = io.BytesIO(json.dumps(report, indent=2).encode())
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f"{filename}.json",
                         mimetype='application/json')

    elif fmt == 'excel':
        buf = _build_excel_report(report)
        if not buf:
            return jsonify({'success': False, 'error': 'openpyxl not available'}), 500
        return send_file(buf, as_attachment=True,
                         download_name=f"{filename}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif fmt == 'pdf':
        html = _build_html_report(report)
        try:
            import weasyprint
            pdf_buf = io.BytesIO()
            weasyprint.HTML(string=html).write_pdf(pdf_buf)
            pdf_buf.seek(0)
            return send_file(pdf_buf, as_attachment=True,
                             download_name=f"{filename}.pdf",
                             mimetype='application/pdf')
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            # Fallback: send HTML
            return Response(html, mimetype='text/html',
                            headers={'Content-Disposition': f'attachment; filename="{filename}.html"'})

    return jsonify({'success': False, 'error': f'Unknown format: {fmt}'}), 400


@dvr_config_bp.route('/api/dvr/email-report', methods=['POST'])
@login_required
def email_report():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    recipients = data.get('recipients', [])
    formats = data.get('formats', ['pdf'])
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404
    if not recipients:
        return jsonify({'success': False, 'error': 'No recipients specified'}), 400

    report = _build_report_data(sess)
    device_name = report['device_name']
    ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    filename_base = f"DVR_Config_{device_name.replace(' ', '_')}_{ts}"
    overall = report['validation_summary']['overall']

    msg = MIMEMultipart('mixed')
    msg['Subject'] = f"DVR/NVR Configuration Report — {device_name} [{overall}]"
    msg['From'] = Config.SMTP_USERNAME
    msg['To'] = ', '.join(recipients)

    # HTML body
    vs = report['validation_summary']
    body_html = f"""<html><body style="font-family:Arial,sans-serif;">
    <h2 style="color:#0056b3;">DVR/NVR Configuration Complete</h2>
    <p><strong>Device:</strong> {device_name} ({report['ip_address']})</p>
    <p><strong>Model:</strong> {report['manufacturer']} {report['model']}</p>
    <p><strong>Result:</strong> <strong style="color:{'green' if overall=='PASS' else ('orange' if overall=='WARNING' else 'red')}">{overall}</strong></p>
    <p><strong>Validation:</strong> ✅ {vs['passed']} Passed &nbsp; ⚠ {vs['warned']} Warnings &nbsp; ❌ {vs['failed']} Failed</p>
    <p><strong>Technician:</strong> {report['technician']}</p>
    <p><strong>Configured At:</strong> {report['configured_at']}</p>
    <p><strong>Duration:</strong> {report['total_duration']}</p>
    <hr>
    <p style="color:#999;font-size:0.85em;">REBIZ Sentinel X — AI DVR/NVR Auto-Configuration Module</p>
    </body></html>"""
    msg.attach(MIMEText(body_html, 'html'))

    # Attachments
    if 'pdf' in formats:
        try:
            html = _build_html_report(report)
            import weasyprint
            pdf_bytes = io.BytesIO()
            weasyprint.HTML(string=html).write_pdf(pdf_bytes)
            part = MIMEApplication(pdf_bytes.getvalue(), Name=f"{filename_base}.pdf")
            part['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
            msg.attach(part)
        except Exception as e:
            logger.warning(f"PDF attachment failed: {e}")

    if 'excel' in formats:
        try:
            buf = _build_excel_report(report)
            if buf:
                part = MIMEApplication(buf.read(), Name=f"{filename_base}.xlsx")
                part['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
                msg.attach(part)
        except Exception as e:
            logger.warning(f"Excel attachment failed: {e}")

    if 'json' in formats:
        json_bytes = json.dumps(report, indent=2).encode()
        part = MIMEApplication(json_bytes, Name=f"{filename_base}.json")
        part['Content-Disposition'] = f'attachment; filename="{filename_base}.json"'
        msg.attach(part)

    try:
        with smtplib.SMTP(Config.SMTP_SERVER, Config.SMTP_PORT) as smtp:
            if Config.SMTP_USE_TLS:
                smtp.starttls()
            smtp.login(Config.SMTP_USERNAME, Config.SMTP_PASSWORD)
            smtp.sendmail(Config.SMTP_USERNAME, recipients, msg.as_string())
        return jsonify({'success': True, 'message': f'Report sent to {", ".join(recipients)}'})
    except Exception as e:
        logger.error(f"Email send failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'Email failed: {str(e)}'}), 500


# ─────────────────────────────────────────────
# Session History
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/sessions', methods=['GET'])
@login_required
def list_sessions():
    sessions = DVRConfigSession.query.order_by(DVRConfigSession.updated_at.desc()).limit(20).all()
    result = []
    for s in sessions:
        d = s.get_data()
        conn   = d.get('connection', {})
        val    = d.get('steps', {}).get('validation', {})
        checks = val.get('checks', [])
        result.append({
            'id':          s.id,
            'device_name': conn.get('device_name', 'N/A'),
            'ip':          conn.get('ip', 'N/A'),
            'overall':     val.get('overall', 'Incomplete'),
            'passed':      sum(1 for c in checks if c.get('status') == 'pass'),
            'warned':      sum(1 for c in checks if c.get('status') == 'warn'),
            'failed':      sum(1 for c in checks if c.get('status') == 'fail'),
            'technician':  d.get('technician', 'N/A'),
            'started_at':  d.get('started_at', ''),
            'updated_at':  s.updated_at.isoformat() if s.updated_at else '',
        })
    return jsonify({'sessions': result})


@dvr_config_bp.route('/api/dvr/session/<session_id>', methods=['GET'])
@login_required
def get_session(session_id):
    sess = _get_session_data(session_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404
    return jsonify({'success': True, 'session': sess})


# ─────────────────────────────────────────────
# Camera Name Management (standalone feature)
# ─────────────────────────────────────────────

@dvr_config_bp.route('/camera-names')
@login_required
def camera_names():
    return render_template('camera_names.html')


@dvr_config_bp.route('/api/dvr/camera-names/connect', methods=['POST'])
@login_required
def camera_names_connect():
    data = request.get_json() or {}
    required = ['ip', 'port', 'username', 'password']
    for f in required:
        if not data.get(f):
            return jsonify({'success': False, 'error': f'Missing field: {f}'}), 400

    try:
        client = _make_client(data)
        ok = client.test_connection()
        if not ok:
            return jsonify({'success': False,
                            'error': 'Authentication failed. Check IP, port, username and password.'}), 401

        channels = client.list_active_channels_for_naming()

        sess_id = f"camname_{current_user.id}_{int(time.time())}"
        sess_data = {
            'session_id': sess_id,
            'technician': current_user.username,
            'started_at': datetime.utcnow().isoformat(),
            'connection': {
                'ip': data['ip'],
                'port': data['port'],
                'username': data['username'],
                'password': data['password'],
                'use_https': data.get('use_https', False),
            },
            'steps': {'channels_detected': channels},
        }
        _save_session_data(sess_id, sess_data)

        return jsonify({
            'success': True,
            'session_id': sess_id,
            'channels': channels,
            'skipped_note': 'Offline or disabled channels are automatically excluded from this list.',
        })
    except Exception as e:
        logger.error(f"Camera name management connect error: {e}", exc_info=True)
        return jsonify({'success': False,
                        'error': f'Connection failed: {str(e)}. Verify the device is reachable.'}), 500


@dvr_config_bp.route('/api/dvr/camera-names/apply', methods=['POST'])
@login_required
def camera_names_apply():
    """
    Apply a naming template to all currently-active channels.
    Body: {session_id, template, assignments?: {channel_id: name}}
    `template` supports the placeholder {n} for a 1-based sequence number,
    e.g. "Camera {n}", "Front Lobby {n}". If `assignments` is provided it
    takes precedence per-channel (manual override of the template result).
    """
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    template = (data.get('template') or 'Camera {n}').strip()
    assignments = data.get('assignments') or {}

    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    channels = sess.get('steps', {}).get('channels_detected', [])
    if not channels:
        return jsonify({'success': False, 'error': 'No active channels detected for this session'}), 400

    try:
        client = _make_client(sess['connection'])
        results = []
        for i, ch in enumerate(channels, start=1):
            ch_id = ch['channel_id']
            new_name = assignments.get(ch_id) or template.replace('{n}', str(i).zfill(2))
            ok, detail = client.apply_channel_name(ch_id, new_name)
            results.append({
                'channel_id': ch_id,
                'previous_name': ch.get('name', ''),
                'new_name': new_name,
                'status': 'success' if ok else 'failed',
                'detail': detail,
            })

        sess['steps']['naming_report'] = {
            'template': template,
            'applied_at': datetime.utcnow().isoformat(),
            'results': results,
        }
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'results': results})
    except Exception as e:
        logger.error(f"Camera name management apply error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@dvr_config_bp.route('/api/dvr/camera-names/report/<session_id>', methods=['GET'])
@login_required
def camera_names_report(session_id):
    sess = _get_session_data(session_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404
    report = sess.get('steps', {}).get('naming_report')
    if not report:
        return jsonify({'success': False, 'error': 'No naming report available yet for this session'}), 404

    total = len(report['results'])
    succeeded = sum(1 for r in report['results'] if r['status'] == 'success')
    return jsonify({
        'success': True,
        'session_id': session_id,
        'connection_ip': sess.get('connection', {}).get('ip'),
        'template': report['template'],
        'applied_at': report['applied_at'],
        'total_channels': total,
        'succeeded': succeeded,
        'failed': total - succeeded,
        'results': report['results'],
    })


# ─────────────────────────────────────────────────────────────────────────────
# DVR Diagnostic Tool  — /dvr/diagnostic
# Lets you test user-creation against a device directly from the browser.
# Tries every known approach and shows you the raw device XML responses.
# ─────────────────────────────────────────────────────────────────────────────

@dvr_config_bp.route('/dvr/diagnostic', methods=['GET', 'POST'])
@login_required
def dvr_diagnostic():
    """Browser-based diagnostic: test user creation against any DVR/NVR."""
    if request.method == 'GET':
        return render_template('dvr_diagnostic.html')

    # ── Collect form inputs ──────────────────────────────────────────────────
    ip       = (request.form.get('ip') or '').strip()
    port     = int(request.form.get('port') or 80)
    username = (request.form.get('username') or '').strip()
    password = (request.form.get('password') or '').strip()

    if not ip or not username or not password:
        return render_template('dvr_diagnostic.html',
                               error='IP, username and password are required.')

    base     = f"http://{ip}:{port}/ISAPI"
    auth     = HTTPDigestAuth(username, password)
    sess_req = requests.Session()
    headers  = {"Content-Type": "application/xml"}
    timeout  = 15

    TEST_USER       = "_dvr_diag_"
    TEST_PASS_PLAIN = "Test@12345"
    TEST_PASS_MD5   = hashlib.md5(TEST_PASS_PLAIN.encode()).hexdigest()

    steps = []   # list of dicts: {label, sent, status, response, success}

    def _do_get(path):
        url = base + path
        try:
            r = sess_req.get(url, auth=auth, verify=False, timeout=timeout)
            return r.status_code, r.text
        except Exception as exc:
            return None, str(exc)

    def _do_post(path, body):
        url = base + path
        try:
            r = sess_req.post(url, auth=auth, verify=False, timeout=timeout,
                              data=body, headers=headers)
            return r.status_code, r.text
        except Exception as exc:
            return None, str(exc)

    def _do_put(path, body):
        url = base + path
        try:
            r = sess_req.put(url, auth=auth, verify=False, timeout=timeout,
                             data=body, headers=headers)
            return r.status_code, r.text
        except Exception as exc:
            return None, str(exc)

    def _user_exists(raw_xml, uname):
        return f"<userName>{uname}</userName>" in raw_xml or \
               f"<userName xmlns" in raw_xml and uname in raw_xml

    def _find_user_id(raw_xml, uname):
        """Return the id text node that appears just before userName=uname."""
        pattern = r'<id>(\d+)</id>\s*<userName>' + re.escape(uname) + r'</userName>'
        m = re.search(pattern, raw_xml)
        return m.group(1) if m else None

    def _cleanup(uid):
        if uid:
            _do_delete(f"/Security/users/{uid}")

    def _do_delete(path):
        url = base + path
        try:
            sess_req.delete(url, auth=auth, verify=False, timeout=timeout)
        except Exception:
            pass

    # ── Step 1: GET /Security/users ──────────────────────────────────────────
    st1, resp1 = _do_get("/Security/users")
    steps.append({
        'label': 'GET /ISAPI/Security/users  (raw device XML)',
        'sent': None,
        'status': st1,
        'response': resp1,
        'success': st1 == 200,
    })
    raw_userlist = resp1 if st1 == 200 else ''

    # Detect namespace
    ns_match = re.search(r'xmlns="([^"]+)"', raw_userlist)
    detected_ns = ns_match.group(1) if ns_match else ''

    # Shared fields used in every user block
    def _bond():
        return """  <bondIpList>
    <bondIp>
      <id>1</id>
      <ipAddress>0.0.0.0</ipAddress>
      <ipv6Address>::</ipv6Address>
    </bondIp>
  </bondIpList>"""

    created_uid = None
    success_method = None

    # ── Step 2: POST — plain password, no xmlns ──────────────────────────────
    if not success_method:
        body = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<User>\n'
            '  <id>0</id>\n'
            f'  <userName>{TEST_USER}</userName>\n'
            f'  <password>{TEST_PASS_PLAIN}</password>\n'
            + _bond() + '\n'
            '  <macAddress></macAddress>\n'
            '  <userLevel>Viewer</userLevel>\n'
            '  <attribute><inherent>false</inherent></attribute>\n'
            '</User>'
        )
        st, resp = _do_post("/Security/users", body)
        ok = st == 200 and 'statusCode' not in resp
        steps.append({
            'label': 'POST /ISAPI/Security/users — plain password, no xmlns',
            'sent': body, 'status': st, 'response': resp, 'success': ok,
        })
        if ok:
            _, fresh = _do_get("/Security/users")
            created_uid = _find_user_id(fresh, TEST_USER)
            success_method = 'POST plain password'

    # ── Step 3: POST — MD5 password, no xmlns ───────────────────────────────
    if not success_method:
        body = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<User>\n'
            '  <id>0</id>\n'
            f'  <userName>{TEST_USER}</userName>\n'
            f'  <password>{TEST_PASS_MD5}</password>\n'
            + _bond() + '\n'
            '  <macAddress></macAddress>\n'
            '  <userLevel>Viewer</userLevel>\n'
            '  <attribute><inherent>false</inherent></attribute>\n'
            '</User>'
        )
        st, resp = _do_post("/Security/users", body)
        ok = st == 200 and 'statusCode' not in resp
        steps.append({
            'label': f'POST /ISAPI/Security/users — MD5 password ({TEST_PASS_MD5})',
            'sent': body, 'status': st, 'response': resp, 'success': ok,
        })
        if ok:
            _, fresh = _do_get("/Security/users")
            created_uid = _find_user_id(fresh, TEST_USER)
            success_method = 'POST MD5 password'

    # ── Step 4: POST — plain password, with xmlns ────────────────────────────
    if not success_method and detected_ns:
        body = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            f'<User xmlns="{detected_ns}">\n'
            '  <id>0</id>\n'
            f'  <userName>{TEST_USER}</userName>\n'
            f'  <password>{TEST_PASS_PLAIN}</password>\n'
            + _bond() + '\n'
            '  <macAddress></macAddress>\n'
            '  <userLevel>Viewer</userLevel>\n'
            '  <attribute><inherent>false</inherent></attribute>\n'
            '</User>'
        )
        st, resp = _do_post("/Security/users", body)
        ok = st == 200 and 'statusCode' not in resp
        steps.append({
            'label': f'POST — plain password, xmlns="{detected_ns}"',
            'sent': body, 'status': st, 'response': resp, 'success': ok,
        })
        if ok:
            _, fresh = _do_get("/Security/users")
            created_uid = _find_user_id(fresh, TEST_USER)
            success_method = f'POST plain+xmlns'

    # ── Step 5: POST — MD5 password, with xmlns ──────────────────────────────
    if not success_method and detected_ns:
        body = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            f'<User xmlns="{detected_ns}">\n'
            '  <id>0</id>\n'
            f'  <userName>{TEST_USER}</userName>\n'
            f'  <password>{TEST_PASS_MD5}</password>\n'
            + _bond() + '\n'
            '  <macAddress></macAddress>\n'
            '  <userLevel>Viewer</userLevel>\n'
            '  <attribute><inherent>false</inherent></attribute>\n'
            '</User>'
        )
        st, resp = _do_post("/Security/users", body)
        ok = st == 200 and 'statusCode' not in resp
        steps.append({
            'label': f'POST — MD5 password, xmlns="{detected_ns}"',
            'sent': body, 'status': st, 'response': resp, 'success': ok,
        })
        if ok:
            _, fresh = _do_get("/Security/users")
            created_uid = _find_user_id(fresh, TEST_USER)
            success_method = 'POST MD5+xmlns'

    # ── Step 6: Clone first <User> from GET, patch, POST ────────────────────
    if not success_method and raw_userlist:
        m = re.search(
            r'(<(?:[^:>\s/]+:)?User(?:\s[^>]*)?>.*?</(?:[^:>\s/]+:)?User>)',
            raw_userlist, re.DOTALL
        )
        if m:
            cloned = m.group(1)
            for tag, val in [('id', '0'), ('userName', TEST_USER),
                              ('password', TEST_PASS_PLAIN),
                              ('userLevel', 'Viewer'), ('inherent', 'false')]:
                cloned = re.sub(
                    rf'(<(?:[^:>\s/]+:)?{re.escape(tag)}(?:\s[^>]*)?>)'
                    rf'.*?(</(?:[^:>\s/]+:)?{re.escape(tag)}>)',
                    rf'\g<1>{val}\g<2>', cloned, flags=re.DOTALL
                )
            body = '<?xml version="1.0" encoding="UTF-8"?>\n' + cloned
            st, resp = _do_post("/Security/users", body)
            ok = st == 200 and 'statusCode' not in resp
            steps.append({
                'label': 'POST — cloned first <User> from device GET, patched fields',
                'sent': body, 'status': st, 'response': resp, 'success': ok,
            })
            if ok:
                _, fresh = _do_get("/Security/users")
                created_uid = _find_user_id(fresh, TEST_USER)
                success_method = 'POST cloned User block'

    # ── Step 7: PUT full UserList + injected user (plain password) ───────────
    if not success_method and raw_userlist:
        new_block = (
            '  <User>\n'
            '    <id>0</id>\n'
            f'    <userName>{TEST_USER}</userName>\n'
            f'    <password>{TEST_PASS_PLAIN}</password>\n'
            '    <bondIpList><bondIp><id>1</id>'
            '<ipAddress>0.0.0.0</ipAddress>'
            '<ipv6Address>::</ipv6Address></bondIp></bondIpList>\n'
            '    <macAddress></macAddress>\n'
            '    <userLevel>Viewer</userLevel>\n'
            '    <attribute><inherent>false</inherent></attribute>\n'
            '  </User>\n'
        )
        modified, n = re.subn(
            r'(</(?:[^:>\s/]+:)?[Uu]ser[Ll]ist>)',
            new_block + r'\1', raw_userlist
        )
        if n:
            st, resp = _do_put("/Security/users", modified)
            ok = st in (200, 204)
            steps.append({
                'label': 'PUT /ISAPI/Security/users — full UserList with new user injected (plain password)',
                'sent': modified[:3000] + ('…(truncated)' if len(modified) > 3000 else ''),
                'status': st, 'response': resp, 'success': ok,
            })
            if ok:
                _, fresh = _do_get("/Security/users")
                created_uid = _find_user_id(fresh, TEST_USER)
                if created_uid:
                    success_method = 'PUT UserList plain password'
                else:
                    steps[-1]['success'] = False
                    steps[-1]['response'] += '\n\n[Note: PUT returned 200 but user not found in refreshed list]'

    # ── Step 8: PUT full UserList + injected user (MD5 password) ─────────────
    if not success_method and raw_userlist:
        new_block_md5 = (
            '  <User>\n'
            '    <id>0</id>\n'
            f'    <userName>{TEST_USER}</userName>\n'
            f'    <password>{TEST_PASS_MD5}</password>\n'
            '    <bondIpList><bondIp><id>1</id>'
            '<ipAddress>0.0.0.0</ipAddress>'
            '<ipv6Address>::</ipv6Address></bondIp></bondIpList>\n'
            '    <macAddress></macAddress>\n'
            '    <userLevel>Viewer</userLevel>\n'
            '    <attribute><inherent>false</inherent></attribute>\n'
            '  </User>\n'
        )
        modified_md5, n = re.subn(
            r'(</(?:[^:>\s/]+:)?[Uu]ser[Ll]ist>)',
            new_block_md5 + r'\1', raw_userlist
        )
        if n:
            st, resp = _do_put("/Security/users", modified_md5)
            ok = st in (200, 204)
            steps.append({
                'label': 'PUT /ISAPI/Security/users — full UserList with new user injected (MD5 password)',
                'sent': modified_md5[:3000] + ('…(truncated)' if len(modified_md5) > 3000 else ''),
                'status': st, 'response': resp, 'success': ok,
            })
            if ok:
                _, fresh = _do_get("/Security/users")
                created_uid = _find_user_id(fresh, TEST_USER)
                if created_uid:
                    success_method = 'PUT UserList MD5 password'
                else:
                    steps[-1]['success'] = False
                    steps[-1]['response'] += '\n\n[Note: PUT returned 200 but user not found in refreshed list]'

    # ── Cleanup test user ────────────────────────────────────────────────────
    if created_uid:
        _cleanup(created_uid)

    return render_template(
        'dvr_diagnostic.html',
        ran=True,
        ip=ip, port=port, username=username,
        steps=steps,
        success_method=success_method,
        detected_ns=detected_ns,
    )
