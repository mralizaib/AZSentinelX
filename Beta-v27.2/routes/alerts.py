from flask import Blueprint, render_template, request, jsonify, flash, redirect, url_for
from flask_login import login_required, current_user
import logging
from datetime import datetime, timedelta
import json

logger = logging.getLogger(__name__)

alerts_bp = Blueprint('alerts', __name__)
from routes.permissions import make_blueprint_permission_check
alerts_bp.before_request(make_blueprint_permission_check('alerts'))

@alerts_bp.route('/alerts')
@login_required
def index():
    # Get user's alert configurations with error handling
    alerts = []
    try:
        from models import AlertConfig
        alerts = AlertConfig.query.filter_by(user_id=current_user.id).all()
    except Exception as e:
        logger.error(f"Error loading alert configurations: {str(e)}")
        flash('Unable to load alert configurations. Database may need migration.', 'warning')
    # Pass rule_id context from URL so the template can show a filter banner
    rule_id = request.args.get('rule_id', '')
    rule_desc = request.args.get('rule_desc', '')
    return render_template('alerts.html', alerts=alerts, rule_id=rule_id, rule_desc=rule_desc)

@alerts_bp.route('/api/alerts', methods=['GET'])
@login_required
def get_alerts():
    """Get security alerts from OpenSearch"""
    try:
        from opensearch_api import OpenSearchAPI
        opensearch = OpenSearchAPI()

        # Get query parameters
        severity_levels = request.args.getlist('severity_levels[]')
        
        # If no array format, try single severity parameter (for dashboard redirects)
        if not severity_levels:
            single_severity = request.args.get('severity')
            if single_severity == 'all':
                severity_levels = ['critical', 'high', 'medium', 'low', 'none']
            elif single_severity:
                severity_levels = [single_severity.lower()]

        if not severity_levels:
            severity_levels = ['critical', 'high']  # Default levels

        # Time range
        time_range = request.args.get('time_range', '24h')
        end_time = datetime.utcnow().isoformat()

        if time_range == '1h':
            start_time = (datetime.utcnow() - timedelta(hours=1)).isoformat()
        elif time_range == '6h':
            start_time = (datetime.utcnow() - timedelta(hours=6)).isoformat()
        elif time_range == '24h':
            start_time = (datetime.utcnow() - timedelta(days=1)).isoformat()
        elif time_range == '7d':
            start_time = (datetime.utcnow() - timedelta(days=7)).isoformat()
        elif time_range == '12d':
            start_time = (datetime.utcnow() - timedelta(days=12)).isoformat()
        elif time_range == '30d':
            start_time = (datetime.utcnow() - timedelta(days=30)).isoformat()
        elif time_range == '60d':
            start_time = (datetime.utcnow() - timedelta(days=60)).isoformat()
        elif time_range == '90d':
            start_time = (datetime.utcnow() - timedelta(days=90)).isoformat()
        else:
            # Custom time range
            start_time = request.args.get('start_time')
            end_time = request.args.get('end_time', end_time)

        # Pagination
        limit = int(request.args.get('limit', 100))
        offset = int(request.args.get('offset', 0))

        # Other filters
        sort_field = request.args.get('sort_field', '@timestamp')
        sort_order = request.args.get('sort_order', 'desc')

        # Additional filters
        additional_filters = {}
        search_query = request.args.get('search_query')
        if search_query:
            additional_filters['search_query'] = search_query  # Use search_query key for multi-field search

        rule_id = request.args.get('rule_id')
        if rule_id:
            additional_filters['rule.id'] = rule_id

        # FIM alerts filter
        fim_alerts = request.args.get('fim_alerts')
        if fim_alerts == 'true':
            additional_filters['rule.id'] = ['553', '554']

        # Search for alerts
        results = opensearch.search_alerts(
            severity_levels=severity_levels,
            start_time=start_time,
            end_time=end_time,
            limit=limit,
            offset=offset,
            sort_field=sort_field,
            sort_order=sort_order,
            additional_filters=additional_filters
        )

        # When a text search returns 0 results due to a restrictive severity filter,
        # automatically fall back to all severity levels so the user still gets matches.
        ALL_LEVELS = ['critical', 'high', 'medium', 'low', 'fim', 'events']
        if (
            search_query
            and results.get('total', 0) == 0
            and set(severity_levels) != set(ALL_LEVELS)
        ):
            fallback = opensearch.search_alerts(
                severity_levels=ALL_LEVELS,
                start_time=start_time,
                end_time=end_time,
                limit=limit,
                offset=offset,
                sort_field=sort_field,
                sort_order=sort_order,
                additional_filters=additional_filters
            )
            if fallback.get('total', 0) > 0:
                fallback['severity_relaxed'] = True
                fallback['original_severity_levels'] = severity_levels
                return jsonify(fallback)

        return jsonify(results)
    except Exception as e:
        logger.error(f"Error getting alerts: {str(e)}")
        return jsonify({'error': str(e)}), 500

@alerts_bp.route('/alerts/view/<string:alert_id>')
@login_required
def view_alert(alert_id):
    """View alert details page"""
    try:
        from opensearch_api import OpenSearchAPI
        index = request.args.get('index')
        opensearch = OpenSearchAPI()
        
        result = opensearch.get_alert_by_id(alert_id, index)
        
        if not result or 'error' in result:
            error_msg = result.get('error') if result else "Alert not found"
            logger.error(f"Error loading alert: {error_msg}")
            flash(f"Error loading alert: {error_msg}", "danger")
            return redirect(url_for('alerts.index'))
            
        # If the result is from search, it might be the hit object itself
        # opensearch_api.py returns results differently based on if index is provided
        alert_data = result
        if '_source' in result:
            # It's a raw hit from OpenSearch
            alert_data = {
                'id': result.get('_id'),
                'index': result.get('_index'),
                'source': result.get('_source')
            }
            
        # ── Context-specific enrichment ───────────────────────────────────────
        vuln_info    = {}
        defender_info = {}

        source  = alert_data.get('source', {}) or {}
        data    = source.get('data', {}) or {}
        rule    = source.get('rule', {}) or {}
        groups  = rule.get('groups', []) or []
        desc    = rule.get('description', '').lower()
        groups_str = ' '.join(groups).lower()

        # ── 1. Vulnerability alert ─────────────────────────────────────────────
        vuln_data = data.get('vulnerability', {}) or {}
        if vuln_data or 'vulnerability-detector' in groups_str or 'vulnerability' in groups_str:
            pkg      = vuln_data.get('package', {}) or {}
            pkg_name = pkg.get('name', '').strip()
            agent_id = (source.get('agent', {}) or {}).get('id', '').strip()
            cvss3    = (vuln_data.get('cvss', {}) or {}).get('cvss3', {}) or {}
            vector   = cvss3.get('vector', {}) or {}

            # Installation path is fetched asynchronously by the frontend
            # via /api/alerts/syscollector-lookup to avoid blocking page load.
            install_path        = None
            install_time        = None
            syscollector_vendor = None

            raw_status   = vuln_data.get('status', '').strip()
            solved_aliases = {'solved', 'fixed', 'not present', 'patched'}
            status_resolved = raw_status.lower() in solved_aliases

            vuln_info = {
                'cve':          vuln_data.get('cve', ''),
                'severity':     vuln_data.get('severity', ''),
                'score':        (vuln_data.get('score', {}) or {}).get('base', ''),
                'cvss_version': (vuln_data.get('score', {}) or {}).get('version', ''),
                'status':       raw_status or 'Unknown',
                'status_resolved': status_resolved,
                'package_name':    pkg_name,
                'package_version': pkg.get('version', ''),
                'package_arch':    pkg.get('architecture', '').strip(),
                'condition':       pkg.get('condition', ''),
                'cwe':             vuln_data.get('cwe_reference', ''),
                'rationale':       vuln_data.get('rationale', ''),
                'reference':       vuln_data.get('reference', ''),
                'install_path':    install_path,
                'install_time':    install_time,
                'vendor':          syscollector_vendor,
                'assigner':        vuln_data.get('assigner', ''),
                'published':       vuln_data.get('published', ''),
                'attack_vector':   vector.get('attack_vector', ''),
                'privileges_required': vector.get('privileges_required', ''),
                'user_interaction':    vector.get('user_interaction', ''),
            }

        # ── 2. Defender alert ──────────────────────────────────────────────────
        defender_keywords = ('defender', 'windows defender', 'microsoft defender',
                             'malware', 'antivirus', 'anti-virus')
        if any(kw in desc or kw in groups_str for kw in defender_keywords):
            win       = data.get('win', {}) or {}
            eventdata = win.get('eventdata', {}) or {}
            sysdata   = win.get('system',    {}) or {}

            # Field names vary by Wazuh version / decoder — try several aliases
            def _pick(*keys):
                for k in keys:
                    v = eventdata.get(k, '').strip()
                    if v:
                        return v
                return ''

            threat_name  = _pick('threat name', 'threatName', 'Threat Name',
                                 'name', 'detection', 'Detection')
            action       = _pick('action name', 'actionName', 'Action Name',
                                 'action', 'Action', 'result', 'Result')
            file_path    = _pick('path', 'Path', 'process', 'Process',
                                 'processName', 'filePath', 'file')
            sev_name     = _pick('severity name', 'severityName', 'Severity Name',
                                 'severity', 'Severity')
            det_user     = _pick('detection user', 'detectionUser', 'Detection User',
                                 'user', 'User', 'processUser')
            product_name = _pick('product name', 'productName', 'Product Name',
                                 'product', 'Product')

            # Determine resolution status
            remediated_kw = ('quarantine', 'quarantined', 'removed', 'cleaned',
                             'deleted', 'blocked', 'suspended', 'remediated')
            is_resolved = any(kw in action.lower() or kw in desc
                              for kw in remediated_kw)

            # Additional Defender action fields
            action_time = _pick('actionTime', 'action time', 'Action Time',
                                'remediationTime', 'Remediation Time',
                                'detectionTime', 'timeCreated')
            executed_by = _pick('executedBy', 'Executed By', 'provider',
                                'providerName', 'engineVersion')
            if not executed_by and product_name:
                executed_by = product_name
            elif not executed_by:
                executed_by = 'Microsoft Defender'
            execution_result = _pick('executionResult', 'Execution Result',
                                     'statusCode', 'errorDescription',
                                     'result', 'Result', 'outcome')
            reason = _pick('reason', 'Reason', 'noActionReason',
                           'No Action Reason', 'errorCode', 'error')
            category_name = _pick('categoryName', 'Category Name',
                                  'category', 'threatType')

            defender_info = {
                'threat_name':      threat_name  or rule.get('description', ''),
                'action':           action,
                'action_taken':     bool(action and action.lower() not in ('', 'unknown', 'none')),
                'action_time':      action_time,
                'executed_by':      executed_by,
                'execution_result': execution_result,
                'reason':           reason if not action else '',
                'file_path':        file_path,
                'severity_name':    sev_name,
                'detection_user':   det_user,
                'product_name':     product_name or 'Microsoft Defender',
                'event_id':         sysdata.get('eventID', ''),
                'category_name':    category_name,
                'status':           'Resolved' if is_resolved else 'Active',
                'status_resolved':  is_resolved,
                'raw_eventdata':    eventdata,
            }

        return render_template('alert_details.html',
                               alert=alert_data,
                               vuln_info=vuln_info,
                               defender_info=defender_info)
    except Exception as e:
        logger.error(f"Error loading alert details view: {str(e)}", exc_info=True)
        flash(f"Internal server error while loading alert details: {str(e)}", "danger")
        return redirect(url_for('alerts.index'))

@alerts_bp.route('/api/alerts/syscollector-lookup', methods=['GET'])
@login_required
def syscollector_lookup():
    """
    Async endpoint: look up an installed package on a specific Wazuh agent
    via the syscollector API and return its installation path.
    Called by the alert detail page after it has already rendered, so the
    blocking Wazuh API call does not delay the initial page load.

    Query params:
        agent_id   - Wazuh agent ID (e.g. "571")
        pkg_name   - package name to search for (e.g. "Mozilla Firefox (x64 en-US)")
    """
    agent_id = request.args.get('agent_id', '').strip()
    pkg_name = request.args.get('pkg_name', '').strip()

    if not agent_id or not pkg_name:
        return jsonify({'error': 'agent_id and pkg_name are required'}), 400

    try:
        from wazuh_api import WazuhAPI
        wazuh    = WazuhAPI()
        packages = wazuh.get_agent_packages(agent_id, search=pkg_name, limit=50)

        install_path = None
        install_time = None
        vendor       = None
        version      = None

        for p in packages:
            if p.get('location') or p.get('install_time'):
                install_path = p.get('location') or install_path
                install_time = p.get('install_time') or install_time
                vendor       = p.get('vendor') or vendor
                version      = p.get('version') or version
                if install_path:
                    break

        return jsonify({
            'install_path': install_path,
            'install_time': install_time,
            'vendor':       vendor,
            'version':      version,
            'found':        install_path is not None,
        })
    except Exception as e:
        logger.warning(f"Syscollector lookup error (agent={agent_id}, pkg={pkg_name}): {e}")
        return jsonify({'error': str(e)}), 500


@alerts_bp.route('/api/alert_configs', methods=['GET'])
@login_required
def get_alert_configs():
    """Get all alert configurations for the current user"""
    try:
        from models import AlertConfig
        alerts = AlertConfig.query.filter_by(user_id=current_user.id).all()

        alerts_list = []
        for alert in alerts:
            try:
                alert_data = {
                    'id': alert.id,
                    'name': alert.name,
                    'alert_levels': alert.get_alert_levels(),
                    'email_recipient': alert.email_recipient,
                    'notify_time': alert.notify_time,
                    'enabled': alert.enabled,
                    'created_at': alert.created_at.isoformat(),
                    'alert_type': getattr(alert, 'alert_type', 'standard') or 'standard',
                }

                # Add include_fields if available
                try:
                    if hasattr(alert, 'get_include_fields') and callable(getattr(alert, 'get_include_fields')):
                        alert_data['include_fields'] = alert.get_include_fields()
                    else:
                        alert_data['include_fields'] = ["@timestamp", "agent.ip", "agent.labels.location.set", "agent.name", "rule.description", "rule.id"]
                except Exception as field_error:
                    logger.warning(f"Error getting include_fields for alert {alert.id}: {str(field_error)}")
                    alert_data['include_fields'] = ["@timestamp", "agent.ip", "agent.labels.location.set", "agent.name", "rule.description", "rule.id"]

                # Add FIM-specific fields
                if alert_data['alert_type'] == 'fim':
                    try:
                        alert_data['fim_agent_names'] = alert.get_fim_agent_names() if hasattr(alert, 'get_fim_agent_names') else []
                        alert_data['fim_paths'] = alert.get_fim_paths() if hasattr(alert, 'get_fim_paths') else []
                        alert_data['fim_file_names'] = alert.get_fim_file_names() if hasattr(alert, 'get_fim_file_names') else []
                        alert_data['fim_file_extensions'] = alert.get_fim_file_extensions() if hasattr(alert, 'get_fim_file_extensions') else []
                    except Exception as fim_err:
                        logger.warning(f"Error getting FIM fields for alert {alert.id}: {fim_err}")
                        alert_data['fim_agent_names'] = []
                        alert_data['fim_paths'] = []
                        alert_data['fim_file_names'] = []
                        alert_data['fim_file_extensions'] = []

                alerts_list.append(alert_data)
            except Exception as alert_error:
                logger.error(f"Error processing alert config {alert.id}: {str(alert_error)}")
                continue

        return jsonify(alerts_list)
    except Exception as e:
        logger.error(f"Error getting alert configs: {str(e)}")
        # Return empty list instead of error to allow page to load
        return jsonify([])

@alerts_bp.route('/api/alert_configs', methods=['POST'])
@login_required
def create_alert_config():
    """Create a new alert configuration"""
    if not current_user.is_admin():
        return jsonify({'error': 'Access denied. Admin privileges required.'}), 403
    try:
        from models import AlertConfig, db
        
        data = request.json

        # Validate required fields
        if not data.get('name'):
            return jsonify({'error': 'Alert name is required'}), 400

        if data.get('alert_type', 'standard') != 'fim' and not data.get('alert_levels'):
            return jsonify({'error': 'At least one alert level must be selected'}), 400

        if not data.get('email_recipient'):
            return jsonify({'error': 'Email recipient is required'}), 400

        # Create new alert configuration
        new_alert = AlertConfig(
            user_id=current_user.id,
            name=data.get('name'),
            email_recipient=data.get('email_recipient'),
            notify_time=data.get('notify_time'),
            enabled=data.get('enabled', True),
            alert_type=data.get('alert_type', 'standard')
        )

        # Set JSON fields
        new_alert.set_alert_levels(data.get('alert_levels'))

        # Set include_fields if provided
        if 'include_fields' in data and hasattr(new_alert, 'set_include_fields'):
            new_alert.set_include_fields(data.get('include_fields'))

        # FIM-specific fields
        if new_alert.alert_type == 'fim':
            if data.get('fim_agent_names'):
                agents = [a.strip() for a in data['fim_agent_names'] if a.strip()] \
                    if isinstance(data['fim_agent_names'], list) \
                    else [a.strip() for a in data['fim_agent_names'].split(',') if a.strip()]
                new_alert.set_fim_agent_names(agents)
            if data.get('fim_paths'):
                paths = [p.strip() for p in data['fim_paths'] if p.strip()] \
                    if isinstance(data['fim_paths'], list) \
                    else [p.strip() for p in data['fim_paths'].split(',') if p.strip()]
                new_alert.set_fim_paths(paths)
            if data.get('fim_file_names'):
                fnames = [f.strip() for f in data['fim_file_names'] if f.strip()] \
                    if isinstance(data['fim_file_names'], list) \
                    else [f.strip() for f in data['fim_file_names'].split(',') if f.strip()]
                new_alert.set_fim_file_names(fnames)
            if data.get('fim_file_extensions'):
                exts = [e.strip() for e in data['fim_file_extensions'] if e.strip()] \
                    if isinstance(data['fim_file_extensions'], list) \
                    else [e.strip() for e in data['fim_file_extensions'].split(',') if e.strip()]
                new_alert.set_fim_file_extensions(exts)

        # Save to database
        db.session.add(new_alert)
        db.session.commit()

        return jsonify({
            'id': new_alert.id,
            'name': new_alert.name,
            'message': 'Alert configuration created successfully'
        }), 201
    except Exception as e:
        if 'db' in locals() and db:
            db.session.rollback()
        logger.error(f"Error creating alert config: {str(e)}")
        return jsonify({'error': str(e)}), 500

@alerts_bp.route('/api/alert_configs/<int:alert_id>', methods=['PUT'])
@login_required
def update_alert_config(alert_id):
    """Update an existing alert configuration"""
    if not current_user.is_admin():
        return jsonify({'error': 'Access denied. Admin privileges required.'}), 403
    try:
        from models import AlertConfig, db
        
        # Validate request data
        if not request.is_json:
            return jsonify({'error': 'Request must be JSON'}), 400
            
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
        
        alert = AlertConfig.query.filter_by(id=alert_id, user_id=current_user.id).first()

        if not alert:
            return jsonify({'error': 'Alert configuration not found'}), 404

        # Validate required fields
        if 'name' in data and not data['name'].strip():
            return jsonify({'error': 'Alert name cannot be empty'}), 400

        # Determine the effective alert type (prefer incoming value, fall back to stored)
        effective_type = data.get('alert_type', alert.alert_type or 'standard')
        if effective_type != 'fim' and 'alert_levels' in data and not data['alert_levels']:
            return jsonify({'error': 'At least one alert level must be selected'}), 400

        if 'email_recipient' in data and not data['email_recipient'].strip():
            return jsonify({'error': 'Email recipient cannot be empty'}), 400

        # Update fields if provided
        if 'name' in data:
            alert.name = data['name'].strip()

        if 'alert_levels' in data:
            alert.set_alert_levels(data['alert_levels'])

        if 'email_recipient' in data:
            alert.email_recipient = data['email_recipient'].strip()

        if 'notify_time' in data:
            alert.notify_time = data['notify_time']

        if 'enabled' in data:
            alert.enabled = data['enabled']

        # Set include_fields if provided
        if 'include_fields' in data and hasattr(alert, 'set_include_fields'):
            alert.set_include_fields(data.get('include_fields'))

        # FIM-specific fields
        if 'alert_type' in data:
            alert.alert_type = data['alert_type']

        if alert.alert_type == 'fim':
            if 'fim_agent_names' in data:
                raw = data['fim_agent_names']
                agents = [a.strip() for a in raw if a.strip()] if isinstance(raw, list) \
                    else [a.strip() for a in raw.split(',') if a.strip()]
                alert.set_fim_agent_names(agents)
            if 'fim_paths' in data:
                raw = data['fim_paths']
                paths = [p.strip() for p in raw if p.strip()] if isinstance(raw, list) \
                    else [p.strip() for p in raw.split(',') if p.strip()]
                alert.set_fim_paths(paths)
            if 'fim_file_names' in data:
                raw = data['fim_file_names']
                fnames = [f.strip() for f in raw if f.strip()] if isinstance(raw, list) \
                    else [f.strip() for f in raw.split(',') if f.strip()]
                alert.set_fim_file_names(fnames)
            if 'fim_file_extensions' in data:
                raw = data['fim_file_extensions']
                exts = [e.strip() for e in raw if e.strip()] if isinstance(raw, list) \
                    else [e.strip() for e in raw.split(',') if e.strip()]
                alert.set_fim_file_extensions(exts)

        # Save changes
        db.session.commit()

        return jsonify({
            'id': alert.id,
            'message': 'Alert configuration updated successfully'
        })
    except Exception as e:
        if 'db' in locals() and db:
            db.session.rollback()
        logger.error(f"Error updating alert config: {str(e)}")
        return jsonify({'error': f'Failed to update alert configuration: {str(e)}'}), 500

@alerts_bp.route('/api/alert_configs/<int:alert_id>', methods=['DELETE'])
@login_required
def delete_alert_config(alert_id):
    """Delete an alert configuration"""
    try:
        from models import AlertConfig, db
        
        alert = AlertConfig.query.filter_by(id=alert_id, user_id=current_user.id).first()

        if not alert:
            return jsonify({'error': 'Alert configuration not found'}), 404

        # Store alert name for response
        alert_name = alert.name
        
        db.session.delete(alert)
        db.session.commit()

        return jsonify({
            'message': f'Alert configuration "{alert_name}" deleted successfully'
        })
    except Exception as e:
        if 'db' in locals() and db:
            db.session.rollback()
        logger.error(f"Error deleting alert config: {str(e)}")
        return jsonify({'error': f'Failed to delete alert configuration: {str(e)}'}), 500

@alerts_bp.route('/api/alerts/<string:alert_id>', methods=['GET'])
@login_required
def get_alert_details(alert_id):
    """Get details for a specific alert by ID"""
    try:
        from opensearch_api import OpenSearchAPI
        
        index = request.args.get('index')
        opensearch = OpenSearchAPI()

        result = opensearch.get_alert_by_id(alert_id, index)

        if 'error' in result:
            logger.error(f"Error getting alert details: {result['error']}")
            return jsonify(result), 404

        return jsonify(result)
    except Exception as e:
        logger.error(f"Error loading alert details: {str(e)}")
        return jsonify({'error': str(e)}), 500

@alerts_bp.route('/api/alert_configs/<int:alert_id>/reset-dedup', methods=['DELETE'])
@login_required
def reset_alert_dedup(alert_id):
    """
    Delete all SentAlert dedup records for this config so the next check cycle
    will resend any events it finds. Useful when the email was lost or the
    config was corrected.
    """
    try:
        from models import AlertConfig, SentAlert
        alert = AlertConfig.query.filter_by(id=alert_id, user_id=current_user.id).first()
        if not alert:
            return jsonify({'error': 'Alert configuration not found'}), 404
        deleted = SentAlert.query.filter_by(alert_config_id=alert_id).delete()
        from models import db
        db.session.commit()
        logger.info(f"Dedup reset for alert config {alert_id} ('{alert.name}'): {deleted} record(s) removed")
        return jsonify({'success': True, 'deleted': deleted,
                        'message': f'Cleared {deleted} dedup record(s). Next check cycle will resend.'})
    except Exception as e:
        logger.error(f"reset_alert_dedup error: {e}")
        return jsonify({'error': str(e)}), 500


@alerts_bp.route('/api/fim-agents', methods=['GET'])
@login_required
def fim_agents_lookup():
    """
    Return all agent names that have FIM/syscheck events in the last 24 h.
    Used by the UI "Look up agent names" helper so users can find the exact
    registered agent name without leaving the page.
    """
    try:
        from opensearch_api import OpenSearchAPI
        opensearch = OpenSearchAPI()
        agents = opensearch.get_fim_agents_diagnostic(hours=24)
        return jsonify({'agents': agents})
    except Exception as e:
        logger.error(f"FIM agents lookup error: {e}")
        return jsonify({'error': str(e), 'agents': []}), 500


@alerts_bp.route('/api/alert_configs/<int:alert_id>/debug', methods=['GET'])
@login_required
def debug_alert(alert_id):
    """Debug an alert configuration to see why it might not be triggering"""
    try:
        from models import AlertConfig, SentAlert
        from opensearch_api import OpenSearchAPI
        from email_alerts import EmailAlerts

        alert = AlertConfig.query.filter_by(id=alert_id, user_id=current_user.id).first()

        if not alert:
            return jsonify({'error': 'Alert configuration not found'}), 404

        ea = EmailAlerts()
        is_fim = getattr(alert, 'alert_type', 'standard') == 'fim'

        debug_info = {
            'alert_config': {
                'id': alert.id,
                'name': alert.name,
                'enabled': alert.enabled,
                'alert_type': getattr(alert, 'alert_type', 'standard') or 'standard',
                'notify_time': alert.notify_time,
                'email_recipient': alert.email_recipient,
                'alert_levels': alert.get_alert_levels(),
            },
            'current_time': {
                'utc': datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S'),
                'pakistan': (datetime.utcnow() + timedelta(hours=5)).strftime('%Y-%m-%d %H:%M:%S'),
            },
            'recent_alerts': [],
            'sent_alerts': [],
            'smtp_config': {
                'server': ea.smtp_server if ea.smtp_server else 'not configured',
                'username': 'configured' if ea.smtp_username else 'not configured',
            },
        }

        opensearch = OpenSearchAPI()
        end_time = datetime.utcnow().isoformat()

        if is_fim:
            # ── FIM-specific debug ────────────────────────────────────────────
            agent_names = alert.get_fim_agent_names() if hasattr(alert, 'get_fim_agent_names') else []
            paths = alert.get_fim_paths() if hasattr(alert, 'get_fim_paths') else []
            file_names = alert.get_fim_file_names() if hasattr(alert, 'get_fim_file_names') else []
            file_exts = alert.get_fim_file_extensions() if hasattr(alert, 'get_fim_file_extensions') else []

            debug_info['fim_config'] = {
                'agent_names': agent_names,
                'paths': paths,
                'file_names': file_names,
                'file_extensions': file_exts,
            }

            # Check 1: broad FIM events in the last 24 h (no agent/path filter)
            # so the user can see whether ANY FIM events exist in OpenSearch.
            start_broad = (datetime.utcnow() - timedelta(hours=24)).isoformat()
            broad_data = opensearch.search_fim_events(
                start_time=start_broad,
                end_time=end_time,
                agent_names=None,
                paths=None,
                limit=20,
            )
            broad_results = broad_data.get('results', [])
            debug_info['fim_all_agents_last_24h'] = {
                'total_found': broad_data.get('total', 0),
                'sample': [
                    {
                        'timestamp': r.get('source', {}).get('@timestamp'),
                        'agent_name': r.get('source', {}).get('agent', {}).get('name'),
                        'syscheck_path': r.get('source', {}).get('syscheck', {}).get('path'),
                        'rule_id': r.get('source', {}).get('rule', {}).get('id'),
                    }
                    for r in broad_results[:10]
                ],
            }

            # Check 2: narrowed to the configured agents/paths (same as the real alert)
            from models import SystemConfig
            interval_mins = int(SystemConfig.get_value('alert_check_interval', '2'))
            start_narrow = (datetime.utcnow() - timedelta(minutes=interval_mins)).isoformat()
            narrow_data = opensearch.search_fim_events(
                start_time=start_narrow,
                end_time=end_time,
                agent_names=agent_names if agent_names else None,
                paths=paths if paths else None,
                file_names=file_names if file_names else None,
                file_extensions=file_exts if file_exts else None,
                limit=20,
            )
            debug_info['fim_filtered_last_interval'] = {
                'interval_minutes': interval_mins,
                'total_found': narrow_data.get('total', 0),
                'error': narrow_data.get('error'),
                'sample': [
                    {
                        'timestamp': r.get('source', {}).get('@timestamp'),
                        'agent_name': r.get('source', {}).get('agent', {}).get('name'),
                        'syscheck_path': r.get('source', {}).get('syscheck', {}).get('path'),
                        'rule_id': r.get('source', {}).get('rule', {}).get('id'),
                    }
                    for r in narrow_data.get('results', [])[:10]
                ],
            }

            # Deduplications recorded in last 4 h
            dup_window = int(SystemConfig.get_value('alert_duplicate_window', '4'))
            debug_info['dedup_window_hours'] = dup_window
            debug_info['diagnosis'] = []

            if not agent_names:
                debug_info['diagnosis'].append('WARNING: No agent names configured — FIM alert will be skipped.')
            if not paths:
                debug_info['diagnosis'].append('WARNING: No paths configured — FIM alert will be skipped.')
            if broad_data.get('total', 0) == 0:
                debug_info['diagnosis'].append(
                    'No FIM events found in OpenSearch in the last 24 h for ANY agent. '
                    'Check that Wazuh syscheck is enabled and events are being indexed.'
                )
            elif narrow_data.get('total', 0) == 0 and agent_names and paths:
                debug_info['diagnosis'].append(
                    f'FIM events exist in OpenSearch (found {broad_data.get("total",0)} in 24 h) '
                    f'but NONE matched the configured agent(s) {agent_names} and path(s) {paths}. '
                    'Check that the agent name exactly matches what is stored in OpenSearch '
                    '(see fim_all_agents_last_24h for actual agent names) and that the path '
                    'prefix is correct (forward vs back slash, trailing separator).'
                )
            else:
                debug_info['diagnosis'].append(
                    f'Query looks correct — {narrow_data.get("total",0)} event(s) found in the '
                    f'last {interval_mins} min window. If no email was received, check the '
                    'sent_alerts list for deduplication records and verify SMTP credentials.'
                )

        else:
            # ── Standard severity-based debug ────────────────────────────────
            start_time = (datetime.utcnow() - timedelta(hours=1)).isoformat()
            alerts_data = opensearch.search_alerts(
                severity_levels=alert.get_alert_levels(),
                start_time=start_time,
                end_time=end_time,
                limit=10,
            )
            if 'results' in alerts_data:
                debug_info['recent_alerts'] = [
                    {
                        'timestamp': a.get('source', {}).get('@timestamp'),
                        'rule_id': a.get('source', {}).get('rule', {}).get('id'),
                        'agent_ip': a.get('source', {}).get('agent', {}).get('ip'),
                        'level': a.get('source', {}).get('rule', {}).get('level'),
                    }
                    for a in alerts_data['results'][:5]
                ]

        # Sent-alert history (last 24 h)
        recent_sent = SentAlert.query.filter(
            SentAlert.alert_config_id == alert_id,
            SentAlert.timestamp >= datetime.utcnow() - timedelta(hours=24),
        ).order_by(SentAlert.timestamp.desc()).limit(20).all()

        debug_info['sent_alerts'] = [
            {
                'identifier': sent.alert_identifier[:20] + '...',
                'timestamp': sent.timestamp.isoformat(),
            }
            for sent in recent_sent
        ]

        return jsonify(debug_info)

    except Exception as e:
        logger.error(f"Error debugging alert config: {str(e)}")
        return jsonify({'error': str(e)}), 500

@alerts_bp.route('/api/alerts/manual_check', methods=['POST'])
@login_required
def manual_alert_check():
    """Manually trigger alert checking for all enabled configurations"""
    try:
        if not current_user.is_admin():
            return jsonify({"error": "Admin access required"}), 403

        import scheduler

        # Run alert checking manually
        scheduler.check_alerts()

        return jsonify({
            "message": "Manual alert check completed successfully",
            "timestamp": datetime.utcnow().isoformat()
        })

    except Exception as e:
        logger.error(f"Error in manual alert check: {str(e)}")
        return jsonify({"error": str(e)}), 500

@alerts_bp.route('/api/alerts/export', methods=['GET'])
@login_required
def export_alerts():
    """Export alerts in CSV, XLSX, or PDF format"""
    try:
        import csv
        import io
        from flask import make_response
        from opensearch_api import OpenSearchAPI

        opensearch = OpenSearchAPI()

        # Get query parameters (same as get_alerts)
        severity_levels = request.args.getlist('severity_levels[]')
        if not severity_levels:
            severity_levels = ['critical', 'high']

        # Time range
        time_range = request.args.get('time_range', '24h')
        end_time = datetime.utcnow().isoformat()

        if time_range == '1h':
            start_time = (datetime.utcnow() - timedelta(hours=1)).isoformat()
        elif time_range == '6h':
            start_time = (datetime.utcnow() - timedelta(hours=6)).isoformat()
        elif time_range == '24h':
            start_time = (datetime.utcnow() - timedelta(days=1)).isoformat()
        elif time_range == '7d':
            start_time = (datetime.utcnow() - timedelta(days=7)).isoformat()
        elif time_range == '12d':
            start_time = (datetime.utcnow() - timedelta(days=12)).isoformat()
        elif time_range == '30d':
            start_time = (datetime.utcnow() - timedelta(days=30)).isoformat()
        elif time_range == '60d':
            start_time = (datetime.utcnow() - timedelta(days=60)).isoformat()
        elif time_range == '90d':
            start_time = (datetime.utcnow() - timedelta(days=90)).isoformat()
        else:
            start_time = request.args.get('start_time')
            end_time = request.args.get('end_time', end_time)

        # Get export format
        export_format = request.args.get('format', 'csv').lower()

        # Additional filters
        additional_filters = {}
        search_query = request.args.get('search_query')
        if search_query:
            additional_filters['search_query'] = search_query # Use search_query for consistency with search_alerts

        rule_id = request.args.get('rule_id')
        if rule_id:
            additional_filters['rule.id'] = rule_id

         # FIM alerts filter
        fim_alerts = request.args.get('fim_alerts')
        if fim_alerts == 'true':
            additional_filters['rule.id'] = ['553', '554']

        # Get all alerts for export (no pagination limit)
        results = opensearch.search_alerts(
            severity_levels=severity_levels,
            start_time=start_time,
            end_time=end_time,
            limit=10000,  # Large limit for export
            offset=0,
            sort_field='@timestamp',
            sort_order='desc',
            additional_filters=additional_filters
        )

        if 'results' not in results or not results['results']:
            return jsonify({'error': 'No alerts found'}), 404

        alerts_data = results['results']

        if export_format == 'csv':
            return export_alerts_csv(alerts_data)
        elif export_format == 'xlsx':
            return export_alerts_xlsx(alerts_data)
        elif export_format == 'pdf':
            from report_generator import ReportGenerator
            generator = ReportGenerator()
            
            # Prepare config-like object for ReportGenerator
            report_config = {
                'severity_levels': severity_levels
            }
            
            pdf_file = generator.generate_report(
                report_config=report_config,
                start_time=start_time,
                end_time=end_time,
                format='pdf',
                alerts_data=results
            )
            
            if pdf_file:
                response = make_response(pdf_file.getvalue())
                response.headers['Content-Type'] = 'application/pdf'
                response.headers['Content-Disposition'] = f'attachment; filename=alerts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
                return response
            else:
                return jsonify({'error': 'Failed to generate PDF report'}), 500
        else:
            return jsonify({'error': 'Unsupported format'}), 400

    except Exception as e:
        logger.error(f"Error exporting alerts: {str(e)}")
        return jsonify({'error': str(e)}), 500

def export_alerts_csv(alerts_data):
    """Export alerts as CSV"""
    import csv
    import io
    from flask import make_response

    output = io.StringIO()
    writer = csv.writer(output)

    # Write header
    writer.writerow([
        'Timestamp',
        'User',
        'Agent Name',
        'Agent ID', 
        'Agent IP',
        'Rule ID',
        'Rule Description',
        'Severity Level',
        'Location',
        'Full Log'
    ])

    # Write data
    for alert in alerts_data:
        source = alert.get('source', {})
        timestamp = source.get('@timestamp', 'N/A')
        agent = source.get('agent', {})
        rule = source.get('rule', {})

        # Format timestamp
        try:
            if timestamp != 'N/A':
                formatted_time = datetime.fromisoformat(timestamp.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
            else:
                formatted_time = 'N/A'
        except:
            formatted_time = timestamp

        writer.writerow([
            formatted_time,
            current_user.username if hasattr(current_user, 'username') else 'N/A',
            agent.get('name', 'N/A'),
            agent.get('id', 'N/A'),
            agent.get('ip', 'N/A'),
            rule.get('id', 'N/A'),
            rule.get('description', 'N/A'),
            rule.get('level', 'N/A'),
            agent.get('labels', {}).get('location', {}).get('set', 'N/A'),
            source.get('full_log', 'N/A')
        ])

    output.seek(0)
    response = make_response(output.getvalue())
    response.headers['Content-Type'] = 'text/csv'
    response.headers['Content-Disposition'] = f'attachment; filename=alerts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
    return response

def export_alerts_xlsx(alerts_data):
    """Export alerts as Excel file"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        import io
        from flask import make_response

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Security Alerts"

        # Header style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        # Write headers
        headers = [
            'Timestamp', 'User', 'Agent Name', 'Agent ID', 'Agent IP',
            'Rule ID', 'Rule Description', 'Severity Level', 'Location', 'Full Log'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        # Write data
        for row_idx, alert in enumerate(alerts_data, 2):
            source = alert.get('source', {})
            timestamp = source.get('@timestamp', 'N/A')
            agent = source.get('agent', {})
            rule = source.get('rule', {})

            # Format timestamp
            try:
                if timestamp != 'N/A':
                    formatted_time = datetime.fromisoformat(timestamp.replace('Z', '+00:00')).strftime('%Y-%m-%d %H:%M:%S')
                else:
                    formatted_time = 'N/A'
            except:
                formatted_time = timestamp

            row_data = [
                formatted_time,
                current_user.username if hasattr(current_user, 'username') else 'N/A',
                agent.get('name', 'N/A'),
                agent.get('id', 'N/A'),
                agent.get('ip', 'N/A'),
                rule.get('id', 'N/A'),
                rule.get('description', 'N/A'),
                rule.get('level', 'N/A'),
                agent.get('labels', {}).get('location', {}).get('set', 'N/A'),
                source.get('full_log', 'N/A')
            ]

            for col, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col, value=value)

        # Auto-fit columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=alerts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        return response

    except ImportError:
        return jsonify({'error': 'Excel export requires openpyxl package. Please install it.'}), 500

def export_alerts_pdf(alerts_data):
    """Export alerts as PDF"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        import io
        from flask import make_response

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4, topMargin=0.5*inch)

        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=1  # Center alignment
        )

        # Content list
        content = []

        # Title
        title = Paragraph("Security Alerts Export", title_style)
        content.append(title)
        content.append(Spacer(1, 12))

        # Export info
        export_info = Paragraph(f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", styles['Normal'])
        content.append(export_info)
        content.append(Paragraph(f"Total Alerts: {len(alerts_data)}", styles['Normal']))
        content.append(Spacer(1, 20))

        # Prepare table data
        table_data = [['Timestamp', 'Agent', 'Rule ID', 'Description', 'Level']]

        for alert in alerts_data[:100]:  # Limit to first 100 for PDF
            source = alert.get('source', {})
            timestamp = source.get('@timestamp', 'N/A')
            agent = source.get('agent', {})
            rule = source.get('rule', {})

            # Format timestamp
            try:
                if timestamp != 'N/A':
                    formatted_time = datetime.fromisoformat(timestamp.replace('Z', '+00:00')).strftime('%m/%d %H:%M')
                else:
                    formatted_time = 'N/A'
            except:
                formatted_time = timestamp[:10] if len(timestamp) > 10 else timestamp

            # Truncate long descriptions
            description = rule.get('description', 'N/A')
            if len(description) > 40:
                description = description[:37] + '...'

            table_data.append([
                formatted_time,
                agent.get('name', 'N/A')[:15],
                str(rule.get('id', 'N/A')),
                description,
                str(rule.get('level', 'N/A'))
            ])

        # Create table
        table = Table(table_data, colWidths=[1.2*inch, 1.2*inch, 0.8*inch, 2.5*inch, 0.6*inch])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ]))

        content.append(table)

        if len(alerts_data) > 100:
            content.append(Spacer(1, 12))
            content.append(Paragraph(f"Note: Only first 100 alerts shown. Total: {len(alerts_data)}", styles['Italic']))

        # Build PDF
        doc.build(content)
        buffer.seek(0)

        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=alerts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        return response

    except ImportError:
        return jsonify({'error': 'PDF export requires reportlab package. Please install it.'}), 500

@alerts_bp.route('/api/alert_configs/<int:alert_id>/test', methods=['POST'])
@login_required
def test_alert(alert_id):
    """Test an alert configuration by sending a test email"""
    try:
        from models import AlertConfig
        from opensearch_api import OpenSearchAPI
        from email_alerts import EmailAlerts

        alert = AlertConfig.query.filter_by(id=alert_id, user_id=current_user.id).first()
        if not alert:
            return jsonify({'error': 'Alert configuration not found'}), 404

        email_alerts = EmailAlerts()

        # ── FIM alert test ────────────────────────────────────────────────────
        if alert.is_fim():
            opensearch = OpenSearchAPI()
            end_time   = datetime.utcnow().isoformat()
            start_time = (datetime.utcnow() - timedelta(hours=24)).isoformat()

            fim_data = opensearch.search_fim_events(
                start_time=start_time,
                end_time=end_time,
                agent_names=alert.get_fim_agent_names() or None,
                paths=alert.get_fim_paths() or None,
                file_names=alert.get_fim_file_names() or None,
                file_extensions=alert.get_fim_file_extensions() or None,
                limit=5
            )

            if fim_data and 'results' in fim_data and fim_data['results']:
                # Build subject from first real event
                first = fim_data['results'][0]
                sc = first.get('source', {}).get('syscheck', {})
                ag = first.get('source', {}).get('agent', {})
                action_map = {'added': 'CREATED', 'modified': 'MODIFIED',
                              'deleted': 'DELETED', 'renamed': 'RENAMED'}
                action = action_map.get(sc.get('event', ''), 'CHANGED')
                file_path  = sc.get('path', 'Unknown path')
                agent_name = ag.get('name', 'Unknown agent')
                subject = (
                    f"[REBIZ Sentinel X] FIM Alert — File {action}: "
                    f"{file_path} on {agent_name}"
                )

                now_pkt = datetime.utcnow() + timedelta(hours=5)
                period_end   = now_pkt.strftime('%Y-%m-%d %H:%M')
                period_start = (now_pkt - timedelta(hours=24)).strftime('%Y-%m-%d %H:%M')

                body = email_alerts._build_fim_html_email_body(
                    config=alert,
                    fim_data=fim_data,
                    period_start=period_start,
                    period_end=period_end,
                    now_pkt=now_pkt,
                )
                success = email_alerts.send_alert_email(alert.email_recipient, subject, body)
                msg = f"FIM test alert sent with {len(fim_data['results'])} real event(s) to {alert.email_recipient}"
            else:
                # No real FIM events found — send a clearly-labelled test email
                # using the FIM template with synthetic data so the layout is correct.
                now_utc = datetime.utcnow()
                now_pkt = now_utc + timedelta(hours=5)
                ts_display = now_pkt.strftime('%Y-%m-%d %H:%M:%S PKT')

                agent_names = alert.get_fim_agent_names()
                paths       = alert.get_fim_paths()
                agent_name  = agent_names[0] if agent_names else 'demo-agent-01'
                file_path   = (paths[0].rstrip('/\\') + '/testfile.dll') if paths else r'C:\Windows\System32\testfile.dll'
                file_name   = file_path.split('/')[-1].split('\\')[-1]

                synthetic_event = {
                    'id': 'TEST-EVENT-001',
                    'source': {
                        '@timestamp': now_utc.isoformat(),
                        'agent': {'name': agent_name, 'ip': '10.0.0.1', 'id': '001'},
                        'hostname': agent_name,
                        'syscheck': {
                            'path': file_path,
                            'event': 'modified',
                            'size_after': '0',
                            'md5_after':    '(no real event — test mode)',
                            'sha1_after':   '(no real event — test mode)',
                            'sha256_after': '(no real event — test mode)',
                            'md5_before':   '(no real event — test mode)',
                            'uname_after': 'SYSTEM',
                            'audit': {'user': {'name': 'SYSTEM', 'id': 'S-1-5-18'}},
                        },
                        'rule': {
                            'id': '550',
                            'description': 'Integrity checksum changed (TEST MODE — no real events found in last 24 h)',
                            'level': 7,
                            'groups': 'syscheck',
                        },
                    }
                }
                synthetic_data = {'total': 1, 'results': [synthetic_event]}

                subject = (
                    f"[REBIZ Sentinel X] FIM Alert — File MODIFIED: "
                    f"{file_path} on {agent_name} [TEST]"
                )
                period_end   = now_pkt.strftime('%Y-%m-%d %H:%M')
                period_start = (now_pkt - timedelta(hours=24)).strftime('%Y-%m-%d %H:%M')

                body = email_alerts._build_fim_html_email_body(
                    config=alert,
                    fim_data=synthetic_data,
                    period_start=period_start,
                    period_end=period_end,
                    now_pkt=now_pkt,
                )
                success = email_alerts.send_alert_email(alert.email_recipient, subject, body)
                msg = f"FIM test email (synthetic data) sent to {alert.email_recipient} — no real events found in last 24 h"

            if success:
                return jsonify({'message': msg})
            else:
                return jsonify({'error': 'Failed to send FIM test email'}), 500

        # ── Standard (severity) alert test ───────────────────────────────────
        opensearch = OpenSearchAPI()
        severity_levels = alert.get_alert_levels()
        end_time   = datetime.utcnow().isoformat()
        start_time = (datetime.utcnow() - timedelta(hours=1)).isoformat()

        alerts_data = opensearch.search_alerts(
            severity_levels=severity_levels,
            start_time=start_time,
            end_time=end_time,
            limit=5
        )

        if alerts_data and 'results' in alerts_data and len(alerts_data['results']) > 0:
            success = email_alerts.send_severity_alert(alert, alerts_data)
            msg = f"Test alert sent with real data to {alert.email_recipient}"
        else:
            subject = f"[REBIZ Sentinel X] Test Alert — {alert.name}"
            include_fields = alert.get_include_fields() if hasattr(alert, 'get_include_fields') else []
            fields_html = (
                "<p><strong>Configured fields:</strong></p><ul>"
                + ''.join(f'<li>{f}</li>' for f in include_fields)
                + "</ul>"
            ) if include_fields else ""

            body = f"""<!DOCTYPE html><html><body style="font-family:Arial,sans-serif;padding:24px;">
<h2 style="color:#0d1b4b;">REBIZ Sentinel X — Test Alert</h2>
<div style="background:#f0f4ff;border-left:4px solid #4f46e5;padding:14px 18px;border-radius:6px;margin-bottom:16px;">
  <p><strong>Alert Name:</strong> {alert.name}</p>
  <p><strong>Levels:</strong> {', '.join(alert.get_alert_levels()) or 'N/A'}</p>
  <p><strong>Time:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
  {fields_html}
</div>
<p style="color:#555;">No real alerts matching your criteria were found in the last hour.
When actual alerts are found they will be sent automatically.</p>
</body></html>"""

            success = email_alerts.send_alert_email(alert.email_recipient, subject, body)
            msg = f"Test alert (no real data) sent to {alert.email_recipient}"

        if success:
            return jsonify({'message': msg})
        else:
            return jsonify({'error': 'Failed to send test alert email'}), 500

    except Exception as e:
        logger.error(f"Error sending test alert: {str(e)}")
        return jsonify({'error': str(e)}), 500