from flask import Blueprint, render_template, request, jsonify, Response, send_file
from flask_login import login_required, current_user
import logging
import io
import csv
import requests as _requests
from datetime import datetime
from config import Config

logger = logging.getLogger(__name__)

config_assessment_bp = Blueprint('config_assessment', __name__)


# ─────────────────────────────────────────────
#  Page
# ─────────────────────────────────────────────

@config_assessment_bp.route('/config-assessment')
@login_required
def index():
    return render_template('config_assessment.html')


# ─────────────────────────────────────────────
#  API — agent list
# ─────────────────────────────────────────────

@config_assessment_bp.route('/api/config-assessment/agents')
@login_required
def list_agents():
    """Return all Wazuh agents with basic SCA summary (pass/fail counts per agent)."""
    try:
        from wazuh_api import WazuhAPI
        api = WazuhAPI()

        agents_resp = api.get_agents({'limit': 1000})
        if not agents_resp or 'data' not in agents_resp:
            return jsonify({'error': 'Failed to fetch agents from Wazuh'}), 502

        agents = agents_resp['data'].get('affected_items', [])
        total = agents_resp['data'].get('total_affected_items', len(agents))

        result = []
        for agent in agents:
            agent_id = agent.get('id')
            result.append({
                'id': agent_id,
                'name': agent.get('name', 'Unknown'),
                'ip': agent.get('ip', 'N/A'),
                'os': _extract_os(agent),
                'status': agent.get('status', 'unknown'),
                'version': agent.get('version', 'N/A'),
                'group': ', '.join(agent.get('group', [])) if agent.get('group') else 'default',
                'last_keepalive': agent.get('lastKeepAlive', ''),
            })

        return jsonify({'agents': result, 'total': total})
    except _requests.exceptions.ConnectTimeout:
        msg = f"Connection timed out reaching Wazuh at {Config.WAZUH_API_URL}. Ensure the server is reachable from this machine."
        logger.error(msg)
        return jsonify({'error': msg}), 502
    except _requests.exceptions.ConnectionError as e:
        msg = f"Cannot connect to Wazuh at {Config.WAZUH_API_URL}: {e}"
        logger.error(msg)
        return jsonify({'error': msg}), 502
    except Exception as e:
        logger.error(f"Error listing agents for config assessment: {e}")
        return jsonify({'error': str(e)}), 500


# ─────────────────────────────────────────────
#  API — SCA policies for one agent
# ─────────────────────────────────────────────

@config_assessment_bp.route('/api/config-assessment/agent/<agent_id>/policies')
@login_required
def agent_policies(agent_id):
    """Return SCA policies and their summary scores for a given agent."""
    try:
        from wazuh_api import WazuhAPI
        api = WazuhAPI()

        resp = api.get_sca_policies(agent_id, {'limit': 500})
        if not resp or 'data' not in resp:
            return jsonify({'error': 'Failed to fetch SCA policies', 'policies': []}), 200

        policies = resp['data'].get('affected_items', [])
        result = []
        for p in policies:
            total_checks = p.get('total_checks', 0)
            passed = p.get('pass', 0)
            failed = p.get('fail', 0)
            invalid = p.get('invalid', 0)
            score = p.get('score', 0)

            result.append({
                'policy_id': p.get('policy_id', p.get('id', '')),
                'name': p.get('name', 'Unknown Policy'),
                'description': p.get('description', ''),
                'references': p.get('references', ''),
                'total_checks': total_checks,
                'pass': passed,
                'fail': failed,
                'invalid': invalid,
                'score': score,
                'end_scan': p.get('end_scan', ''),
                'start_scan': p.get('start_scan', ''),
                'hash_file': p.get('hash_file', ''),
            })

        return jsonify({'policies': result, 'total': len(result)})
    except Exception as e:
        logger.error(f"Error fetching SCA policies for agent {agent_id}: {e}")
        return jsonify({'error': str(e)}), 500


# ─────────────────────────────────────────────
#  API — SCA checks for one policy
# ─────────────────────────────────────────────

@config_assessment_bp.route('/api/config-assessment/agent/<agent_id>/policy/<policy_id>/checks')
@login_required
def policy_checks(agent_id, policy_id):
    """Return all checks for a specific SCA policy on a given agent."""
    try:
        from wazuh_api import WazuhAPI
        api = WazuhAPI()

        result_filter = request.args.get('result', '')
        search = request.args.get('search', '').strip()
        offset = int(request.args.get('offset', 0))
        limit = int(request.args.get('limit', 1000))

        params = {'limit': limit, 'offset': offset}
        if result_filter:
            params['result'] = result_filter

        resp = api.get_sca_checks(agent_id, policy_id, params)
        if not resp or 'data' not in resp:
            return jsonify({'error': 'Failed to fetch SCA checks', 'checks': []}), 200

        checks = resp['data'].get('affected_items', [])
        total = resp['data'].get('total_affected_items', len(checks))

        result = []
        for c in checks:
            title = c.get('title', 'Untitled Check')
            if search and search.lower() not in title.lower():
                continue

            check_result = c.get('result', 'not applicable')
            compliance = c.get('compliance', [])
            if isinstance(compliance, list):
                compliance_str = '; '.join(
                    f"{item.get('key','')}: {item.get('value','')}"
                    for item in compliance if isinstance(item, dict)
                )
            else:
                compliance_str = str(compliance)

            rules = c.get('rules', [])
            rules_str = '; '.join(
                r.get('rule', '') for r in rules if isinstance(r, dict)
            ) if isinstance(rules, list) else str(rules)

            result.append({
                'id': c.get('id', ''),
                'title': title,
                'description': c.get('description', ''),
                'rationale': c.get('rationale', ''),
                'remediation': c.get('remediation', ''),
                'result': check_result,
                'reason': c.get('reason', ''),
                'compliance': compliance_str,
                'rules': rules_str,
                'condition': c.get('condition', ''),
            })

        return jsonify({'checks': result, 'total': total})
    except Exception as e:
        logger.error(f"Error fetching checks for agent {agent_id} policy {policy_id}: {e}")
        return jsonify({'error': str(e)}), 500


# ─────────────────────────────────────────────
#  Downloads
# ─────────────────────────────────────────────

@config_assessment_bp.route('/api/config-assessment/agent/<agent_id>/policy/<policy_id>/download/csv')
@login_required
def download_csv(agent_id, policy_id):
    """Download all SCA checks for a policy as CSV."""
    try:
        from wazuh_api import WazuhAPI
        api = WazuhAPI()

        policy_resp = api.get_sca_policies(agent_id)
        policy_name = policy_id
        if policy_resp and 'data' in policy_resp:
            for p in policy_resp['data'].get('affected_items', []):
                if p.get('policy_id', p.get('id', '')) == policy_id:
                    policy_name = p.get('name', policy_id)
                    break

        checks_resp = api.get_sca_checks(agent_id, policy_id, {'limit': 2000})
        checks = []
        if checks_resp and 'data' in checks_resp:
            checks = checks_resp['data'].get('affected_items', [])

        output = io.StringIO()
        writer = csv.writer(output)
        writer.writerow(['ByteIT SentinelX — Configuration Assessment Report'])
        writer.writerow([f'Agent ID: {agent_id}', f'Policy: {policy_name}'])
        writer.writerow([f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}'])
        writer.writerow([])

        writer.writerow(['Check ID', 'Title', 'Result', 'Rationale', 'Remediation',
                         'Description', 'Compliance', 'Condition', 'Rules', 'Reason'])
        for c in checks:
            compliance = c.get('compliance', [])
            if isinstance(compliance, list):
                comp_str = '; '.join(
                    f"{i.get('key','')}: {i.get('value','')}"
                    for i in compliance if isinstance(i, dict)
                )
            else:
                comp_str = str(compliance)

            rules = c.get('rules', [])
            rules_str = '; '.join(
                r.get('rule', '') for r in rules if isinstance(r, dict)
            ) if isinstance(rules, list) else str(rules)

            writer.writerow([
                c.get('id', ''),
                c.get('title', ''),
                c.get('result', ''),
                c.get('rationale', ''),
                c.get('remediation', ''),
                c.get('description', ''),
                comp_str,
                c.get('condition', ''),
                rules_str,
                c.get('reason', ''),
            ])

        filename = f"sca_{agent_id}_{policy_id}_{datetime.utcnow().strftime('%Y%m%d')}.csv"
        output.seek(0)
        return Response(
            output.getvalue(),
            mimetype='text/csv',
            headers={'Content-Disposition': f'attachment; filename="{filename}"'}
        )
    except Exception as e:
        logger.error(f"CSV download error agent={agent_id} policy={policy_id}: {e}")
        return jsonify({'error': str(e)}), 500


@config_assessment_bp.route('/api/config-assessment/agent/<agent_id>/policy/<policy_id>/download/xlsx')
@login_required
def download_xlsx(agent_id, policy_id):
    """Download all SCA checks for a policy as XLSX with colour-coded results."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from wazuh_api import WazuhAPI

        api = WazuhAPI()

        policy_name = policy_id
        policy_resp = api.get_sca_policies(agent_id)
        if policy_resp and 'data' in policy_resp:
            for p in policy_resp['data'].get('affected_items', []):
                if p.get('policy_id', p.get('id', '')) == policy_id:
                    policy_name = p.get('name', policy_id)
                    break

        checks_resp = api.get_sca_checks(agent_id, policy_id, {'limit': 2000})
        checks = []
        if checks_resp and 'data' in checks_resp:
            checks = checks_resp['data'].get('affected_items', [])

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'SCA Checks'

        hdr_fill   = PatternFill('solid', fgColor='0F2744')
        pass_fill  = PatternFill('solid', fgColor='0A2A15')
        fail_fill  = PatternFill('solid', fgColor='4A1010')
        na_fill    = PatternFill('solid', fgColor='2A2A2A')
        err_fill   = PatternFill('solid', fgColor='3A2000')
        white_font = Font(color='FFFFFF', bold=True)
        pass_font  = Font(color='4ADE80', bold=True)
        fail_font  = Font(color='FF6B6B', bold=True)
        na_font    = Font(color='94A3B8')
        err_font   = Font(color='F97316', bold=True)
        val_font   = Font(color='E2E8F0')
        label_font = Font(color='94A3B8')

        ws.merge_cells('A1:J1')
        title_cell = ws.cell(row=1, column=1, value='ByteIT SentinelX — Configuration Assessment Report')
        title_cell.font = Font(color='FFFFFF', bold=True, size=14)
        title_cell.fill = hdr_fill
        title_cell.alignment = Alignment(horizontal='center')
        ws.row_dimensions[1].height = 28

        ws.merge_cells('A2:J2')
        meta_cell = ws.cell(row=2, column=1,
            value=f'Agent: {agent_id}  |  Policy: {policy_name}  |  Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}')
        meta_cell.font = Font(color='94A3B8', size=9)
        meta_cell.fill = PatternFill('solid', fgColor='0A0E1A')

        headers = ['Check ID', 'Title', 'Result', 'Rationale', 'Remediation',
                   'Description', 'Compliance', 'Condition', 'Rules', 'Reason']
        col_widths = [10, 50, 14, 55, 55, 40, 40, 16, 40, 30]
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            cell = ws.cell(row=3, column=ci, value=h)
            cell.font = white_font
            cell.fill = hdr_fill
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(ci)].width = w

        for row_i, c in enumerate(checks, 4):
            result = c.get('result', 'not applicable').lower()

            if result == 'passed':
                r_fill, r_font = pass_fill, pass_font
            elif result == 'failed':
                r_fill, r_font = fail_fill, fail_font
            elif result in ('error', 'not checked'):
                r_fill, r_font = err_fill, err_font
            else:
                r_fill, r_font = na_fill, na_font

            compliance = c.get('compliance', [])
            if isinstance(compliance, list):
                comp_str = '; '.join(
                    f"{i.get('key','')}: {i.get('value','')}"
                    for i in compliance if isinstance(i, dict)
                )
            else:
                comp_str = str(compliance)

            rules = c.get('rules', [])
            rules_str = '; '.join(
                r.get('rule', '') for r in rules if isinstance(r, dict)
            ) if isinstance(rules, list) else str(rules)

            row_data = [
                c.get('id', ''),
                c.get('title', ''),
                result.upper(),
                c.get('rationale', ''),
                c.get('remediation', ''),
                c.get('description', ''),
                comp_str,
                c.get('condition', ''),
                rules_str,
                c.get('reason', ''),
            ]
            for ci, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_i, column=ci, value=value)
                if ci == 3:
                    cell.font = r_font
                    cell.fill = r_fill
                else:
                    cell.font = val_font
                    cell.fill = r_fill
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            ws.row_dimensions[row_i].height = 30

        summary_ws = wb.create_sheet('Summary')
        summary_ws.merge_cells('A1:C1')
        sh = summary_ws.cell(row=1, column=1, value='Result Summary')
        sh.font = Font(color='FFFFFF', bold=True, size=12)
        sh.fill = hdr_fill
        sh.alignment = Alignment(horizontal='center')

        from collections import Counter
        counts = Counter(c.get('result', 'not applicable').lower() for c in checks)
        rows = [('passed', counts.get('passed', 0), '4ADE80'),
                ('failed', counts.get('failed', 0), 'FF6B6B'),
                ('not applicable', counts.get('not applicable', 0), '94A3B8'),
                ('error', counts.get('error', 0), 'F97316')]
        for ri, (label, count, color) in enumerate(rows, 2):
            summary_ws.cell(row=ri, column=1, value=label.upper()).font = Font(color=color, bold=True)
            summary_ws.cell(row=ri, column=2, value=count).font = Font(color='E2E8F0')
        summary_ws.column_dimensions['A'].width = 20
        summary_ws.column_dimensions['B'].width = 10

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"sca_{agent_id}_{policy_id}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    except Exception as e:
        logger.error(f"XLSX download error agent={agent_id} policy={policy_id}: {e}")
        return jsonify({'error': str(e)}), 500


@config_assessment_bp.route('/api/config-assessment/agent/<agent_id>/policy/<policy_id>/download/pdf')
@login_required
def download_pdf(agent_id, policy_id):
    """Download all SCA checks for a policy as a management-ready PDF."""
    try:
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.units import mm
        from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                        Paragraph, Spacer, HRFlowable)
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from wazuh_api import WazuhAPI

        api = WazuhAPI()

        policy_name = policy_id
        policy_resp = api.get_sca_policies(agent_id)
        if policy_resp and 'data' in policy_resp:
            for p in policy_resp['data'].get('affected_items', []):
                if p.get('policy_id', p.get('id', '')) == policy_id:
                    policy_name = p.get('name', policy_id)
                    break

        checks_resp = api.get_sca_checks(agent_id, policy_id, {'limit': 2000})
        checks = []
        if checks_resp and 'data' in checks_resp:
            checks = checks_resp['data'].get('affected_items', [])

        from collections import Counter
        counts = Counter(c.get('result', 'not applicable').lower() for c in checks)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                topMargin=15*mm, bottomMargin=15*mm,
                                leftMargin=15*mm, rightMargin=15*mm)

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle('title', parent=styles['Normal'],
                                     fontSize=16, textColor=colors.white,
                                     fontName='Helvetica-Bold', spaceAfter=4)
        sub_style = ParagraphStyle('sub', parent=styles['Normal'],
                                   fontSize=9, textColor=colors.HexColor('#94A3B8'),
                                   spaceAfter=10)
        cell_style = ParagraphStyle('cell', parent=styles['Normal'],
                                    fontSize=7, textColor=colors.HexColor('#E2E8F0'),
                                    fontName='Helvetica', leading=9)

        COLOR_BG    = colors.HexColor('#0A0E1A')
        COLOR_HDR   = colors.HexColor('#0F2744')
        COLOR_PASS  = colors.HexColor('#0A2A15')
        COLOR_FAIL  = colors.HexColor('#4A1010')
        COLOR_NA    = colors.HexColor('#1A1A2A')
        COLOR_ERR   = colors.HexColor('#3A2000')
        COLOR_PASS_TXT = colors.HexColor('#4ADE80')
        COLOR_FAIL_TXT = colors.HexColor('#FF6B6B')
        COLOR_NA_TXT   = colors.HexColor('#94A3B8')
        COLOR_ERR_TXT  = colors.HexColor('#F97316')
        COLOR_WHITE    = colors.white

        story = []

        story.append(Paragraph('ByteIT SentinelX — Configuration Assessment Report', title_style))
        story.append(Paragraph(
            f'Agent: {agent_id}  |  Policy: {policy_name}  |  '
            f'Generated: {datetime.utcnow().strftime("%Y-%m-%d %H:%M UTC")}', sub_style))
        story.append(HRFlowable(width='100%', color=colors.HexColor('#1e3a5f'), spaceAfter=6))

        summary_data = [
            ['Total Checks', 'Passed', 'Failed', 'Not Applicable', 'Error'],
            [str(len(checks)),
             str(counts.get('passed', 0)),
             str(counts.get('failed', 0)),
             str(counts.get('not applicable', 0)),
             str(counts.get('error', 0))]
        ]
        summary_table = Table(summary_data, colWidths=[40*mm]*5)
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), COLOR_HDR),
            ('BACKGROUND', (0,1), (0,1), COLOR_NA),
            ('BACKGROUND', (1,1), (1,1), COLOR_PASS),
            ('BACKGROUND', (2,1), (2,1), COLOR_FAIL),
            ('BACKGROUND', (3,1), (3,1), COLOR_NA),
            ('BACKGROUND', (4,1), (4,1), COLOR_ERR),
            ('TEXTCOLOR', (0,0), (-1,0), COLOR_WHITE),
            ('TEXTCOLOR', (0,1), (0,1), COLOR_WHITE),
            ('TEXTCOLOR', (1,1), (1,1), COLOR_PASS_TXT),
            ('TEXTCOLOR', (2,1), (2,1), COLOR_FAIL_TXT),
            ('TEXTCOLOR', (3,1), (3,1), COLOR_NA_TXT),
            ('TEXTCOLOR', (4,1), (4,1), COLOR_ERR_TXT),
            ('FONTNAME', (0,0), (-1,-1), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 10),
            ('ALIGN', (0,0), (-1,-1), 'CENTER'),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,0), (-1,-1), [None]),
            ('ROWHEIGHT', (0,0), (-1,-1), 18),
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#1e3a5f')),
            ('INNERGRID', (0,0), (-1,-1), 0.5, colors.HexColor('#1e3a5f')),
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 8*mm))

        col_widths_main = [12*mm, 65*mm, 22*mm, 60*mm, 68*mm, 32*mm]
        table_data = [[
            Paragraph('ID', cell_style),
            Paragraph('Title', cell_style),
            Paragraph('Result', cell_style),
            Paragraph('Rationale', cell_style),
            Paragraph('Remediation', cell_style),
            Paragraph('Compliance', cell_style),
        ]]

        for c in checks:
            result = c.get('result', 'not applicable').lower()
            if result == 'passed':
                row_bg, res_color = COLOR_PASS, COLOR_PASS_TXT
            elif result == 'failed':
                row_bg, res_color = COLOR_FAIL, COLOR_FAIL_TXT
            elif result in ('error', 'not checked'):
                row_bg, res_color = COLOR_ERR, COLOR_ERR_TXT
            else:
                row_bg, res_color = COLOR_NA, COLOR_NA_TXT

            compliance = c.get('compliance', [])
            if isinstance(compliance, list):
                comp_str = '; '.join(
                    f"{i.get('key','')}: {i.get('value','')}"
                    for i in compliance if isinstance(i, dict)
                )
            else:
                comp_str = str(compliance)

            res_style = ParagraphStyle('res', parent=cell_style, textColor=res_color, fontName='Helvetica-Bold')
            table_data.append([
                Paragraph(str(c.get('id', '')), cell_style),
                Paragraph(str(c.get('title', ''))[:200], cell_style),
                Paragraph(result.upper(), res_style),
                Paragraph(str(c.get('rationale', ''))[:300], cell_style),
                Paragraph(str(c.get('remediation', ''))[:300], cell_style),
                Paragraph(comp_str[:150], cell_style),
            ])

        main_table = Table(table_data, colWidths=col_widths_main, repeatRows=1)
        style_cmds = [
            ('BACKGROUND', (0,0), (-1,0), COLOR_HDR),
            ('TEXTCOLOR', (0,0), (-1,0), COLOR_WHITE),
            ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE', (0,0), (-1,-1), 7),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
            ('ALIGN', (0,0), (-1,-1), 'LEFT'),
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#1e3a5f')),
            ('INNERGRID', (0,0), (-1,-1), 0.3, colors.HexColor('#1e3a5f')),
            ('ROWHEIGHT', (0,0), (0,-1), 16),
        ]
        for ri, c in enumerate(checks, 1):
            result = c.get('result', 'not applicable').lower()
            if result == 'passed':
                bg = COLOR_PASS
            elif result == 'failed':
                bg = COLOR_FAIL
            elif result in ('error', 'not checked'):
                bg = COLOR_ERR
            else:
                bg = COLOR_NA
            style_cmds.append(('BACKGROUND', (0, ri), (-1, ri), bg))

        main_table.setStyle(TableStyle(style_cmds))
        story.append(main_table)

        doc.build(story)
        buf.seek(0)
        filename = f"sca_{agent_id}_{policy_id}_{datetime.utcnow().strftime('%Y%m%d')}.pdf"
        return send_file(buf, mimetype='application/pdf',
                         as_attachment=True, download_name=filename)
    except Exception as e:
        logger.error(f"PDF download error agent={agent_id} policy={policy_id}: {e}")
        return jsonify({'error': str(e)}), 500


# ─────────────────────────────────────────────
#  Helpers
# ─────────────────────────────────────────────

def _extract_os(agent):
    os_info = agent.get('os', {})
    if isinstance(os_info, dict):
        name = os_info.get('name', '')
        version = os_info.get('version', '')
        arch = os_info.get('arch', '')
        parts = [p for p in [name, version, arch] if p]
        return ' '.join(parts) if parts else 'Unknown'
    return str(os_info) if os_info else 'Unknown'
