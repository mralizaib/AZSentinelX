"""
AI DVR/NVR Auto-Configuration — Blueprint
Handles the 11-step guided workflow for configuring Hikvision / Platinum / LTS devices.
"""
import hashlib
import io
import json
import logging
import re
import smtplib
import time
from datetime import datetime
from email.mime.application import MIMEApplication
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import requests
from requests.auth import HTTPDigestAuth

from flask import (Blueprint, Response, jsonify, render_template,
                   request, send_file, session)
from flask_login import current_user, login_required

from config import Config
from models import DVRConfigSession, db
from routes.permissions import make_blueprint_permission_check

logger = logging.getLogger(__name__)

dvr_config_bp = Blueprint('dvr_config', __name__)
dvr_config_bp.before_request(make_blueprint_permission_check('dvr_config'))


# ─────────────────────────────────────────────
# Helper: build DVRClient from session data
# ─────────────────────────────────────────────

def _make_client(conn):
    from dvr_api import DVRClient
    return DVRClient(
        ip=conn['ip'],
        port=int(conn.get('port', 80)),
        username=conn['username'],
        password=conn['password'],
        use_https=conn.get('use_https', False),
    )


def _get_session_data(session_id):
    rec = DVRConfigSession.query.get(session_id)
    if not rec:
        return None
    return rec.get_data()


def _save_session_data(session_id, data):
    rec = DVRConfigSession.query.get(session_id)
    if not rec:
        rec = DVRConfigSession(id=session_id)
        db.session.add(rec)
    rec.set_data(data)
    rec.updated_at = datetime.utcnow()
    db.session.commit()


# ─────────────────────────────────────────────
# Main Page
# ─────────────────────────────────────────────

@dvr_config_bp.route('/dvr-config')
@login_required
def index():
    from dvr_api import STATE_TIMEZONE, TZ_TO_HIKVISION
    return render_template(
        'dvr_config.html',
        state_timezone_json=json.dumps(STATE_TIMEZONE),
        tz_to_hikvision_json=json.dumps(TZ_TO_HIKVISION),
    )


@dvr_config_bp.route('/dvr/history')
@login_required
def history_page():
    return render_template('dvr_history.html')


@dvr_config_bp.route('/dvr/templates')
@login_required
def templates_page():
    return render_template('dvr_templates.html')


@dvr_config_bp.route('/dvr/reports')
@login_required
def reports_page():
    return render_template('dvr_reports.html')


# ─────────────────────────────────────────────
# Step 1: Connect & Authenticate
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/connect', methods=['POST'])
@login_required
def connect():
    data = request.get_json() or {}
    required = ['device_name', 'ip', 'port', 'username', 'password']
    for f in required:
        if not data.get(f):
            return jsonify({'success': False, 'error': f'Missing field: {f}'}), 400

    try:
        client = _make_client(data)
        ok = client.test_connection()
        if not ok:
            return jsonify({'success': False,
                            'error': 'Authentication failed. Check IP, port, username and password.'}), 401

        # Push the friendly name onto the device itself — this previously
        # only lived in our local session/report data and never reached the
        # device, so the DVR/NVR's own name never actually changed.
        name_applied = client.configure_device_name(data['device_name'])
        if not name_applied:
            logger.warning(f"Could not set device name on {data['ip']}; continuing anyway")

        # Create a new config session
        sess_id = f"{current_user.id}_{int(time.time())}"
        sess_data = {
            'session_id': sess_id,
            'technician': current_user.username,
            'started_at': datetime.utcnow().isoformat(),
            'connection': {
                'device_name': data['device_name'],
                'ip': data['ip'],
                'port': data['port'],
                'username': data['username'],
                'password': data['password'],
                'use_https': data.get('use_https', False),
            },
            'steps': {},
        }
        _save_session_data(sess_id, sess_data)

        return jsonify({'success': True, 'session_id': sess_id})
    except Exception as e:
        logger.error(f"DVR connect error: {e}", exc_info=True)
        return jsonify({'success': False,
                        'error': f'Connection failed: {str(e)}. Verify the device is reachable.'}), 500


# ─────────────────────────────────────────────
# Step 2: Device Discovery
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/discover', methods=['POST'])
@login_required
def discover():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        info = client.get_device_info()

        sess['steps']['discovery'] = info
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'device_info': info})
    except Exception as e:
        logger.error(f"DVR discover error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 3: Location / Time / NTP
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-time', methods=['POST'])
@login_required
def configure_time():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    city = data.get('city', '')
    state = data.get('state', '')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        result = client.configure_time(city, state)
        result['city'] = city
        result['state'] = state

        sess['steps']['time'] = result
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'time_config': result})
    except Exception as e:
        logger.error(f"DVR time config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 4: Storage Detection & Formatting
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/storage', methods=['POST'])
@login_required
def storage_info():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        hdds = client.get_storage_info()

        sess['steps']['storage'] = {'hdds': hdds}
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'hdds': hdds})
    except Exception as e:
        logger.error(f"DVR storage error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@dvr_config_bp.route('/api/dvr/format-hdd', methods=['POST'])
@login_required
def format_hdd():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    hdd_id = data.get('hdd_id', '1')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        client.format_hdd(hdd_id)
        time.sleep(3)  # Brief pause; formatting is async on device
        hdds = client.get_storage_info()

        sess['steps']['storage'] = {'hdds': hdds, 'formatted': [hdd_id]}
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'message': f'HDD {hdd_id} formatting initiated.', 'hdds': hdds})
    except Exception as e:
        logger.error(f"DVR format HDD error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 5: Camera Stream Configuration
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-streams', methods=['POST'])
@login_required
def configure_streams():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        discovery = sess['steps'].get('discovery', {})
        # Prefer real channel IDs discovered from /Streaming/channels during
        # device discovery (e.g. ["101","201","301"]).  Fall back to synthesising
        # them from total_channels only when discovery didn't capture them.
        channel_ids = discovery.get('channel_ids') or []
        if not channel_ids:
            total_channels = int(discovery.get('total_channels', 0)) or 4
            channel_ids = [f"{ch}01" for ch in range(1, total_channels + 1)]

        results = []
        for main_id in channel_ids:
            ch_label = main_id[:-2] if len(main_id) > 2 else main_id
            main_ok  = client.configure_main_stream(main_id)
            sub_res  = client.configure_sub_stream(main_id)
            results.append({
                'channel':              ch_label,
                'main_stream':          'configured' if main_ok else 'error',
                'sub_stream':           sub_res['status'],
                'sub_resolution':       sub_res['applied_resolution'],
                'sub_note':             sub_res['note'],
            })

        client.disable_channel_zero()

        sess['steps']['streams'] = {'channels': results, 'channel_zero_disabled': True}
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'channels': results})
    except Exception as e:
        logger.error(f"DVR stream config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 6: Recording Schedule
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-recording', methods=['POST'])
@login_required
def configure_recording():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        discovery = sess['steps'].get('discovery', {})
        channel_ids = discovery.get('channel_ids') or []
        if not channel_ids:
            total_channels = int(discovery.get('total_channels', 0)) or 4
            channel_ids = [f"{ch}01" for ch in range(1, total_channels + 1)]

        results = []
        for main_id in channel_ids:
            ch_label = main_id[:-2] if len(main_id) > 2 else main_id
            ok = client.configure_recording_schedule(main_id)
            results.append({'channel': ch_label, 'status': 'configured' if ok else 'error'})

        sess['steps']['recording'] = {'channels': results}
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'channels': results})
    except Exception as e:
        logger.error(f"DVR recording config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 7: Holiday Schedule
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-holiday', methods=['POST'])
@login_required
def configure_holiday():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        ok = client.configure_holiday()

        holiday_info = {
            'name': 'Thanksgiving & Black Friday',
            'type': 'By Week',
            'start': '4th Monday of November',
            'end': '1st Monday of December',
            'recording_mode': 'Continuous',
            'status': 'configured' if ok else 'warn',
        }
        sess['steps']['holiday'] = holiday_info
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'holiday': holiday_info})
    except Exception as e:
        logger.error(f"DVR holiday config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 8 & 9: Users + Permissions
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/configure-users', methods=['POST'])
@login_required
def configure_users():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    initials = (data.get('client_initials') or '').strip().upper()
    # User Selection: caller chooses which standard accounts to actually
    # create on the device. Defaults to both for backward compatibility.
    selected_users = data.get('selected_users')
    if not selected_users:
        selected_users = ['cms', 'dlt']
    selected_users = [u for u in selected_users if u in ('cms', 'dlt', 'manager')]

    # Manager account: optional. The operator either ticks the "create
    # manager" checkbox, or types a manager username/password directly into
    # the wizard — either counts as "credentials provided" and triggers
    # creation. If nothing is provided, the manager step is skipped silently
    # (not an error, not a failed check).
    manager_username_input = (data.get('manager_username') or '').strip()
    manager_password_input = (data.get('manager_password') or '').strip()
    create_manager = (
        bool(data.get('create_manager', False))
        or 'manager' in selected_users
        or bool(manager_username_input)
        or bool(manager_password_input)
    )
    if create_manager and 'manager' not in selected_users:
        selected_users.append('manager')

    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404
    if not initials:
        return jsonify({'success': False, 'error': 'Client initials are required'}), 400
    core_users = [u for u in selected_users if u in ('cms', 'dlt')]
    if not core_users:
        return jsonify({'success': False, 'error': 'Select at least one account to create'}), 400

    cms_password     = f"{initials}_cam12"
    dlt_password     = f"{initials}9722IDT!"
    # Default manager username is "manager" (editable via manager_username);
    # default password follows the client-initials standard (editable via
    # manager_password), e.g. "RCman_12" for client initials "RC".
    manager_username = manager_username_input or 'manager'
    manager_password = manager_password_input or f"{initials}man_12"

    results = []
    try:
        client = _make_client(sess['connection'])

        def _account_status(result, created_label='created'):
            """Derive a human-readable status string from configure_standard_account result."""
            if not result.get('ok'):
                return 'error'
            already = result.get('already_existed', False)
            perm_st = result.get('permission_status', '')
            if perm_st == 'update_failed':
                return 'updated_permissions_failed' if already else 'created_permissions_failed'
            return 'updated' if already else created_label

        # CMS user (Viewer-level monitoring account).
        # Uses configure_standard_account: checks if the account exists first;
        # creates it if missing, updates permissions if they differ, and
        # continues without error either way.
        if 'cms' in selected_users:
            try:
                cms = client.configure_standard_account("cms", cms_password, "cms")
                status = _account_status(cms)
                entry = {
                    'username': 'cms',
                    'password': cms_password,
                    'role': 'viewer',
                    'status': status,
                    'already_existed': cms.get('already_existed', False),
                    'permission_status': cms.get('permission_status'),
                    'verified': cms.get('verified', False),
                    'permissions': 'Log Search, Live View, Playback, Video Export (Local); Log Search, Live View, Playback/Download (Remote)',
                }
                if status == 'error':
                    entry['error'] = cms.get('error', 'Unknown error')
                results.append(entry)
            except Exception as e:
                results.append({
                    'username': 'cms', 'password': cms_password,
                    'status': 'error', 'error': str(e),
                })

        # DLT user (Operator-level remote-admin account).
        # Same smart check-create-update flow as CMS.
        if 'dlt' in selected_users:
            try:
                dlt = client.configure_standard_account("dlt", dlt_password, "dlt")
                status = _account_status(dlt)
                entry = {
                    'username': 'dlt',
                    'password': dlt_password,
                    'role': 'operator',
                    'status': status,
                    'already_existed': dlt.get('already_existed', False),
                    'permission_status': dlt.get('permission_status'),
                    'verified': dlt.get('verified', False),
                    'permissions': 'Parameter Settings, Log Search, Shutdown/Reboot, Live View, Playback/Download (Remote)',
                }
                if status == 'error':
                    entry['error'] = dlt.get('error', 'Unknown error')
                results.append(entry)
            except Exception as e:
                results.append({
                    'username': 'dlt', 'password': dlt_password,
                    'status': 'error', 'error': str(e),
                })

        # Manager user (optional — Viewer-level, same permissions as CMS).
        # Creates the account if new, updates password if it already exists,
        # and only rewrites permissions when they differ from the required set.
        if 'manager' in selected_users:
            try:
                mgr = client.configure_standard_account(manager_username, manager_password, "manager")
                status = _account_status(mgr)
                entry = {
                    'username': manager_username,
                    'password': manager_password,
                    'role': 'viewer',
                    'status': status,
                    'already_existed': mgr.get('already_existed', False),
                    'permission_status': mgr.get('permission_status'),
                    'verified': mgr.get('verified', False),
                    'permissions': 'Log Search, Live View, Playback, Video Export (Local); Log Search, Live View, Playback/Download (Remote)',
                }
                if status == 'error':
                    entry['error'] = mgr.get('error', 'Unknown error')
                results.append(entry)
            except Exception as e:
                results.append({
                    'username': manager_username, 'password': manager_password,
                    'status': 'error', 'error': str(e),
                })

        sess['steps']['users'] = {
            'client_initials': initials,
            'selected_users': selected_users,
            'manager_created': 'manager' in selected_users,
            'accounts': results,
        }
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'users': results})
    except Exception as e:
        logger.error(f"DVR user config error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 10: Validation
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/validate', methods=['POST'])
@login_required
def validate():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    try:
        client = _make_client(sess['connection'])
        expected = dict(sess['steps'].get('time', {}))
        expected['device_name'] = sess['connection'].get('device_name', '')
        users_step = sess['steps'].get('users', {})
        expected['users_created'] = [
            a['username'] for a in users_step.get('accounts', [])
            if a.get('status') in ('created', 'created_permissions_failed')
        ]
        # Use the same channel IDs discovered/configured earlier in the wizard
        # so the validation report checks every real channel on this device
        # instead of falling back to a single hardcoded channel.
        discovery = sess['steps'].get('discovery', {})
        expected['channel_ids'] = discovery.get('channel_ids') or []
        checks = client.validate_configuration(expected)

        passed = sum(1 for c in checks if c['status'] == 'pass')
        warned = sum(1 for c in checks if c['status'] == 'warn')
        failed = sum(1 for c in checks if c['status'] == 'fail')

        overall = 'PASS' if failed == 0 and warned == 0 else ('WARNING' if failed == 0 else 'FAIL')

        sess['steps']['validation'] = {
            'checks': checks,
            'passed': passed,
            'warned': warned,
            'failed': failed,
            'overall': overall,
            'completed_at': datetime.utcnow().isoformat(),
        }
        _save_session_data(sess_id, sess)

        return jsonify({
            'success': True,
            'checks': checks,
            'summary': {'passed': passed, 'warned': warned, 'failed': failed, 'overall': overall},
        })
    except Exception as e:
        logger.error(f"DVR validate error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# Step 11: Report Generation
# ─────────────────────────────────────────────

def _build_report_data(sess):
    conn = sess.get('connection', {})
    disc = sess['steps'].get('discovery', {})
    time_cfg = sess['steps'].get('time', {})
    storage = sess['steps'].get('storage', {})
    streams = sess['steps'].get('streams', {})
    recording = sess['steps'].get('recording', {})
    holiday = sess['steps'].get('holiday', {})
    users_step = sess['steps'].get('users', {})
    validation = sess['steps'].get('validation', {})

    started = datetime.fromisoformat(sess.get('started_at', datetime.utcnow().isoformat()))
    completed_str = validation.get('completed_at', datetime.utcnow().isoformat())
    completed = datetime.fromisoformat(completed_str)
    duration_sec = int((completed - started).total_seconds())
    duration = f"{duration_sec // 60}m {duration_sec % 60}s"

    return {
        'device_name': conn.get('device_name', 'N/A'),
        'ip_address': conn.get('ip', 'N/A'),
        'manufacturer': disc.get('manufacturer', 'N/A'),
        'model': disc.get('model', 'N/A'),
        'device_type': disc.get('device_type', 'N/A'),
        'serial_number': disc.get('serial_number', 'N/A'),
        'firmware_version': disc.get('firmware_version', 'N/A'),
        'total_channels': disc.get('total_channels', 'N/A'),
        'active_channels': disc.get('active_channels', 'N/A'),
        'mac_address': disc.get('mac_address', 'N/A'),
        'timezone': time_cfg.get('iana_timezone', 'N/A'),
        'ntp_server': time_cfg.get('ntp_server', 'pool.ntp.org'),
        'city': time_cfg.get('city', 'N/A'),
        'state': time_cfg.get('state', 'N/A'),
        'hdds': storage.get('hdds', []),
        'stream_channels': streams.get('channels', []),
        'channel_zero_disabled': streams.get('channel_zero_disabled', False),
        'recording_channels': recording.get('channels', []),
        'holiday': holiday,
        'users': users_step.get('accounts', []),
        'client_initials': users_step.get('client_initials', ''),
        'validation': validation.get('checks', []),
        'validation_summary': {
            'passed': validation.get('passed', 0),
            'warned': validation.get('warned', 0),
            'failed': validation.get('failed', 0),
            'overall': validation.get('overall', 'N/A'),
        },
        'technician': sess.get('technician', 'N/A'),
        'configured_at': completed.strftime('%Y-%m-%d %H:%M:%S'),
        'total_duration': duration,
    }


def _build_html_report(report):
    status_icon = {'pass': '✅', 'warn': '⚠', 'fail': '❌'}
    status_color = {'pass': '#28a745', 'warn': '#ffc107', 'fail': '#dc3545'}
    overall_color = {'PASS': '#28a745', 'WARNING': '#ffc107', 'FAIL': '#dc3545'}
    overall = report['validation_summary']['overall']

    checks_html = ""
    for c in report['validation']:
        icon = status_icon.get(c['status'], '?')
        color = status_color.get(c['status'], '#999')
        action = f"<br><small style='color:#aaa;'>Action: {c['action']}</small>" if c.get('action') else ""
        checks_html += f"""
        <tr>
          <td style="color:{color};font-size:1.2em;">{icon}</td>
          <td><strong>{c['check']}</strong></td>
          <td>{c['detail']}{action}</td>
        </tr>"""

    users_html = ""
    for u in report['users']:
        users_html += f"""
        <tr>
          <td><strong>{u['username']}</strong></td>
          <td><code>{u.get('password','N/A')}</code></td>
          <td>{u.get('role','N/A')}</td>
          <td style="color:{'#28a745' if u.get('status')=='created' else '#dc3545'}">
            {u.get('status','N/A')}
          </td>
        </tr>"""

    hdds_html = ""
    for h in report['hdds']:
        hdds_html += f"<tr><td>{h['id']}</td><td>{h['name']}</td><td>{h['capacity']} MB</td><td>{h['status']}</td><td>{h['property']}</td></tr>"

    vs = report['validation_summary']
    return f"""<!DOCTYPE html>
<html><head><meta charset="UTF-8">
<style>
  body{{font-family:Arial,sans-serif;margin:30px;color:#222;}}
  h1{{color:#0056b3;}} h2{{color:#0056b3;border-bottom:2px solid #0056b3;padding-bottom:4px;margin-top:30px;}}
  table{{width:100%;border-collapse:collapse;margin-bottom:20px;}}
  th{{background:#0056b3;color:#fff;padding:8px;text-align:left;}}
  td{{padding:7px;border-bottom:1px solid #ddd;}}
  tr:hover{{background:#f5f5f5;}}
  .badge{{display:inline-block;padding:4px 12px;border-radius:4px;font-weight:bold;color:#fff;font-size:1.1em;}}
  .info-grid{{display:grid;grid-template-columns:1fr 1fr;gap:10px;margin-bottom:20px;}}
  .info-item{{background:#f9f9f9;border:1px solid #ddd;border-radius:6px;padding:10px;}}
  .info-item label{{font-size:0.8em;color:#666;display:block;}}
  .info-item span{{font-size:1em;font-weight:bold;}}
  code{{background:#eee;padding:2px 5px;border-radius:3px;font-size:0.95em;}}
</style></head><body>
<h1>🎥 ByteIT SentinalX — DVR/NVR Configuration Report</h1>
<p style="color:#666;">Generated: {report['configured_at']} &nbsp;|&nbsp; Technician: <strong>{report['technician']}</strong> &nbsp;|&nbsp; Duration: <strong>{report['total_duration']}</strong></p>

<div style="text-align:center;margin:20px 0;">
  <span class="badge" style="background:{overall_color.get(overall,'#999')};font-size:1.4em;">
    Overall Result: {overall}
  </span>
  &nbsp;
  <span class="badge" style="background:#28a745;">✅ {vs['passed']} Passed</span>
  &nbsp;
  <span class="badge" style="background:#ffc107;">⚠ {vs['warned']} Warnings</span>
  &nbsp;
  <span class="badge" style="background:#dc3545;">❌ {vs['failed']} Failed</span>
</div>

<h2>Device Information</h2>
<div class="info-grid">
  <div class="info-item"><label>Device Name</label><span>{report['device_name']}</span></div>
  <div class="info-item"><label>IP Address</label><span>{report['ip_address']}</span></div>
  <div class="info-item"><label>Manufacturer</label><span>{report['manufacturer']}</span></div>
  <div class="info-item"><label>Model</label><span>{report['model']}</span></div>
  <div class="info-item"><label>Device Type</label><span>{report['device_type']}</span></div>
  <div class="info-item"><label>Serial Number</label><span>{report['serial_number']}</span></div>
  <div class="info-item"><label>Firmware Version</label><span>{report['firmware_version']}</span></div>
  <div class="info-item"><label>MAC Address</label><span>{report['mac_address']}</span></div>
  <div class="info-item"><label>Total Channels</label><span>{report['total_channels']}</span></div>
  <div class="info-item"><label>Active Channels</label><span>{report['active_channels']}</span></div>
  <div class="info-item"><label>Location</label><span>{report['city']}, {report['state']}</span></div>
  <div class="info-item"><label>Timezone</label><span>{report['timezone']}</span></div>
</div>

<h2>HDD Information</h2>
<table><tr><th>ID</th><th>Name</th><th>Capacity</th><th>Status</th><th>Property</th></tr>
{hdds_html if hdds_html else '<tr><td colspan=5>No HDD data available</td></tr>'}
</table>

<h2>User Accounts</h2>
<table><tr><th>Username</th><th>Password</th><th>Role</th><th>Status</th></tr>
{users_html}</table>

<h2>Validation Results</h2>
<table><tr><th width="40">Status</th><th width="200">Check</th><th>Detail</th></tr>
{checks_html}</table>

<hr style="margin-top:40px;">
<p style="color:#999;font-size:0.85em;text-align:center;">
  ByteIT SentinalX — AI DVR/NVR Auto-Configuration Module &nbsp;|&nbsp; {report['configured_at']}
</p>
</body></html>"""


def _build_excel_report(report):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return None

    wb = openpyxl.Workbook()

    # Sheet 1: Device Info
    ws1 = wb.active
    ws1.title = "Device Info"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0056B3")

    ws1.append(["Field", "Value"])
    for cell in ws1[1]:
        cell.font = header_font
        cell.fill = header_fill

    fields = [
        ("Device Name", report['device_name']),
        ("IP Address", report['ip_address']),
        ("Manufacturer", report['manufacturer']),
        ("Model", report['model']),
        ("Device Type", report['device_type']),
        ("Serial Number", report['serial_number']),
        ("Firmware", report['firmware_version']),
        ("MAC Address", report['mac_address']),
        ("Total Channels", str(report['total_channels'])),
        ("Active Channels", str(report['active_channels'])),
        ("Location", f"{report['city']}, {report['state']}"),
        ("Timezone", report['timezone']),
        ("NTP Server", report['ntp_server']),
        ("Technician", report['technician']),
        ("Configured At", report['configured_at']),
        ("Duration", report['total_duration']),
    ]
    for row in fields:
        ws1.append(row)
    ws1.column_dimensions['A'].width = 22
    ws1.column_dimensions['B'].width = 40

    # Sheet 2: Validation
    ws2 = wb.create_sheet("Validation")
    ws2.append(["Status", "Check", "Detail", "Corrective Action"])
    for cell in ws2[1]:
        cell.font = header_font
        cell.fill = header_fill
    for c in report['validation']:
        icon = {'pass': '✅', 'warn': '⚠', 'fail': '❌'}.get(c['status'], '?')
        ws2.append([f"{icon} {c['status'].upper()}", c['check'], c['detail'], c.get('action', '')])
    for col in ['A', 'B', 'C', 'D']:
        ws2.column_dimensions[col].width = 28

    # Sheet 3: Users
    ws3 = wb.create_sheet("User Accounts")
    ws3.append(["Username", "Password", "Role", "Status", "Permissions"])
    for cell in ws3[1]:
        cell.font = header_font
        cell.fill = header_fill
    for u in report['users']:
        ws3.append([u['username'], u.get('password', 'N/A'), u.get('role', 'N/A'),
                     u.get('status', 'N/A'), u.get('permissions', '')])
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws3.column_dimensions[col].width = 30

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@dvr_config_bp.route('/api/dvr/report/<fmt>', methods=['POST'])
@login_required
def generate_report(fmt):
    """Generate report in pdf / excel / json format."""
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    report = _build_report_data(sess)
    device_name = report['device_name'].replace(' ', '_')
    ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    filename = f"DVR_Config_{device_name}_{ts}"

    if fmt == 'json':
        buf = io.BytesIO(json.dumps(report, indent=2).encode())
        buf.seek(0)
        return send_file(buf, as_attachment=True,
                         download_name=f"{filename}.json",
                         mimetype='application/json')

    elif fmt == 'excel':
        buf = _build_excel_report(report)
        if not buf:
            return jsonify({'success': False, 'error': 'openpyxl not available'}), 500
        return send_file(buf, as_attachment=True,
                         download_name=f"{filename}.xlsx",
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    elif fmt == 'pdf':
        html = _build_html_report(report)
        try:
            import weasyprint
            pdf_buf = io.BytesIO()
            weasyprint.HTML(string=html).write_pdf(pdf_buf)
            pdf_buf.seek(0)
            return send_file(pdf_buf, as_attachment=True,
                             download_name=f"{filename}.pdf",
                             mimetype='application/pdf')
        except Exception as e:
            logger.error(f"PDF generation failed: {e}")
            # Fallback: send HTML
            return Response(html, mimetype='text/html',
                            headers={'Content-Disposition': f'attachment; filename="{filename}.html"'})

    return jsonify({'success': False, 'error': f'Unknown format: {fmt}'}), 400


@dvr_config_bp.route('/api/dvr/email-report', methods=['POST'])
@login_required
def email_report():
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    recipients = data.get('recipients', [])
    formats = data.get('formats', ['pdf'])
    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404
    if not recipients:
        return jsonify({'success': False, 'error': 'No recipients specified'}), 400

    report = _build_report_data(sess)
    device_name = report['device_name']
    ts = datetime.utcnow().strftime('%Y%m%d_%H%M%S')
    filename_base = f"DVR_Config_{device_name.replace(' ', '_')}_{ts}"
    overall = report['validation_summary']['overall']

    msg = MIMEMultipart('mixed')
    msg['Subject'] = f"DVR/NVR Configuration Report — {device_name} [{overall}]"
    msg['From'] = Config.SMTP_USERNAME
    msg['To'] = ', '.join(recipients)

    # HTML body
    vs = report['validation_summary']
    body_html = f"""<html><body style="font-family:Arial,sans-serif;">
    <h2 style="color:#0056b3;">DVR/NVR Configuration Complete</h2>
    <p><strong>Device:</strong> {device_name} ({report['ip_address']})</p>
    <p><strong>Model:</strong> {report['manufacturer']} {report['model']}</p>
    <p><strong>Result:</strong> <strong style="color:{'green' if overall=='PASS' else ('orange' if overall=='WARNING' else 'red')}">{overall}</strong></p>
    <p><strong>Validation:</strong> ✅ {vs['passed']} Passed &nbsp; ⚠ {vs['warned']} Warnings &nbsp; ❌ {vs['failed']} Failed</p>
    <p><strong>Technician:</strong> {report['technician']}</p>
    <p><strong>Configured At:</strong> {report['configured_at']}</p>
    <p><strong>Duration:</strong> {report['total_duration']}</p>
    <hr>
    <p style="color:#999;font-size:0.85em;">ByteIT SentinalX — AI DVR/NVR Auto-Configuration Module</p>
    </body></html>"""
    msg.attach(MIMEText(body_html, 'html'))

    # Attachments
    if 'pdf' in formats:
        try:
            html = _build_html_report(report)
            import weasyprint
            pdf_bytes = io.BytesIO()
            weasyprint.HTML(string=html).write_pdf(pdf_bytes)
            part = MIMEApplication(pdf_bytes.getvalue(), Name=f"{filename_base}.pdf")
            part['Content-Disposition'] = f'attachment; filename="{filename_base}.pdf"'
            msg.attach(part)
        except Exception as e:
            logger.warning(f"PDF attachment failed: {e}")

    if 'excel' in formats:
        try:
            buf = _build_excel_report(report)
            if buf:
                part = MIMEApplication(buf.read(), Name=f"{filename_base}.xlsx")
                part['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
                msg.attach(part)
        except Exception as e:
            logger.warning(f"Excel attachment failed: {e}")

    if 'json' in formats:
        json_bytes = json.dumps(report, indent=2).encode()
        part = MIMEApplication(json_bytes, Name=f"{filename_base}.json")
        part['Content-Disposition'] = f'attachment; filename="{filename_base}.json"'
        msg.attach(part)

    try:
        with smtplib.SMTP(Config.SMTP_SERVER, Config.SMTP_PORT) as smtp:
            if Config.SMTP_USE_TLS:
                smtp.starttls()
            smtp.login(Config.SMTP_USERNAME, Config.SMTP_PASSWORD)
            smtp.sendmail(Config.SMTP_USERNAME, recipients, msg.as_string())
        return jsonify({'success': True, 'message': f'Report sent to {", ".join(recipients)}'})
    except Exception as e:
        logger.error(f"Email send failed: {e}", exc_info=True)
        return jsonify({'success': False, 'error': f'Email failed: {str(e)}'}), 500


# ─────────────────────────────────────────────
# Session History
# ─────────────────────────────────────────────

@dvr_config_bp.route('/api/dvr/sessions', methods=['GET'])
@login_required
def list_sessions():
    sessions = DVRConfigSession.query.order_by(DVRConfigSession.updated_at.desc()).limit(20).all()
    result = []
    for s in sessions:
        d = s.get_data()
        conn   = d.get('connection', {})
        val    = d.get('steps', {}).get('validation', {})
        checks = val.get('checks', [])
        result.append({
            'id':          s.id,
            'device_name': conn.get('device_name', 'N/A'),
            'ip':          conn.get('ip', 'N/A'),
            'overall':     val.get('overall', 'Incomplete'),
            'passed':      sum(1 for c in checks if c.get('status') == 'pass'),
            'warned':      sum(1 for c in checks if c.get('status') == 'warn'),
            'failed':      sum(1 for c in checks if c.get('status') == 'fail'),
            'technician':  d.get('technician', 'N/A'),
            'started_at':  d.get('started_at', ''),
            'updated_at':  s.updated_at.isoformat() if s.updated_at else '',
        })
    return jsonify({'sessions': result})


@dvr_config_bp.route('/api/dvr/session/<session_id>', methods=['GET'])
@login_required
def get_session(session_id):
    sess = _get_session_data(session_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404
    return jsonify({'success': True, 'session': sess})


# ─────────────────────────────────────────────
# Camera Name Management (standalone feature)
# ─────────────────────────────────────────────

@dvr_config_bp.route('/camera-names')
@login_required
def camera_names():
    return render_template('camera_names.html')


@dvr_config_bp.route('/api/dvr/camera-names/connect', methods=['POST'])
@login_required
def camera_names_connect():
    data = request.get_json() or {}
    required = ['ip', 'port', 'username', 'password']
    for f in required:
        if not data.get(f):
            return jsonify({'success': False, 'error': f'Missing field: {f}'}), 400

    try:
        client = _make_client(data)
        ok = client.test_connection()
        if not ok:
            return jsonify({'success': False,
                            'error': 'Authentication failed. Check IP, port, username and password.'}), 401

        channels = client.list_active_channels_for_naming()

        sess_id = f"camname_{current_user.id}_{int(time.time())}"
        sess_data = {
            'session_id': sess_id,
            'technician': current_user.username,
            'started_at': datetime.utcnow().isoformat(),
            'connection': {
                'ip': data['ip'],
                'port': data['port'],
                'username': data['username'],
                'password': data['password'],
                'use_https': data.get('use_https', False),
            },
            'steps': {'channels_detected': channels},
        }
        _save_session_data(sess_id, sess_data)

        return jsonify({
            'success': True,
            'session_id': sess_id,
            'channels': channels,
            'skipped_note': 'Offline or disabled channels are automatically excluded from this list.',
        })
    except Exception as e:
        logger.error(f"Camera name management connect error: {e}", exc_info=True)
        return jsonify({'success': False,
                        'error': f'Connection failed: {str(e)}. Verify the device is reachable.'}), 500


@dvr_config_bp.route('/api/dvr/camera-names/apply', methods=['POST'])
@login_required
def camera_names_apply():
    """
    Apply a naming template to all currently-active channels.
    Body: {session_id, template, assignments?: {channel_id: name}}
    `template` supports the placeholder {n} for a 1-based sequence number,
    e.g. "Camera {n}", "Front Lobby {n}". If `assignments` is provided it
    takes precedence per-channel (manual override of the template result).
    """
    data = request.get_json() or {}
    sess_id = data.get('session_id')
    template = (data.get('template') or 'Camera {n}').strip()
    assignments = data.get('assignments') or {}

    sess = _get_session_data(sess_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404

    channels = sess.get('steps', {}).get('channels_detected', [])
    if not channels:
        return jsonify({'success': False, 'error': 'No active channels detected for this session'}), 400

    try:
        client = _make_client(sess['connection'])
        results = []
        for i, ch in enumerate(channels, start=1):
            ch_id = ch['channel_id']
            new_name = assignments.get(ch_id) or template.replace('{n}', str(i).zfill(2))
            ok, detail = client.apply_channel_name(ch_id, new_name)
            results.append({
                'channel_id': ch_id,
                'previous_name': ch.get('name', ''),
                'new_name': new_name,
                'status': 'success' if ok else 'failed',
                'detail': detail,
            })

        sess['steps']['naming_report'] = {
            'template': template,
            'applied_at': datetime.utcnow().isoformat(),
            'results': results,
        }
        _save_session_data(sess_id, sess)

        return jsonify({'success': True, 'results': results})
    except Exception as e:
        logger.error(f"Camera name management apply error: {e}", exc_info=True)
        return jsonify({'success': False, 'error': str(e)}), 500


@dvr_config_bp.route('/api/dvr/camera-names/report/<session_id>', methods=['GET'])
@login_required
def camera_names_report(session_id):
    sess = _get_session_data(session_id)
    if not sess:
        return jsonify({'success': False, 'error': 'Session not found'}), 404
    report = sess.get('steps', {}).get('naming_report')
    if not report:
        return jsonify({'success': False, 'error': 'No naming report available yet for this session'}), 404

    total = len(report['results'])
    succeeded = sum(1 for r in report['results'] if r['status'] == 'success')
    return jsonify({
        'success': True,
        'session_id': session_id,
        'connection_ip': sess.get('connection', {}).get('ip'),
        'template': report['template'],
        'applied_at': report['applied_at'],
        'total_channels': total,
        'succeeded': succeeded,
        'failed': total - succeeded,
        'results': report['results'],
    })


# ─────────────────────────────────────────────────────────────────────────────
# DVR Diagnostic Tool  — /dvr/diagnostic
# Lets you test user-creation against a device directly from the browser.
# Tries every known approach and shows you the raw device XML responses.
# ─────────────────────────────────────────────────────────────────────────────

@dvr_config_bp.route('/dvr/diagnostic', methods=['GET', 'POST'])
@login_required
def dvr_diagnostic():
    """Browser-based diagnostic: test user creation against any DVR/NVR."""
    if request.method == 'GET':
        return render_template('dvr_diagnostic.html')

    # ── Collect form inputs ──────────────────────────────────────────────────
    ip       = (request.form.get('ip') or '').strip()
    port     = int(request.form.get('port') or 80)
    username = (request.form.get('username') or '').strip()
    password = (request.form.get('password') or '').strip()

    if not ip or not username or not password:
        return render_template('dvr_diagnostic.html',
                               error='IP, username and password are required.')

    base     = f"http://{ip}:{port}/ISAPI"
    auth     = HTTPDigestAuth(username, password)
    sess_req = requests.Session()
    headers  = {"Content-Type": "application/xml"}
    timeout  = 15

    TEST_USER       = "_dvr_diag_"
    TEST_PASS_PLAIN = "Test@12345"
    TEST_PASS_MD5   = hashlib.md5(TEST_PASS_PLAIN.encode()).hexdigest()

    steps = []   # list of dicts: {label, sent, status, response, success}

    def _do_get(path):
        url = base + path
        try:
            r = sess_req.get(url, auth=auth, verify=False, timeout=timeout)
            return r.status_code, r.text
        except Exception as exc:
            return None, str(exc)

    def _do_post(path, body):
        url = base + path
        try:
            r = sess_req.post(url, auth=auth, verify=False, timeout=timeout,
                              data=body, headers=headers)
            return r.status_code, r.text
        except Exception as exc:
            return None, str(exc)

    def _do_put(path, body):
        url = base + path
        try:
            r = sess_req.put(url, auth=auth, verify=False, timeout=timeout,
                             data=body, headers=headers)
            return r.status_code, r.text
        except Exception as exc:
            return None, str(exc)

    def _user_exists(raw_xml, uname):
        return f"<userName>{uname}</userName>" in raw_xml or \
               f"<userName xmlns" in raw_xml and uname in raw_xml

    def _find_user_id(raw_xml, uname):
        """Return the id text node that appears just before userName=uname."""
        pattern = r'<id>(\d+)</id>\s*<userName>' + re.escape(uname) + r'</userName>'
        m = re.search(pattern, raw_xml)
        return m.group(1) if m else None

    def _cleanup(uid):
        if uid:
            _do_delete(f"/Security/users/{uid}")

    def _do_delete(path):
        url = base + path
        try:
            sess_req.delete(url, auth=auth, verify=False, timeout=timeout)
        except Exception:
            pass

    # ── Step 1: GET /Security/users ──────────────────────────────────────────
    st1, resp1 = _do_get("/Security/users")
    steps.append({
        'label': 'GET /ISAPI/Security/users  (raw device XML)',
        'sent': None,
        'status': st1,
        'response': resp1,
        'success': st1 == 200,
    })
    raw_userlist = resp1 if st1 == 200 else ''

    # Detect namespace
    ns_match = re.search(r'xmlns="([^"]+)"', raw_userlist)
    detected_ns = ns_match.group(1) if ns_match else ''

    # Shared fields used in every user block
    def _bond():
        return """  <bondIpList>
    <bondIp>
      <id>1</id>
      <ipAddress>0.0.0.0</ipAddress>
      <ipv6Address>::</ipv6Address>
    </bondIp>
  </bondIpList>"""

    created_uid = None
    success_method = None

    # ── Step 2: POST — plain password, no xmlns ──────────────────────────────
    if not success_method:
        body = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<User>\n'
            '  <id>0</id>\n'
            f'  <userName>{TEST_USER}</userName>\n'
            f'  <password>{TEST_PASS_PLAIN}</password>\n'
            + _bond() + '\n'
            '  <macAddress></macAddress>\n'
            '  <userLevel>Viewer</userLevel>\n'
            '  <attribute><inherent>false</inherent></attribute>\n'
            '</User>'
        )
        st, resp = _do_post("/Security/users", body)
        ok = st == 200 and 'statusCode' not in resp
        steps.append({
            'label': 'POST /ISAPI/Security/users — plain password, no xmlns',
            'sent': body, 'status': st, 'response': resp, 'success': ok,
        })
        if ok:
            _, fresh = _do_get("/Security/users")
            created_uid = _find_user_id(fresh, TEST_USER)
            success_method = 'POST plain password'

    # ── Step 3: POST — MD5 password, no xmlns ───────────────────────────────
    if not success_method:
        body = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            '<User>\n'
            '  <id>0</id>\n'
            f'  <userName>{TEST_USER}</userName>\n'
            f'  <password>{TEST_PASS_MD5}</password>\n'
            + _bond() + '\n'
            '  <macAddress></macAddress>\n'
            '  <userLevel>Viewer</userLevel>\n'
            '  <attribute><inherent>false</inherent></attribute>\n'
            '</User>'
        )
        st, resp = _do_post("/Security/users", body)
        ok = st == 200 and 'statusCode' not in resp
        steps.append({
            'label': f'POST /ISAPI/Security/users — MD5 password ({TEST_PASS_MD5})',
            'sent': body, 'status': st, 'response': resp, 'success': ok,
        })
        if ok:
            _, fresh = _do_get("/Security/users")
            created_uid = _find_user_id(fresh, TEST_USER)
            success_method = 'POST MD5 password'

    # ── Step 4: POST — plain password, with xmlns ────────────────────────────
    if not success_method and detected_ns:
        body = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            f'<User xmlns="{detected_ns}">\n'
            '  <id>0</id>\n'
            f'  <userName>{TEST_USER}</userName>\n'
            f'  <password>{TEST_PASS_PLAIN}</password>\n'
            + _bond() + '\n'
            '  <macAddress></macAddress>\n'
            '  <userLevel>Viewer</userLevel>\n'
            '  <attribute><inherent>false</inherent></attribute>\n'
            '</User>'
        )
        st, resp = _do_post("/Security/users", body)
        ok = st == 200 and 'statusCode' not in resp
        steps.append({
            'label': f'POST — plain password, xmlns="{detected_ns}"',
            'sent': body, 'status': st, 'response': resp, 'success': ok,
        })
        if ok:
            _, fresh = _do_get("/Security/users")
            created_uid = _find_user_id(fresh, TEST_USER)
            success_method = f'POST plain+xmlns'

    # ── Step 5: POST — MD5 password, with xmlns ──────────────────────────────
    if not success_method and detected_ns:
        body = (
            '<?xml version="1.0" encoding="UTF-8"?>\n'
            f'<User xmlns="{detected_ns}">\n'
            '  <id>0</id>\n'
            f'  <userName>{TEST_USER}</userName>\n'
            f'  <password>{TEST_PASS_MD5}</password>\n'
            + _bond() + '\n'
            '  <macAddress></macAddress>\n'
            '  <userLevel>Viewer</userLevel>\n'
            '  <attribute><inherent>false</inherent></attribute>\n'
            '</User>'
        )
        st, resp = _do_post("/Security/users", body)
        ok = st == 200 and 'statusCode' not in resp
        steps.append({
            'label': f'POST — MD5 password, xmlns="{detected_ns}"',
            'sent': body, 'status': st, 'response': resp, 'success': ok,
        })
        if ok:
            _, fresh = _do_get("/Security/users")
            created_uid = _find_user_id(fresh, TEST_USER)
            success_method = 'POST MD5+xmlns'

    # ── Step 6: Clone first <User> from GET, patch, POST ────────────────────
    if not success_method and raw_userlist:
        m = re.search(
            r'(<(?:[^:>\s/]+:)?User(?:\s[^>]*)?>.*?</(?:[^:>\s/]+:)?User>)',
            raw_userlist, re.DOTALL
        )
        if m:
            cloned = m.group(1)
            for tag, val in [('id', '0'), ('userName', TEST_USER),
                              ('password', TEST_PASS_PLAIN),
                              ('userLevel', 'Viewer'), ('inherent', 'false')]:
                cloned = re.sub(
                    rf'(<(?:[^:>\s/]+:)?{re.escape(tag)}(?:\s[^>]*)?>)'
                    rf'.*?(</(?:[^:>\s/]+:)?{re.escape(tag)}>)',
                    rf'\g<1>{val}\g<2>', cloned, flags=re.DOTALL
                )
            body = '<?xml version="1.0" encoding="UTF-8"?>\n' + cloned
            st, resp = _do_post("/Security/users", body)
            ok = st == 200 and 'statusCode' not in resp
            steps.append({
                'label': 'POST — cloned first <User> from device GET, patched fields',
                'sent': body, 'status': st, 'response': resp, 'success': ok,
            })
            if ok:
                _, fresh = _do_get("/Security/users")
                created_uid = _find_user_id(fresh, TEST_USER)
                success_method = 'POST cloned User block'

    # ── Step 7: PUT full UserList + injected user (plain password) ───────────
    if not success_method and raw_userlist:
        new_block = (
            '  <User>\n'
            '    <id>0</id>\n'
            f'    <userName>{TEST_USER}</userName>\n'
            f'    <password>{TEST_PASS_PLAIN}</password>\n'
            '    <bondIpList><bondIp><id>1</id>'
            '<ipAddress>0.0.0.0</ipAddress>'
            '<ipv6Address>::</ipv6Address></bondIp></bondIpList>\n'
            '    <macAddress></macAddress>\n'
            '    <userLevel>Viewer</userLevel>\n'
            '    <attribute><inherent>false</inherent></attribute>\n'
            '  </User>\n'
        )
        modified, n = re.subn(
            r'(</(?:[^:>\s/]+:)?[Uu]ser[Ll]ist>)',
            new_block + r'\1', raw_userlist
        )
        if n:
            st, resp = _do_put("/Security/users", modified)
            ok = st in (200, 204)
            steps.append({
                'label': 'PUT /ISAPI/Security/users — full UserList with new user injected (plain password)',
                'sent': modified[:3000] + ('…(truncated)' if len(modified) > 3000 else ''),
                'status': st, 'response': resp, 'success': ok,
            })
            if ok:
                _, fresh = _do_get("/Security/users")
                created_uid = _find_user_id(fresh, TEST_USER)
                if created_uid:
                    success_method = 'PUT UserList plain password'
                else:
                    steps[-1]['success'] = False
                    steps[-1]['response'] += '\n\n[Note: PUT returned 200 but user not found in refreshed list]'

    # ── Step 8: PUT full UserList + injected user (MD5 password) ─────────────
    if not success_method and raw_userlist:
        new_block_md5 = (
            '  <User>\n'
            '    <id>0</id>\n'
            f'    <userName>{TEST_USER}</userName>\n'
            f'    <password>{TEST_PASS_MD5}</password>\n'
            '    <bondIpList><bondIp><id>1</id>'
            '<ipAddress>0.0.0.0</ipAddress>'
            '<ipv6Address>::</ipv6Address></bondIp></bondIpList>\n'
            '    <macAddress></macAddress>\n'
            '    <userLevel>Viewer</userLevel>\n'
            '    <attribute><inherent>false</inherent></attribute>\n'
            '  </User>\n'
        )
        modified_md5, n = re.subn(
            r'(</(?:[^:>\s/]+:)?[Uu]ser[Ll]ist>)',
            new_block_md5 + r'\1', raw_userlist
        )
        if n:
            st, resp = _do_put("/Security/users", modified_md5)
            ok = st in (200, 204)
            steps.append({
                'label': 'PUT /ISAPI/Security/users — full UserList with new user injected (MD5 password)',
                'sent': modified_md5[:3000] + ('…(truncated)' if len(modified_md5) > 3000 else ''),
                'status': st, 'response': resp, 'success': ok,
            })
            if ok:
                _, fresh = _do_get("/Security/users")
                created_uid = _find_user_id(fresh, TEST_USER)
                if created_uid:
                    success_method = 'PUT UserList MD5 password'
                else:
                    steps[-1]['success'] = False
                    steps[-1]['response'] += '\n\n[Note: PUT returned 200 but user not found in refreshed list]'

    # ── Cleanup test user ────────────────────────────────────────────────────
    if created_uid:
        _cleanup(created_uid)

    return render_template(
        'dvr_diagnostic.html',
        ran=True,
        ip=ip, port=port, username=username,
        steps=steps,
        success_method=success_method,
        detected_ns=detected_ns,
    )
